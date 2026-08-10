"""Tests for the PDF→Excel Protocol v2.1 capability.

Covers:
- dual-channel extraction + preflight on a real (hand-built) PDF
- channel comparison: disagreements become exception rows, agreements
  don't; a missing channel B text layer routes to single_channel
- workbook build: protocol sheets + real cross-sheet check formulas
- acceptance gates: pass on a clean build, fail on page loss (O-4) and on
  broken references / recalculation errors (O-2)
- run_transcription publishes v1 verdict "transcribed" and chains the
  automatic analysis; failures record last_error and release the claim
- transcribe route: guards, background dispatch, 503 on saturation
- auto_review_upload: setting gate, project creation + analysis
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient

from tests.conftest import AUTH_HEADERS


def _make_app(tmp: str):
    from orivellum.api import _deps
    from orivellum.api.app import app
    from orivellum.configuration.config import OrivellumConfig
    from orivellum.database.db import OrivellumDB

    cfg = OrivellumConfig(data_dir=tmp)
    db = OrivellumDB(str(Path(tmp) / "test.db"))
    _deps.init(db=db, cfg=cfg)
    return app, db, cfg


def _mini_pdf(page_texts: list[str]) -> bytes:
    """Hand-built valid PDF: one Helvetica text line per page."""
    n = len(page_texts)
    kids = " ".join(f"{3 + 2 * i} 0 R" for i in range(n))
    font_num = 3 + 2 * n
    objs: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{kids}] /Count {n} >>".encode(),
    ]
    for i, text in enumerate(page_texts):
        content_num = 4 + 2 * i
        objs.append(
            (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
                f"/Contents {content_num} 0 R "
                f"/Resources << /Font << /F1 {font_num} 0 R >> >> >>"
            ).encode()
        )
        stream = f"BT /F1 12 Tf 72 720 Td ({text}) Tj ET".encode()
        objs.append(b"<< /Length %d >>\nstream\n%s\nendstream" % (len(stream), stream))
    objs.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    out = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n".encode() + body + b"\nendobj\n"
    xref_pos = len(out)
    out += f"xref\n0 {len(objs) + 1}\n".encode()
    out += b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (
        f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF"
    ).encode()
    return bytes(out)


def _synthetic_pages():
    from orivellum.capabilities.pdf_excel import PageExtract

    return [
        PageExtract(
            page=1,
            width=612,
            height=792,
            text_a="Revenue 1234 Costs 567",
            text_b="Revenue 1234 Costs 567",
            tables=[[["Item", "Amount"], ["Revenue", "1234"], ["Costs", "567"]]],
        ),
        PageExtract(
            page=2,
            width=612,
            height=792,
            text_a="Notes on methodology for the quarter.",
            text_b="Notes on methodology for the quarter.",
        ),
    ]


def _manifest_for(pdf_path: Path, pages):
    from orivellum.capabilities.pdf_excel import preflight

    return preflight(pdf_path, pages)


class ExtractionAndPreflightTest(unittest.TestCase):
    def test_dual_channel_and_preflight(self):
        from orivellum.capabilities.pdf_excel import dual_channel_extract, preflight

        with tempfile.TemporaryDirectory() as tmp:
            pdf = Path(tmp) / "mini.pdf"
            pdf.write_bytes(
                _mini_pdf(["Revenue 1234 Costs 567", "Second page narrative text here"])
            )
            pages = dual_channel_extract(pdf)
            self.assertEqual(len(pages), 2)
            self.assertIn("1234", pages[0].text_a)
            self.assertIn("1234", pages[0].text_b)
            manifest = preflight(pdf, pages)
            self.assertEqual(manifest["page_count"], 2)
            self.assertEqual(manifest["source_type"], "born-digital")
            self.assertFalse(manifest["encrypted"])
            self.assertEqual(len(manifest["source_sha256"]), 64)
            self.assertIn(manifest["risk"]["class"], ("Low", "Moderate", "High", "Critical"))
            self.assertIn(manifest["risk"]["qa_tier"], (1, 2, 3, 4))


class ChannelComparisonTest(unittest.TestCase):
    def test_agreement_produces_no_exceptions(self):
        from orivellum.capabilities.pdf_excel import compare_channels

        self.assertEqual(compare_channels(_synthetic_pages()), [])

    def test_disagreement_is_bidirectional(self):
        from orivellum.capabilities.pdf_excel import compare_channels

        pages = _synthetic_pages()
        pages[0].text_b = "Revenue 1234 Costs 999"  # 567 uncorroborated, 999 unmatched
        exc = compare_channels(pages)
        self.assertEqual(len(exc), 2)
        a_only = next(e for e in exc if e.exception_type == "channel_disagreement")
        self.assertEqual(a_only.value_a, "567")
        self.assertEqual(a_only.pdf_page, 1)
        b_only = next(e for e in exc if e.exception_type == "channel_b_only")
        self.assertEqual(b_only.value_b, "999")

    def test_duplicate_occurrences_need_duplicate_corroboration(self):
        from orivellum.capabilities.pdf_excel import PageExtract, compare_channels

        pages = [
            PageExtract(page=1, text_a="Total 42 and again 42", text_b="Total 42 once")
        ]
        exc = compare_channels(pages)
        token_rows = [e for e in exc if e.value_a == "42"]
        self.assertEqual(len(token_rows), 1)
        self.assertEqual(token_rows[0].exception_type, "channel_disagreement")
        self.assertIn("2×", token_rows[0].description)

    def test_exception_flood_fails_closed(self):
        from orivellum.capabilities import pdf_excel
        from orivellum.capabilities.pdf_excel import PageExtract, compare_channels

        a = " ".join(str(1000 + i) for i in range(60))
        pages = [PageExtract(page=1, text_a=a, text_b="nothing numeric here at all")]
        with patch.object(pdf_excel, "MAX_EXCEPTIONS", 10), self.assertRaises(ValueError):
            compare_channels(pages)

    def test_missing_channel_b_is_single_channel(self):
        from orivellum.capabilities.pdf_excel import compare_channels

        pages = _synthetic_pages()
        pages[0].text_b = ""
        exc = [e for e in compare_channels(pages) if e.pdf_page == 1]
        self.assertEqual(len(exc), 1)
        self.assertEqual(exc[0].exception_type, "single_channel")

    def test_narrative_divergence_flagged(self):
        from orivellum.capabilities.pdf_excel import compare_channels

        pages = _synthetic_pages()
        pages[1].text_b = "Completely unrelated wording that shares nothing."
        exc = [e for e in compare_channels(pages) if e.pdf_page == 2]
        self.assertEqual(len(exc), 1)
        self.assertEqual(exc[0].exception_type, "channel_disagreement")


class WorkbookBuildAndGatesTest(unittest.TestCase):
    def _build(self, tmp: str, pages=None, exceptions=None):
        from orivellum.capabilities.pdf_excel import compare_channels
        from orivellum.capabilities.pdf_excel_build import build_workbook

        pdf = Path(tmp) / "mini.pdf"
        pdf.write_bytes(_mini_pdf(["Revenue 1234 Costs 567", "Second page"]))
        pages = pages or _synthetic_pages()
        exceptions = compare_channels(pages) if exceptions is None else exceptions
        manifest = _manifest_for(pdf, pages)
        xlsx = Path(tmp) / "out.xlsx"
        build_workbook(xlsx, "mini.pdf", manifest, pages, exceptions)
        return xlsx, manifest

    def test_workbook_has_protocol_sheets_and_formulas(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, _ = self._build(tmp)
            wb = load_workbook(xlsx)
            expected = ("README", "Page_Register", "P001", "Narrative",
                        "Exceptions", "Checks", "Changelog")
            for sheet in expected:
                self.assertIn(sheet, wb.sheetnames, sheet)
            checks = wb["Checks"]
            observed = [c.value for c in checks["C"] if isinstance(c.value, str)]
            self.assertTrue(any(v.startswith("=COUNTA(Page_Register!") for v in observed))
            statuses = [c.value for c in checks["E"] if isinstance(c.value, str)]
            self.assertTrue(any(v.startswith("=IF(") for v in statuses))

    def test_gates_pass_on_clean_build(self):
        from orivellum.capabilities.pdf_excel_build import run_acceptance_gates

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, manifest = self._build(tmp)
            gates = run_acceptance_gates(xlsx, manifest)
            self.assertTrue(gates["passed"], gates)

    def test_completeness_gate_catches_page_loss(self):
        from orivellum.capabilities.pdf_excel_build import run_acceptance_gates

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, manifest = self._build(tmp)
            manifest = dict(manifest, page_count=manifest["page_count"] + 1)
            gates = run_acceptance_gates(xlsx, manifest)
            self.assertFalse(gates["passed"])
            bad = next(g for g in gates["gates"] if g["name"] == "page_completeness")
            self.assertFalse(bad["ok"])

    def test_oversized_table_refuses_to_truncate(self):
        from orivellum.capabilities import pdf_excel_build
        from orivellum.capabilities.pdf_excel import PageExtract
        from orivellum.capabilities.pdf_excel_build import build_workbook

        with tempfile.TemporaryDirectory() as tmp:
            pdf = Path(tmp) / "mini.pdf"
            pdf.write_bytes(_mini_pdf(["x"]))
            big = PageExtract(page=1, text_a="t", text_b="t",
                              tables=[[["a"], ["b"], ["c"], ["d"]]])
            manifest = _manifest_for(pdf, [big])
            with (
                patch.object(pdf_excel_build, "_MAX_TABLE_ROWS", 3),
                self.assertRaises(ValueError),
            ):
                build_workbook(Path(tmp) / "o.xlsx", "mini.pdf", manifest, [big], [])

    def test_completeness_gate_catches_missing_data_row(self):
        from openpyxl import load_workbook

        from orivellum.capabilities.pdf_excel_build import run_acceptance_gates

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, manifest = self._build(tmp)
            wb = load_workbook(xlsx)
            ws = wb["P001"]
            # remove one transcribed record row — the gate must notice
            for row in ws.iter_rows():
                cell = row[0]
                if isinstance(cell.value, str) and cell.value.startswith("R001-"):
                    ws.delete_rows(cell.row, 1)
                    break
            wb.save(xlsx)
            gates = run_acceptance_gates(xlsx, manifest)
            self.assertFalse(gates["passed"])
            bad = next(g for g in gates["gates"] if g["name"] == "page_completeness")
            self.assertIn("row", bad["detail"])

    def test_completeness_gate_catches_truncated_narrative(self):
        from openpyxl import load_workbook

        from orivellum.capabilities.pdf_excel_build import run_acceptance_gates

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, manifest = self._build(tmp)
            wb = load_workbook(xlsx)
            ws = wb["Narrative"]
            ws["B2"] = str(ws["B2"].value)[:-5]  # cut the tail off the text
            wb.save(xlsx)
            gates = run_acceptance_gates(xlsx, manifest)
            self.assertFalse(gates["passed"])

    def test_page_cap_fails_closed(self):
        from orivellum.capabilities import pdf_excel
        from orivellum.capabilities.pdf_excel import dual_channel_extract

        with tempfile.TemporaryDirectory() as tmp:
            pdf = Path(tmp) / "two.pdf"
            pdf.write_bytes(_mini_pdf(["one", "two"]))
            with patch.object(pdf_excel, "MAX_PAGES", 1), self.assertRaises(ValueError):
                dual_channel_extract(pdf)

    def test_recalc_gate_catches_broken_reference(self):
        from openpyxl import load_workbook

        from orivellum.capabilities.pdf_excel_build import run_acceptance_gates

        with tempfile.TemporaryDirectory() as tmp:
            xlsx, manifest = self._build(tmp)
            wb = load_workbook(xlsx)
            wb["Checks"]["C9"] = "='Gone Sheet'!A1+1"
            wb.save(xlsx)
            gates = run_acceptance_gates(xlsx, manifest)
            self.assertFalse(gates["passed"])
            bad = next(g for g in gates["gates"] if g["name"] == "recalc")
            self.assertFalse(bad["ok"])


class RunTranscriptionTest(unittest.TestCase):
    def test_publishes_v1_and_chains_analysis(self):
        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            from orivellum.capabilities.pdf_excel import run_transcription

            pdf = Path(tmp) / "upload.pdf"
            pdf.write_bytes(_mini_pdf(["Revenue 1234 Costs 567"]))
            proj = db.create_wb_project("T", "xlsx", "b")
            self.assertTrue(db.claim_wb_build(proj["id"]))
            with patch("orivellum.capabilities.workbench_analyze.run_analysis") as ra:
                run_transcription(db, cfg, proj["id"], pdf, "upload.pdf")
            ra.assert_called_once()
            versions = db.list_wb_versions(proj["id"])
            self.assertEqual(len(versions), 1)
            self.assertEqual(versions[0]["verdict"], "transcribed")
            self.assertFalse(pdf.exists())

    def test_failure_records_error_and_releases_claim(self):
        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            from orivellum.capabilities.pdf_excel import run_transcription

            bad = Path(tmp) / "bad.pdf"
            bad.write_bytes(b"not a pdf at all")
            proj = db.create_wb_project("T", "xlsx", "b")
            self.assertTrue(db.claim_wb_build(proj["id"]))
            with patch("orivellum.capabilities.workbench_analyze.run_analysis") as ra:
                run_transcription(db, cfg, proj["id"], bad, "bad.pdf")
            ra.assert_not_called()
            row = db.get_wb_project(proj["id"])
            self.assertEqual(row["building"], 0)
            self.assertTrue(row["last_error"])
            self.assertFalse(bad.exists())


class TranscribeRouteTest(unittest.TestCase):
    def test_rejects_non_pdf(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, _, _ = _make_app(tmp)
            client = TestClient(app)
            r = client.post(
                "/api/workbench/transcribe",
                files={"file": ("data.xlsx", b"x", "application/octet-stream")},
                headers=AUTH_HEADERS,
            )
            self.assertEqual(r.status_code, 422)

    def test_dispatches_background_job(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, _ = _make_app(tmp)
            client = TestClient(app)
            with patch("orivellum.api.executor.submit_bg", return_value=True) as sb:
                r = client.post(
                    "/api/workbench/transcribe",
                    files={"file": ("report.pdf", _mini_pdf(["Hello 42"]), "application/pdf")},
                    headers=AUTH_HEADERS,
                )
            self.assertEqual(r.status_code, 200, r.text)
            sb.assert_called_once()
            proj = db.get_wb_project(r.json()["id"])
            self.assertEqual(proj["building"], 1)

    def test_saturation_returns_503_and_releases(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, _ = _make_app(tmp)
            client = TestClient(app)
            with patch("orivellum.api.executor.submit_bg", return_value=False):
                r = client.post(
                    "/api/workbench/transcribe",
                    files={"file": ("report.pdf", _mini_pdf(["Hello 42"]), "application/pdf")},
                    headers=AUTH_HEADERS,
                )
            self.assertEqual(r.status_code, 503)
            projects = db.list_wb_projects()
            self.assertEqual(projects[0]["building"], 0)


class AutoReviewTest(unittest.TestCase):
    def _xlsx(self, tmp: str) -> Path:
        from openpyxl import Workbook

        wb = Workbook()
        wb.active["A1"] = 1
        p = Path(tmp) / "book.xlsx"
        wb.save(p)
        return p

    def test_setting_off_skips(self):
        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            from orivellum.capabilities.workbench import auto_review_upload

            db.set_setting("workbench_auto_review", "false")
            with patch("orivellum.capabilities.workbench_analyze.run_analysis") as ra:
                auto_review_upload(db, cfg, self._xlsx(tmp), "book.xlsx")
            ra.assert_not_called()
            self.assertEqual(db.list_wb_projects(), [])

    def test_default_on_imports_and_analyzes(self):
        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            from orivellum.capabilities.workbench import auto_review_upload

            with patch("orivellum.capabilities.workbench_analyze.run_analysis") as ra:
                auto_review_upload(db, cfg, self._xlsx(tmp), "book.xlsx")
            ra.assert_called_once()
            projects = db.list_wb_projects()
            self.assertEqual(len(projects), 1)
            self.assertTrue(projects[0]["title"].startswith("Review:"))

    def test_import_route_reports_auto_review(self):
        import io

        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            app, _, _ = _make_app(tmp)
            client = TestClient(app)
            wb = Workbook()
            wb.active["A1"] = 1
            buf = io.BytesIO()
            wb.save(buf)
            with patch("orivellum.api.executor.submit_bg", return_value=True):
                r = client.post(
                    "/api/workbench/projects/import",
                    files={"file": ("b.xlsx", buf.getvalue(), "application/octet-stream")},
                    headers=AUTH_HEADERS,
                )
            self.assertEqual(r.status_code, 200, r.text)
            self.assertTrue(r.json().get("auto_review_started"))


if __name__ == "__main__":
    unittest.main()
