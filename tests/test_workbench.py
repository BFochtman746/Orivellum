"""Tests for the Project Workbench (build / edit / repair / version / archive).

Covers:
- DB layer: project CRUD, atomic version numbering, update whitelist
- Output verification: xlsx load checks, code syntax/JSON checks, empty output
- run_build happy path with a stubbed LLM (real sandbox execution)
- run_build failure path: error lands on the project row, no version created
- revert_to copies an old version forward as a NEW version
- archive_project: zip + manifest with hashes, project flips to archived
- Routes: create/list/detail, iterate guards, revert, complete, download,
  delete, archived projects are read-only
"""

from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient

from tests.conftest import AUTH_HEADERS


def _make_app(tmp: str):
    from orivellum.api import _deps
    from orivellum.api.app import app
    from orivellum.configuration.config import OrivellumConfig
    from orivellum.database.db import OrivellumDB

    cfg = OrivellumConfig(data_dir=tmp)
    db = OrivellumDB(str(Path(tmp) / "test.db"))
    _deps.init(db=db, cfg=cfg)
    return app, db, cfg


def _make_db(tmp: str):
    from orivellum.database.db import OrivellumDB

    return OrivellumDB(str(Path(tmp) / "test.db"))


# ── DB layer ──────────────────────────────────────────────────────────────────


class TestWorkbenchDB(unittest.TestCase):
    def test_project_crud_and_version_numbering(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = _make_db(tmp)
            p = db.create_wb_project("Budget", "xlsx", "a budget")
            self.assertEqual(p["status"], "active")
            self.assertEqual(p["kind"], "xlsx")

            v1 = db.create_wb_version(
                p["id"], "first", [{"name": "a.xlsx", "size": 1, "sha256": "x"}]
            )
            v2 = db.create_wb_version(p["id"], "second", [])
            self.assertEqual((v1["version_no"], v2["version_no"]), (1, 2))
            self.assertEqual([v["version_no"] for v in db.list_wb_versions(p["id"])], [1, 2])
            self.assertIsNotNone(db.get_wb_version(p["id"], 2))
            self.assertIsNone(db.get_wb_version(p["id"], 99))

            db.update_wb_project(p["id"], status="archived", nonsense="ignored")
            self.assertEqual(db.get_wb_project(p["id"])["status"], "archived")

            db.delete_wb_project(p["id"])
            self.assertIsNone(db.get_wb_project(p["id"]))
            self.assertEqual(db.list_wb_versions(p["id"]), [])

    def test_claim_is_atomic_and_exclusive(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = _make_db(tmp)
            p = db.create_wb_project("Tool", "code", "t")
            self.assertTrue(db.claim_wb_build(p["id"]))
            self.assertFalse(db.claim_wb_build(p["id"]))  # already claimed
            db.update_wb_project(p["id"], building=0)
            self.assertTrue(db.claim_wb_build(p["id"]))  # released → claimable
            db.update_wb_project(p["id"], building=0, status="archived")
            self.assertFalse(db.claim_wb_build(p["id"]))  # archived + active-only
            self.assertTrue(db.claim_wb_build(p["id"], require_active=False))
            self.assertFalse(db.claim_wb_build("no-such-id"))


# ── Verification ──────────────────────────────────────────────────────────────


class TestVerifyOutput(unittest.TestCase):
    def test_empty_output_fails(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            ok, checks = _verify_output("xlsx", Path(tmp))
            self.assertFalse(ok)
            self.assertIn("no files", checks["error"])

    def test_xlsx_good_and_corrupt(self):
        from openpyxl import Workbook

        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            wb = Workbook()
            wb.active["A1"] = "hello"
            wb.active["A2"] = "=SUM(1,2)"
            wb.save(out / "good.xlsx")
            ok, checks = _verify_output("xlsx", out)
            self.assertTrue(ok, checks)

            (out / "bad.xlsx").write_bytes(b"this is not a workbook")
            ok, checks = _verify_output("xlsx", out)
            self.assertFalse(ok)
            self.assertTrue(any("bad.xlsx" in p for p in checks["problems"]))

    def test_xlsx_requires_a_workbook(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / "readme.txt").write_text("no workbook here")
            ok, checks = _verify_output("xlsx", Path(tmp))
            self.assertFalse(ok)
            self.assertIn("no .xlsx file produced", checks["problems"])

    def test_code_syntax_and_json_checks(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "main.py").write_text("def ok():\n    return 1\n")
            (out / "cfg.json").write_text('{"a": 1}')
            ok, checks = _verify_output("code", out)
            self.assertTrue(ok, checks)

            (out / "broken.py").write_text("def broken(:\n")
            (out / "broken.json").write_text("{nope")
            ok, checks = _verify_output("code", out)
            self.assertFalse(ok)
            self.assertEqual(len(checks["problems"]), 2)


# ── Build loop (stubbed LLM, real sandbox) ────────────────────────────────────

_GOOD_XLSX_SCRIPT = """
import shutil, pathlib
from openpyxl import Workbook
out = pathlib.Path("out")
wb = Workbook()
ws = wb.active
ws.title = "Budget"
ws["A1"] = "Item"; ws["B1"] = "Cost"
ws["A2"] = "Rent"; ws["B2"] = 1200
ws["B4"] = "=SUM(B2:B3)"
wb.save(out / "budget.xlsx")
print("built budget workbook")
"""


class TestRunBuild(unittest.TestCase):
    def test_happy_path_creates_verified_version(self):
        from orivellum.capabilities.llm import LLMResult
        from orivellum.capabilities.workbench import run_build, version_dir

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "a budget workbook")
            with patch(
                "orivellum.capabilities.llm.llm_call",
                return_value=LLMResult(_GOOD_XLSX_SCRIPT, True, "test", 0),
            ):
                run_build(db, cfg, p["id"], "build it")

            proj = db.get_wb_project(p["id"])
            self.assertEqual(proj["building"], 0)
            self.assertIsNone(proj["last_error"], proj["last_error"])
            versions = db.list_wb_versions(p["id"])
            self.assertEqual(len(versions), 1)
            v = versions[0]
            # xlsx builds now run the six-gate proof harness and promote
            self.assertEqual(v["verdict"], "proven")
            checks = json.loads(v["checks_json"])
            self.assertEqual(checks["proof"]["verdict"], "proven")
            files = json.loads(v["files_json"])
            self.assertEqual(files[0]["name"], "budget.xlsx")
            self.assertEqual(len(files[0]["sha256"]), 64)
            self.assertTrue((version_dir(cfg, p["id"], 1) / "budget.xlsx").is_file())

    def test_llm_failure_sets_error_and_no_version(self):
        from orivellum.capabilities.llm import LLMResult
        from orivellum.capabilities.workbench import run_build

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "a budget workbook")
            with patch(
                "orivellum.capabilities.llm.llm_call",
                return_value=LLMResult(None, False, "test", 0, error="down"),
            ):
                run_build(db, cfg, p["id"], "build it")

            proj = db.get_wb_project(p["id"])
            self.assertEqual(proj["building"], 0)
            self.assertIn("down", proj["last_error"])
            self.assertEqual(db.list_wb_versions(p["id"]), [])

    def test_archived_project_never_builds(self):
        from orivellum.capabilities.workbench import run_build

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            db.update_wb_project(p["id"], status="archived")
            run_build(db, cfg, p["id"], "should be a no-op")  # must not raise
            self.assertEqual(db.list_wb_versions(p["id"]), [])


# ── Revert + archive ──────────────────────────────────────────────────────────


def _write_version(db, cfg, project_id: str, content: str):
    """Create a version row + on-disk files directly (bypasses the LLM)."""
    import shutil

    from orivellum.capabilities.workbench import _snapshot, version_dir

    with tempfile.TemporaryDirectory() as t:
        f = Path(t) / "main.py"
        f.write_text(content)
        files = _snapshot(Path(t))
        row = db.create_wb_version(project_id, f"write {content!r}", files)
        dest = version_dir(cfg, project_id, row["version_no"])
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(t, dest, dirs_exist_ok=True)
    return row


class TestRevertAndArchive(unittest.TestCase):
    def test_revert_copies_forward_as_new_version(self):
        from orivellum.capabilities.workbench import revert_to, version_dir

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Tool", "code", "a tool")
            _write_version(db, cfg, p["id"], "x = 1\n")
            _write_version(db, cfg, p["id"], "x = 2\n")

            row = revert_to(db, cfg, p["id"], 1)
            self.assertEqual(row["version_no"], 3)
            text = (version_dir(cfg, p["id"], 3) / "main.py").read_text()
            self.assertEqual(text, "x = 1\n")
            # history untouched
            self.assertEqual(len(db.list_wb_versions(p["id"])), 3)

    def test_archive_builds_zip_with_manifest_and_locks_project(self):
        from orivellum.capabilities.workbench import archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("My Tool!", "code", "a tool")
            _write_version(db, cfg, p["id"], "x = 1\n")
            _write_version(db, cfg, p["id"], "x = 2\n")

            path = archive_project(db, cfg, p["id"])
            self.assertTrue(Path(path).is_file())
            with zipfile.ZipFile(path) as z:
                names = set(z.namelist())
                self.assertIn("manifest.json", names)
                self.assertIn("v1/main.py", names)
                self.assertIn("v2/main.py", names)
                manifest = json.loads(z.read("manifest.json"))
            self.assertEqual(len(manifest["versions"]), 2)
            self.assertEqual(len(manifest["versions"][0]["files"][0]["sha256"]), 64)

            proj = db.get_wb_project(p["id"])
            self.assertEqual(proj["status"], "archived")
            self.assertEqual(proj["archive_path"], path)

    def test_archive_refuses_on_tampered_or_missing_files(self):
        from orivellum.capabilities.workbench import archive_project, version_dir

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Tool", "code", "a tool")
            _write_version(db, cfg, p["id"], "x = 1\n")
            target = version_dir(cfg, p["id"], 1) / "main.py"

            target.write_text("x = 999  # tampered\n")
            with self.assertRaises(RuntimeError):
                archive_project(db, cfg, p["id"])

            target.unlink()
            with self.assertRaises(RuntimeError):
                archive_project(db, cfg, p["id"])
            # project must NOT have been archived
            self.assertEqual(db.get_wb_project(p["id"])["status"], "active")

    def test_archive_empty_project_rejected(self):
        from orivellum.capabilities.workbench import archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Empty", "code", "nothing")
            with self.assertRaises(ValueError):
                archive_project(db, cfg, p["id"])


# ── Routes ────────────────────────────────────────────────────────────────────


class TestWorkbenchRoutes(unittest.TestCase):
    def test_full_route_flow(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, cfg = _make_app(tmp)
            client = TestClient(app)

            with patch("orivellum.api.routes.workbench._start_build") as start:
                # create
                r = client.post(
                    "/api/workbench/projects",
                    headers=AUTH_HEADERS,
                    json={"title": "Budget", "kind": "xlsx", "brief": "a budget"},
                )
                self.assertEqual(r.status_code, 200, r.text)
                pid = r.json()["id"]
                self.assertTrue(r.json()["building"])
                start.assert_called_once()

                # invalid kind
                r = client.post(
                    "/api/workbench/projects",
                    headers=AUTH_HEADERS,
                    json={"title": "X", "kind": "pptx", "brief": "b"},
                )
                self.assertEqual(r.status_code, 422)

                # iterate while building → 409
                r = client.post(
                    f"/api/workbench/projects/{pid}/iterate",
                    headers=AUTH_HEADERS,
                    json={"instruction": "more"},
                )
                self.assertEqual(r.status_code, 409)

                # finish the "build", then iterate works
                db.update_wb_project(pid, building=0)
                r = client.post(
                    f"/api/workbench/projects/{pid}/iterate",
                    headers=AUTH_HEADERS,
                    json={"instruction": "more"},
                )
                self.assertEqual(r.status_code, 200, r.text)
                db.update_wb_project(pid, building=0)

            # complete with zero versions → 422
            r = client.post(f"/api/workbench/projects/{pid}/complete", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 422)

            # add a real version, then download + complete
            _write_version(db, cfg, pid, "x = 1\n")
            r = client.get(
                f"/api/workbench/projects/{pid}/versions/1/download", headers=AUTH_HEADERS
            )
            self.assertEqual(r.status_code, 200)
            self.assertEqual(r.headers["content-type"], "application/zip")

            # xlsx project whose latest version was never proven → refused
            r = client.post(f"/api/workbench/projects/{pid}/complete", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 409, r.text)
            self.assertEqual(r.json()["detail"]["code"], "unproven")

            r = client.post(
                f"/api/workbench/projects/{pid}/complete",
                headers=AUTH_HEADERS,
                json={"force": True},
            )
            self.assertEqual(r.status_code, 200, r.text)

            # archived project is read-only
            r = client.post(
                f"/api/workbench/projects/{pid}/iterate",
                headers=AUTH_HEADERS,
                json={"instruction": "more"},
            )
            self.assertEqual(r.status_code, 409)

            # archive download works
            r = client.get(f"/api/workbench/projects/{pid}/archive/download", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 200)

            # detail lists versions
            r = client.get(f"/api/workbench/projects/{pid}", headers=AUTH_HEADERS)
            self.assertEqual(
                r.json()["version_count"], 1
            )  # builds were mocked; only the direct write exists
            # delete
            r = client.delete(f"/api/workbench/projects/{pid}", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 200)
            r = client.get(f"/api/workbench/projects/{pid}", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 404)

    def test_requires_auth(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, _, _ = _make_app(tmp)
            client = TestClient(app)
            r = client.get("/api/workbench/projects")
            self.assertIn(r.status_code, (401, 403))


if __name__ == "__main__":
    unittest.main()
