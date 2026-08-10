"""Tests for Workbench import (upload → v1) and analyze mode.

Covers:
- kind auto-detection from staged file names
- import via the multipart route: single .xlsx and .zip uploads
- zip safety: path traversal rejected, file-count limit enforced
- wrong file types and empty uploads rejected
- analyze_workbook deterministic findings (broken sheet refs, volatile
  functions, hidden sheets, README detection)
- analyze_code_tree deterministic findings (syntax errors, invalid JSON,
  TODO markers, test detection)
- run_analysis publishes ANALYSIS_REPORT.md as a new 'analyzed' version
  and releases the build claim
- analyze/report routes: guards and report retrieval
"""

from __future__ import annotations

import io
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient

from tests.conftest import AUTH_HEADERS


def _make_app(tmp: str):
    from orivellum.api import _deps
    from orivellum.api.app import app
    from orivellum.configuration.config import OrivellumConfig
    from orivellum.database.db import OrivellumDB

    cfg = OrivellumConfig(data_dir=tmp)
    db = OrivellumDB(str(Path(tmp) / "test.db"))
    _deps.init(db=db, cfg=cfg)
    return app, db, cfg


def _xlsx_bytes(broken_ref: bool = False, volatile: bool = False) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=A1+A2"
    if broken_ref:
        ws["B1"] = "='Gone Sheet'!A1"
        ws["B2"] = "=Missing!C3"
    if volatile:
        ws["C1"] = "=NOW()"
    hidden = wb.create_sheet("Secrets")
    hidden.sheet_state = "hidden"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zip_bytes(entries: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for name, data in entries.items():
            zf.writestr(name, data)
    return buf.getvalue()


# ── kind detection & extraction safety ────────────────────────────────────────


class TestKindAndExtraction(unittest.TestCase):
    def test_detect_kind(self):
        from orivellum.capabilities.workbench import _detect_kind

        self.assertEqual(_detect_kind(["Book1.xlsx", "notes.md"]), "xlsx")
        self.assertEqual(_detect_kind(["Book1.xlsx", "tool.py"]), "code")
        self.assertEqual(_detect_kind(["main.py", "data.json"]), "code")

    def test_zip_slip_rejected(self):
        from orivellum.capabilities.workbench import _extract_upload

        with tempfile.TemporaryDirectory() as tmp:
            up = Path(tmp) / "evil.zip"
            up.write_bytes(_zip_bytes({"../evil.txt": b"x", "ok.py": b"pass"}))
            stage = Path(tmp) / "stage"
            stage.mkdir()
            with self.assertRaises(ValueError):
                _extract_upload(up, "evil.zip", stage)

    def test_too_many_files_rejected(self):
        from orivellum.capabilities.workbench import _MAX_OUTPUT_FILES, _extract_upload

        with tempfile.TemporaryDirectory() as tmp:
            up = Path(tmp) / "many.zip"
            up.write_bytes(_zip_bytes({f"f{i}.txt": b"x" for i in range(_MAX_OUTPUT_FILES + 1)}))
            stage = Path(tmp) / "stage"
            stage.mkdir()
            with self.assertRaises(ValueError):
                _extract_upload(up, "many.zip", stage)

    def test_junk_entries_skipped_and_dirs_preserved(self):
        from orivellum.capabilities.workbench import _extract_upload

        with tempfile.TemporaryDirectory() as tmp:
            up = Path(tmp) / "proj.zip"
            up.write_bytes(
                _zip_bytes(
                    {
                        "src/main.py": b"print('hi')\n",
                        "src/util/helpers.py": b"pass\n",
                        "__MACOSX/._main.py": b"junk",
                        ".DS_Store": b"junk",
                    }
                )
            )
            stage = Path(tmp) / "stage"
            stage.mkdir()
            names = sorted(_extract_upload(up, "proj.zip", stage))
            self.assertEqual(names, ["src/main.py", "src/util/helpers.py"])
            self.assertTrue((stage / "src" / "util" / "helpers.py").is_file())

    def test_xlsx_zip_bomb_rejected(self):
        from orivellum.capabilities import workbench as wbmod

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "b.xlsx"
            p.write_bytes(_xlsx_bytes())
            self.assertIsNone(wbmod.check_xlsx_zip_safety(p))
            with patch.object(wbmod, "_XLSX_MAX_UNCOMPRESSED", 10):
                self.assertIn("uncompressed", wbmod.check_xlsx_zip_safety(p))
            with patch.object(wbmod, "_XLSX_MAX_MEMBERS", 1):
                self.assertIn("members", wbmod.check_xlsx_zip_safety(p))
            # analyze treats a bomb as a load_error finding, never opens it
            from orivellum.capabilities.workbench_analyze import analyze_workbook

            with patch.object(wbmod, "_XLSX_MAX_UNCOMPRESSED", 10):
                f = analyze_workbook(p)
            self.assertIn("uncompressed", f["load_error"])


# ── import route ──────────────────────────────────────────────────────────────


class TestImportRoute(unittest.TestCase):
    def test_import_xlsx_becomes_v1(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, _cfg = _make_app(tmp)
            client = TestClient(app)
            # auto-review dispatch is dropped so v1 import state is observable
            with patch("orivellum.api.executor.submit_bg", return_value=False):
                r = client.post(
                    "/api/workbench/projects/import",
                    files={"file": ("Budget 2026.xlsx", _xlsx_bytes(), "application/octet-stream")},
                    headers=AUTH_HEADERS,
                )
            self.assertEqual(r.status_code, 200, r.text)
            body = r.json()
            self.assertEqual(body["kind"], "xlsx")
            self.assertEqual(body["title"], "Budget 2026")
            self.assertFalse(body["building"])
            self.assertEqual(body["version_count"], 1)
            v1 = body["versions"][0]
            self.assertEqual(v1["verdict"], "imported")
            self.assertEqual([f["name"] for f in v1["files"]], ["Budget 2026.xlsx"])
            self.assertTrue(v1["checks"]["imported"])
            # files really are on disk as v1
            from orivellum.capabilities.workbench import version_dir

            self.assertTrue((version_dir(_cfg, body["id"], 1) / "Budget 2026.xlsx").is_file())

    def test_import_zip_detects_code_kind(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, _db, _cfg = _make_app(tmp)
            client = TestClient(app)
            payload = _zip_bytes({"main.py": b"print(1)\n", "cfg.json": b"{}"})
            r = client.post(
                "/api/workbench/projects/import",
                files={"file": ("tool.zip", payload, "application/zip")},
                data={"brief": "a small tool"},
                headers=AUTH_HEADERS,
            )
            self.assertEqual(r.status_code, 200, r.text)
            body = r.json()
            self.assertEqual(body["kind"], "code")
            self.assertEqual(body["brief"], "a small tool")
            names = {f["name"] for f in body["versions"][0]["files"]}
            self.assertEqual(names, {"main.py", "cfg.json"})

    def test_import_rejects_bad_uploads(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, _db, _cfg = _make_app(tmp)
            client = TestClient(app)
            r = client.post(
                "/api/workbench/projects/import",
                files={"file": ("notes.txt", b"hello", "text/plain")},
                headers=AUTH_HEADERS,
            )
            self.assertEqual(r.status_code, 422)
            r = client.post(
                "/api/workbench/projects/import",
                files={"file": ("empty.xlsx", b"", "application/octet-stream")},
                headers=AUTH_HEADERS,
            )
            self.assertEqual(r.status_code, 400)
            r = client.post(
                "/api/workbench/projects/import",
                files={"file": ("bad.zip", b"not a zip", "application/zip")},
                headers=AUTH_HEADERS,
            )
            self.assertEqual(r.status_code, 422)
            # no half-created projects left behind
            self.assertEqual(
                client.get("/api/workbench/projects", headers=AUTH_HEADERS).json()["projects"], []
            )

    def test_import_publish_failure_leaves_nothing_behind(self):
        from orivellum.capabilities import workbench as wbmod

        with tempfile.TemporaryDirectory() as tmp:
            _app, db, cfg = _make_app(tmp)
            up = Path(tmp) / "b.xlsx"
            up.write_bytes(_xlsx_bytes())
            with (
                patch.object(wbmod, "_publish_version", side_effect=RuntimeError("disk full")),
                self.assertRaises(RuntimeError),
            ):
                wbmod.import_upload(db, cfg, "Doomed", "x", up, "b.xlsx")
            self.assertEqual(db.list_wb_projects(), [])
            # no project directories left under the workbench data dir
            root = wbmod.project_dir(cfg, "x").parent
            leftovers = [p for p in root.iterdir() if p.is_dir()] if root.is_dir() else []
            self.assertEqual(leftovers, [])


# ── deterministic analyzers ───────────────────────────────────────────────────


class TestAnalyzers(unittest.TestCase):
    def test_analyze_workbook_findings(self):
        from orivellum.capabilities.workbench_analyze import analyze_workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "b.xlsx"
            p.write_bytes(_xlsx_bytes(broken_ref=True, volatile=True))
            f = analyze_workbook(p)
            self.assertEqual(f["broken_reference_count"], 2)
            self.assertTrue(any("Gone Sheet" in b for b in f["broken_references"]))
            self.assertEqual(f["volatile_count"], 1)
            self.assertEqual(f["hidden_sheets"], ["Secrets"])
            self.assertFalse(f["has_readme_sheet"])
            self.assertGreaterEqual(f["formulas"], 3)

    def test_analyze_workbook_unreadable_is_a_finding(self):
        from orivellum.capabilities.workbench_analyze import analyze_workbook, summarize_issues

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "broken.xlsx"
            p.write_bytes(b"this is not a workbook")
            f = analyze_workbook(p)
            self.assertIn("load_error", f)
            issues = summarize_issues("xlsx", {"workbooks": [f]})
            self.assertTrue(any("cannot be opened" in i for i in issues))

    def test_analyze_code_tree_findings(self):
        from orivellum.capabilities.workbench_analyze import analyze_code_tree, summarize_issues

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "good.py").write_text("x = 1  # TODO tidy\n")
            (root / "bad.py").write_text("def broken(:\n")
            (root / "bad.json").write_text("{nope")
            f = analyze_code_tree(root)
            self.assertEqual(len(f["python_syntax_errors"]), 1)
            self.assertEqual(len(f["invalid_json"]), 1)
            self.assertEqual(f["todo_marker_count"], 1)
            self.assertFalse(f["has_tests"])
            issues = summarize_issues("code", f)
            self.assertTrue(any("do not parse" in i for i in issues))
            self.assertTrue(any("no tests" in i for i in issues))


# ── run_analysis ──────────────────────────────────────────────────────────────


class TestRunAnalysis(unittest.TestCase):
    def _imported_project(self, db, cfg, xlsx: bytes):
        from orivellum.capabilities.workbench import import_upload

        with tempfile.TemporaryDirectory() as t2:
            up = Path(t2) / "b.xlsx"
            up.write_bytes(xlsx)
            return import_upload(db, cfg, "Book", "a book", up, "b.xlsx")

    def test_analysis_publishes_report_version(self):
        from orivellum.capabilities import workbench_analyze as wa
        from orivellum.capabilities.workbench import version_dir

        with tempfile.TemporaryDirectory() as tmp:
            _app, db, cfg = _make_app(tmp)
            proj = self._imported_project(db, cfg, _xlsx_bytes(broken_ref=True))
            self.assertTrue(db.claim_wb_build(proj["id"]))
            with patch.object(wa, "_narrative", return_value="stub narrative"):
                wa.run_analysis(db, cfg, proj["id"], focus="check the refs")
            proj2 = db.get_wb_project(proj["id"])
            self.assertEqual(proj2["building"], 0)
            self.assertIsNone(proj2["last_error"])
            versions = db.list_wb_versions(proj["id"])
            self.assertEqual(len(versions), 2)
            v2 = versions[-1]
            self.assertEqual(v2["verdict"], "analyzed")
            report = version_dir(cfg, proj["id"], 2) / "ANALYSIS_REPORT.md"
            self.assertTrue(report.is_file())
            text = report.read_text(encoding="utf-8")
            self.assertIn("missing sheet", text)
            self.assertIn("stub narrative", text)
            # original workbook travels with the report version
            self.assertTrue((version_dir(cfg, proj["id"], 2) / "b.xlsx").is_file())
            import json as _json

            checks = _json.loads(v2["checks_json"])
            self.assertGreaterEqual(checks["analysis"]["issue_count"], 1)

    def test_analysis_failure_lands_on_project_row(self):
        from orivellum.capabilities import workbench_analyze as wa

        with tempfile.TemporaryDirectory() as tmp:
            _app, db, cfg = _make_app(tmp)
            proj = db.create_wb_project("Empty", "xlsx", "no versions")
            self.assertTrue(db.claim_wb_build(proj["id"]))
            wa.run_analysis(db, cfg, proj["id"])
            proj2 = db.get_wb_project(proj["id"])
            self.assertEqual(proj2["building"], 0)
            self.assertIn("nothing to analyze", proj2["last_error"])
            self.assertEqual(db.list_wb_versions(proj["id"]), [])


# ── analyze & report routes ───────────────────────────────────────────────────


class TestAnalyzeRoutes(unittest.TestCase):
    def test_route_guards_and_report_fetch(self):
        from orivellum.capabilities import workbench_analyze as wa

        with tempfile.TemporaryDirectory() as tmp:
            app, db, _cfg = _make_app(tmp)
            client = TestClient(app)
            # import a workbook first (drop the auto-review dispatch)
            with patch("orivellum.api.executor.submit_bg", return_value=False):
                r = client.post(
                    "/api/workbench/projects/import",
                    files={"file": ("b.xlsx", _xlsx_bytes(), "application/octet-stream")},
                    headers=AUTH_HEADERS,
                )
            pid = r.json()["id"]

            # report on a version without one → 404
            r = client.get(f"/api/workbench/projects/{pid}/versions/1/report", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 404)

            # analyze runs synchronously via a stubbed executor
            def _sync_submit(fn, **_kw):
                fn()
                return True

            with (
                patch("orivellum.api.executor.submit_bg", _sync_submit),
                patch.object(wa, "_narrative", return_value="stub narrative"),
            ):
                r = client.post(
                    f"/api/workbench/projects/{pid}/analyze",
                    json={"focus": ""},
                    headers=AUTH_HEADERS,
                )
            self.assertEqual(r.status_code, 200, r.text)

            r = client.get(f"/api/workbench/projects/{pid}/versions/2/report", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 200)
            self.assertIn("Analysis Report", r.json()["report"])

            # analyzing while a build claim is held → 409
            self.assertTrue(db.claim_wb_build(pid))
            r = client.post(f"/api/workbench/projects/{pid}/analyze", json={}, headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 409)
            db.update_wb_project(pid, building=0)

            # a project with no versions → 409
            p2 = db.create_wb_project("Empty", "code", "x")
            r = client.post(
                f"/api/workbench/projects/{p2['id']}/analyze", json={}, headers=AUTH_HEADERS
            )
            self.assertEqual(r.status_code, 409)

    def test_dropped_dispatch_releases_claim(self):
        """If the executor drops the work, the project must not stay
        stranded as 'building'."""
        with tempfile.TemporaryDirectory() as tmp:
            app, db, _cfg = _make_app(tmp)
            client = TestClient(app)
            with patch("orivellum.api.executor.submit_bg", return_value=False):
                r = client.post(
                    "/api/workbench/projects/import",
                    files={"file": ("b.xlsx", _xlsx_bytes(), "application/octet-stream")},
                    headers=AUTH_HEADERS,
                )
            pid = r.json()["id"]
            with patch("orivellum.api.executor.submit_bg", return_value=False):
                r = client.post(
                    f"/api/workbench/projects/{pid}/analyze", json={}, headers=AUTH_HEADERS
                )
            self.assertEqual(r.status_code, 503)
            proj = db.get_wb_project(pid)
            self.assertEqual(proj["building"], 0)
            self.assertIn("saturated", proj["last_error"])
            # and the project is usable again afterwards
            self.assertTrue(db.claim_wb_build(pid))
            db.update_wb_project(pid, building=0)


if __name__ == "__main__":
    unittest.main()
