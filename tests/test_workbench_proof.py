"""Tests for the six-gate workbook proving bridge (workbench_proof).

Covers:
- prove_workbook: an LLM-style workbook (formulas, no cached values) is
  repaired, gated, and atomically promoted — data_only reads show real numbers
- prove_workbook failure: a genuinely erroring formula fails the gates and the
  original file stays byte-for-byte untouched
- promote=False (imports): gates run, verdict recorded, file stays verbatim
- unavailable harness degrades to 'unverified', never 'proven'
- _verify_output(prove=True): proof result lands in checks; failed proof
  rejects the version
- archive gate: an unproven / failed latest version refuses to archive unless
  allow_unproven=True; the complete route maps that to 409 code=unproven
"""

from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from tests.conftest import AUTH_HEADERS


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _good_workbook(path: Path) -> None:
    """Formulas with NO cached values — how openpyxl-built files come out."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 10
    ws["A2"] = 32
    ws["A3"] = "=SUM(A1:A2)"
    ws["B1"] = "=A1*2"
    wb.save(path)


def _error_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 0
    ws["A2"] = "=1/A1"  # genuine #DIV/0! — never cache-repairable
    wb.save(path)


def _make_app(tmp: str):
    from orivellum.api import _deps
    from orivellum.api.app import app
    from orivellum.configuration.config import OrivellumConfig
    from orivellum.database.db import OrivellumDB

    cfg = OrivellumConfig(data_dir=tmp)
    db = OrivellumDB(str(Path(tmp) / "test.db"))
    _deps.init(db=db, cfg=cfg)
    return app, db, cfg


class TestProveWorkbook(unittest.TestCase):
    def test_good_workbook_is_proven_and_promoted(self):
        from orivellum.capabilities.workbench_proof import GATE_NAMES, prove_workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "book.xlsx"
            _good_workbook(p)
            before = _sha(p)

            res = prove_workbook(p)
            self.assertEqual(res["verdict"], "proven", res)
            for g in GATE_NAMES:
                self.assertTrue(res["gates"][g], g)
            # promotion happened: repaired file differs, caches are real values
            self.assertNotEqual(_sha(p), before)
            wb = load_workbook(p, data_only=True)
            self.assertEqual(wb["Data"]["A3"].value, 42)
            self.assertEqual(wb["Data"]["B1"].value, 20)
            wb.close()
            self.assertGreaterEqual(res["repairs"]["refreshed_cells"], 2)
            # no candidate debris
            self.assertEqual(list(Path(tmp).glob(".candidate_*")), [])

    def test_error_formula_fails_proof_and_original_untouched(self):
        from orivellum.capabilities.workbench_proof import prove_workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "broken.xlsx"
            _error_workbook(p)
            before = _sha(p)

            res = prove_workbook(p)
            self.assertEqual(res["verdict"], "failed")
            self.assertIn("G2_values_match", res["failed_gates"])
            self.assertTrue(res["problems"])
            self.assertEqual(_sha(p), before)  # never mutated on failure
            self.assertEqual(list(Path(tmp).glob(".candidate_*")), [])

    def test_promote_false_keeps_file_verbatim_and_never_lies(self):
        from orivellum.capabilities.workbench_proof import prove_workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "book.xlsx"
            _good_workbook(p)
            before = _sha(p)
            # gates pass only after cache repairs the file never received →
            # 'provable', NOT 'proven' (the verbatim bytes were never gated)
            res = prove_workbook(p, promote=False)
            self.assertEqual(res["verdict"], "provable")
            self.assertEqual(_sha(p), before)
            self.assertTrue(any("verbatim" in msg for msg in res["problems"]))

            # once actually promoted, a promote=False re-run needs no repairs
            # → the file itself is proven
            self.assertEqual(prove_workbook(p)["verdict"], "proven")
            after_promotion = _sha(p)
            res2 = prove_workbook(p, promote=False)
            self.assertEqual(res2["verdict"], "proven")
            self.assertEqual(_sha(p), after_promotion)

    def test_missing_harness_is_unverified_never_proven(self):
        from orivellum.capabilities import workbench_proof

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "book.xlsx"
            _good_workbook(p)
            with patch.object(workbench_proof, "_load_runner_modules", return_value=(None, None)):
                res = workbench_proof.prove_workbook(p)
            self.assertEqual(res["verdict"], "unverified")
            self.assertIn("unavailable", res["error"])


class TestVerifyOutputProof(unittest.TestCase):
    def test_verify_records_proof_and_accepts(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _good_workbook(out / "model.xlsx")
            ok, checks = _verify_output("xlsx", out, prove=True)
            self.assertTrue(ok, checks)
            self.assertEqual(checks["proof"]["verdict"], "proven")
            self.assertIn("model.xlsx", checks["proof"]["workbooks"])

    def test_verify_rejects_failed_proof(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _error_workbook(out / "broken.xlsx")
            ok, checks = _verify_output("xlsx", out, prove=True)
            self.assertFalse(ok)
            self.assertTrue(any("failed proof" in p for p in checks["problems"]))

    def test_verify_without_prove_keeps_old_behavior(self):
        from orivellum.capabilities.workbench import _verify_output

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _error_workbook(out / "broken.xlsx")
            ok, checks = _verify_output("xlsx", out)
            self.assertTrue(ok)  # loads fine — old contract for imports
            self.assertNotIn("proof", checks)


def _publish_xlsx_version(db, cfg, project_id: str, checks: dict, good: bool = True):
    from orivellum.capabilities.workbench import _publish_version, _snapshot

    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "out"
        src.mkdir()
        (_good_workbook if good else _error_workbook)(src / "model.xlsx")
        files = _snapshot(src)
        return _publish_version(db, cfg, project_id, src, "test", files, checks)


class TestArchiveProofGate(unittest.TestCase):
    def test_archive_refuses_unproven_latest_unless_forced(self):
        from orivellum.capabilities.workbench import UnprovenError, archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")

            # no proof recorded at all → unproven
            _publish_xlsx_version(db, cfg, p["id"], checks={"workbooks": 1})
            with self.assertRaises(UnprovenError):
                archive_project(db, cfg, p["id"])
            self.assertEqual(db.get_wb_project(p["id"])["status"], "active")

            path = archive_project(db, cfg, p["id"], allow_unproven=True)
            self.assertTrue(Path(path).is_file())
            self.assertEqual(db.get_wb_project(p["id"])["status"], "archived")

    def test_archive_refuses_failed_proof(self):
        from orivellum.capabilities.workbench import UnprovenError, archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_xlsx_version(
                db,
                cfg,
                p["id"],
                checks={
                    "proof": {
                        "verdict": "failed",
                        "workbooks": {
                            "model.xlsx": {"verdict": "failed", "problems": ["bad cell"]}
                        },
                    }
                },
                good=False,
            )
            with self.assertRaises(UnprovenError) as ctx:
                archive_project(db, cfg, p["id"])
            self.assertIn("FAILED", str(ctx.exception))

    def test_archive_allows_proven_latest(self):
        from orivellum.capabilities.workbench import archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_xlsx_version(
                db,
                cfg,
                p["id"],
                checks={"proof": {"verdict": "proven", "workbooks": {}}},
            )
            path = archive_project(db, cfg, p["id"])
            self.assertTrue(Path(path).is_file())

    def test_archive_refuses_provable_import(self):
        """An import that would pass only after repairs is NOT archivable as
        proven — the verbatim bytes were never certified."""
        from orivellum.capabilities.workbench import UnprovenError, archive_project

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_xlsx_version(
                db,
                cfg,
                p["id"],
                checks={"proof": {"verdict": "provable", "workbooks": {}}},
            )
            with self.assertRaises(UnprovenError) as ctx:
                archive_project(db, cfg, p["id"])
            self.assertIn("verbatim", str(ctx.exception))

    def test_proof_inherited_when_workbook_bytes_unchanged(self):
        """Analysis / revert versions copy the workbook forward without
        re-gating; identical bytes must carry the earlier proof — and
        different bytes must NOT."""
        from orivellum.capabilities.workbench import (
            _publish_version,
            _snapshot,
            archive_project,
            latest_proof_status,
        )

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")

            with tempfile.TemporaryDirectory() as tmp2:
                src = Path(tmp2) / "out"
                src.mkdir()
                _good_workbook(src / "model.xlsx")
                v1_files = _snapshot(src)
                _publish_version(
                    db,
                    cfg,
                    p["id"],
                    src,
                    "build",
                    v1_files,
                    {"proof": {"verdict": "proven", "workbooks": {}}},
                )
                # analysis-style version: same workbook bytes + a report
                (src / "ANALYSIS_REPORT.md").write_text("# findings\n")
                _publish_version(
                    db, cfg, p["id"], src, "analyze", _snapshot(src), {"analysis": True}
                )

            proj = db.get_wb_project(p["id"])
            versions = db.list_wb_versions(p["id"])
            self.assertEqual(latest_proof_status(proj, versions)[0], "proven")
            self.assertTrue(Path(archive_project(db, cfg, p["id"])).is_file())

    def test_proof_not_inherited_when_bytes_differ(self):
        from orivellum.capabilities.workbench import latest_proof_status

        proj = {"kind": "xlsx"}
        versions = [
            {
                "version_no": 1,
                "files_json": json.dumps([{"name": "m.xlsx", "sha256": "a" * 64}]),
                "checks_json": json.dumps({"proof": {"verdict": "proven", "workbooks": {}}}),
            },
            {
                "version_no": 2,
                "files_json": json.dumps([{"name": "m.xlsx", "sha256": "b" * 64}]),
                "checks_json": json.dumps({}),
            },
        ]
        self.assertEqual(latest_proof_status(proj, versions)[0], "unproven")

    def test_non_xlsx_projects_unaffected(self):
        from orivellum.capabilities.workbench import archive_project, project_dir

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Tool", "code", "t")
            with tempfile.TemporaryDirectory() as tmp2:
                src = Path(tmp2) / "out"
                src.mkdir()
                (src / "main.py").write_text("x = 1\n")
                from orivellum.capabilities.workbench import _publish_version, _snapshot

                _publish_version(db, cfg, p["id"], src, "t", _snapshot(src), {})
            path = archive_project(db, cfg, p["id"])
            self.assertTrue(Path(path).is_file())
            self.assertTrue(project_dir(cfg, p["id"]).exists())


class TestCompleteRouteProofGate(unittest.TestCase):
    def test_complete_409_unproven_then_force(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, cfg = _make_app(tmp)
            client = TestClient(app)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_xlsx_version(db, cfg, p["id"], checks={"workbooks": 1})

            r = client.post(f"/api/workbench/projects/{p['id']}/complete", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 409, r.text)
            self.assertEqual(r.json()["detail"]["code"], "unproven")
            self.assertEqual(db.get_wb_project(p["id"])["status"], "active")

            r = client.post(
                f"/api/workbench/projects/{p['id']}/complete",
                headers=AUTH_HEADERS,
                json={"force": True},
            )
            self.assertEqual(r.status_code, 200, r.text)
            self.assertTrue(r.json()["archived"])
            self.assertEqual(db.get_wb_project(p["id"])["status"], "archived")


def _publish_provable_import(db, cfg, project_id: str) -> dict:
    """Mimic import_upload's v1: verbatim workbook + real promote=False proof."""
    from orivellum.capabilities.workbench import _publish_version, _snapshot
    from orivellum.capabilities.workbench_proof import prove_outputs

    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "out"
        src.mkdir()
        _good_workbook(src / "model.xlsx")
        proof = prove_outputs(src, [src / "model.xlsx"], promote=False)
        return _publish_version(
            db,
            cfg,
            project_id,
            src,
            "Imported from model.xlsx",
            _snapshot(src),
            {"imported": True, "proof": proof},
            verdict="imported",
        )


class TestRepairAndProve(unittest.TestCase):
    def test_provable_import_becomes_proven_new_version(self):
        from orivellum.capabilities.workbench import (
            archive_project,
            latest_proof_status,
            repair_and_prove,
            version_dir,
        )

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            v1 = _publish_provable_import(db, cfg, p["id"])
            self.assertEqual(json.loads(v1["checks_json"])["proof"]["verdict"], "provable", v1)
            v1_sha = _sha(version_dir(cfg, p["id"], 1) / "model.xlsx")

            row = repair_and_prove(db, cfg, p["id"], 1)
            self.assertEqual(row["version_no"], 2)
            self.assertEqual(row["verdict"], "proven")
            checks = json.loads(row["checks_json"])
            self.assertEqual(checks["proof"]["verdict"], "proven")
            self.assertEqual(checks["repaired_from"], 1)

            # the original imported version stays byte-for-byte verbatim
            self.assertEqual(_sha(version_dir(cfg, p["id"], 1) / "model.xlsx"), v1_sha)
            # the new version's workbook is the repaired one — real cached values
            wb = load_workbook(version_dir(cfg, p["id"], 2) / "model.xlsx", data_only=True)
            self.assertEqual(wb["Data"]["A3"].value, 42)
            wb.close()

            # Complete & archive now passes without force
            proj = db.get_wb_project(p["id"])
            versions = db.list_wb_versions(p["id"])
            self.assertEqual(latest_proof_status(proj, versions)[0], "proven")
            self.assertTrue(Path(archive_project(db, cfg, p["id"])).is_file())

    def test_uppercase_xlsx_import_gets_proof_and_repairs(self):
        """Excel files arrive as .XLSX too — the import must record a
        'provable' proof and repair & prove must find the workbook."""
        from orivellum.capabilities.workbench import (
            import_upload,
            repair_and_prove,
            version_dir,
        )

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            upload = Path(tmp) / "MODEL.XLSX"
            _good_workbook(upload)

            proj = import_upload(
                db, cfg, title="Upper", brief="b", upload_path=upload, filename="MODEL.XLSX"
            )
            self.assertEqual(proj["kind"], "xlsx")
            v1 = db.list_wb_versions(proj["id"])[0]
            self.assertEqual(json.loads(v1["checks_json"])["proof"]["verdict"], "provable")
            v1_sha = _sha(version_dir(cfg, proj["id"], 1) / "MODEL.XLSX")

            row = repair_and_prove(db, cfg, proj["id"], 1)
            self.assertEqual(row["verdict"], "proven")
            self.assertEqual(_sha(version_dir(cfg, proj["id"], 1) / "MODEL.XLSX"), v1_sha)
            wb = load_workbook(version_dir(cfg, proj["id"], 2) / "MODEL.XLSX", data_only=True)
            self.assertEqual(wb["Data"]["A3"].value, 42)
            wb.close()

    def test_failing_workbook_publishes_nothing(self):
        from orivellum.capabilities.workbench import (
            _publish_version,
            _snapshot,
            repair_and_prove,
        )

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Broken", "xlsx", "b")
            with tempfile.TemporaryDirectory() as tmp2:
                src = Path(tmp2) / "out"
                src.mkdir()
                _error_workbook(src / "model.xlsx")
                _publish_version(
                    db, cfg, p["id"], src, "import", _snapshot(src), {}, verdict="imported"
                )
            with self.assertRaises(ValueError) as ctx:
                repair_and_prove(db, cfg, p["id"], 1)
            self.assertIn("could not certify", str(ctx.exception))
            self.assertEqual(len(db.list_wb_versions(p["id"])), 1)

    def test_non_xlsx_and_missing_version_refused(self):
        from orivellum.capabilities.workbench import repair_and_prove

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            code = db.create_wb_project("Tool", "code", "t")
            with self.assertRaises(ValueError):
                repair_and_prove(db, cfg, code["id"], 1)
            xlsx = db.create_wb_project("Budget", "xlsx", "b")
            with self.assertRaises(FileNotFoundError):
                repair_and_prove(db, cfg, xlsx["id"], 7)


class TestRepairProveRoute(unittest.TestCase):
    def test_route_publishes_proven_version(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, cfg = _make_app(tmp)
            client = TestClient(app)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_provable_import(db, cfg, p["id"])

            r = client.post(
                f"/api/workbench/projects/{p['id']}/repair-prove",
                headers=AUTH_HEADERS,
                json={"version_no": 1},
            )
            self.assertEqual(r.status_code, 200, r.text)
            body = r.json()
            self.assertEqual(body["version_no"], 2)
            self.assertEqual(body["verdict"], "proven")
            # claim released — a follow-up mutation is not blocked
            self.assertEqual(db.get_wb_project(p["id"])["building"], 0)

            r = client.post(f"/api/workbench/projects/{p['id']}/complete", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 200, r.text)

    def test_route_404_unknown_version_and_409_while_building(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, cfg = _make_app(tmp)
            client = TestClient(app)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_provable_import(db, cfg, p["id"])

            r = client.post(
                f"/api/workbench/projects/{p['id']}/repair-prove",
                headers=AUTH_HEADERS,
                json={"version_no": 9},
            )
            self.assertEqual(r.status_code, 404, r.text)

            db.claim_wb_build(p["id"])
            r = client.post(
                f"/api/workbench/projects/{p['id']}/repair-prove",
                headers=AUTH_HEADERS,
                json={"version_no": 1},
            )
            self.assertEqual(r.status_code, 409, r.text)


def _publish_proven_build(db, cfg, project_id: str) -> dict:
    """Mimic run_build's accept path: prove with promotion (which writes the
    workbook_tests.json regression manifest), then snapshot + publish."""
    from orivellum.capabilities.workbench import _publish_version, _snapshot
    from orivellum.capabilities.workbench_proof import prove_outputs

    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "out"
        src.mkdir()
        _good_workbook(src / "model.xlsx")
        proof = prove_outputs(src, [src / "model.xlsx"], promote=True)
        assert proof["verdict"] == "proven", proof
        return _publish_version(
            db, cfg, project_id, src, "build", _snapshot(src), {"proof": proof}, verdict="proven"
        )


class TestRegressionManifest(unittest.TestCase):
    def test_proven_build_ships_manifest_import_does_not(self):
        from orivellum.capabilities.workbench_proof import MANIFEST_FILENAME, prove_outputs

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _good_workbook(out / "model.xlsx")
            # promote=False (import): verbatim rule — no manifest file
            prove_outputs(out, [out / "model.xlsx"], promote=False)
            self.assertFalse((out / MANIFEST_FILENAME).exists())

            proof = prove_outputs(out, [out / "model.xlsx"], promote=True)
            self.assertEqual(proof["verdict"], "proven")
            doc = json.loads((out / MANIFEST_FILENAME).read_text())
            manifest = doc["workbooks"]["model.xlsx"]
            self.assertEqual(manifest["formula_cells"], 2)
            cases = {(c["sheet"], c["cell"]): c["expected"] for c in manifest["cases"]}
            self.assertEqual(cases[("DATA", "A3")], 42)
            self.assertEqual(cases[("DATA", "B1")], 20)
            # the manifest never leaks into the persisted proof result
            self.assertNotIn("_manifest", proof["workbooks"]["model.xlsx"])
            # provenance: the proof records the digest of the exact bytes written
            payload = (out / MANIFEST_FILENAME).read_bytes()
            self.assertEqual(proof["manifest_sha256"], hashlib.sha256(payload).hexdigest())

    def test_manifest_included_in_published_version_and_verify_passes(self):
        from orivellum.capabilities.workbench import verify_latest, version_dir
        from orivellum.capabilities.workbench_proof import MANIFEST_FILENAME

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            v1 = _publish_proven_build(db, cfg, p["id"])
            names = [f["name"] for f in json.loads(v1["files_json"])]
            self.assertIn(MANIFEST_FILENAME, names)
            self.assertTrue((version_dir(cfg, p["id"], 1) / MANIFEST_FILENAME).is_file())

            res = verify_latest(db, cfg, p["id"])
            self.assertEqual(res["status"], "PASS", res)
            self.assertEqual(res["manifest_version"], 1)
            self.assertEqual(res["target_version"], 1)
            self.assertEqual(res["workbooks"]["model.xlsx"]["passed"], 2)

    def test_broken_formula_in_later_version_fails_manifest(self):
        from orivellum.capabilities.workbench import (
            _publish_version,
            _snapshot,
            verify_latest,
            version_dir,
        )

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_proven_build(db, cfg, p["id"])

            # someone edits the workbook and publishes v2 without re-proving
            with tempfile.TemporaryDirectory() as tmp2:
                src = Path(tmp2) / "out"
                shutil.copytree(version_dir(cfg, p["id"], 1), src)
                wb = load_workbook(src / "model.xlsx")  # keeps formulas
                wb["Data"]["A1"] = 11  # A3 now computes 43, manifest expects 42
                wb.save(src / "model.xlsx")
                _publish_version(db, cfg, p["id"], src, "manual edit", _snapshot(src), {})

            res = verify_latest(db, cfg, p["id"])
            self.assertEqual(res["status"], "FAIL", res)
            self.assertEqual(res["target_version"], 2)
            run = res["workbooks"]["model.xlsx"]
            failed_refs = {(f["sheet"], f["cell"]) for f in run["failed"]}
            self.assertIn(("DATA", "A3"), failed_refs)

    def test_verify_refuses_without_manifest_or_versions(self):
        from orivellum.capabilities.workbench import verify_latest

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            with self.assertRaises(FileNotFoundError):
                verify_latest(db, cfg, p["id"])  # no versions
            _publish_provable_import(db, cfg, p["id"])  # import: no manifest
            with self.assertRaises(ValueError) as ctx:
                verify_latest(db, cfg, p["id"])
            self.assertIn("workbook_tests.json", str(ctx.exception))
            code = db.create_wb_project("Tool", "code", "t")
            with self.assertRaises(ValueError):
                verify_latest(db, cfg, code["id"])

    def test_verify_route(self):
        with tempfile.TemporaryDirectory() as tmp:
            app, db, cfg = _make_app(tmp)
            client = TestClient(app)
            p = db.create_wb_project("Budget", "xlsx", "b")

            r = client.post("/api/workbench/projects/nope/verify", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 404)

            _publish_provable_import(db, cfg, p["id"])
            r = client.post(f"/api/workbench/projects/{p['id']}/verify", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 422, r.text)

            _publish_proven_build(db, cfg, p["id"])
            r = client.post(f"/api/workbench/projects/{p['id']}/verify", headers=AUTH_HEADERS)
            self.assertEqual(r.status_code, 200, r.text)
            body = r.json()
            self.assertEqual(body["status"], "PASS")
            self.assertEqual(body["manifest_version"], 2)

    def test_forged_import_manifest_is_never_trusted(self):
        """An import zip can carry a file named workbook_tests.json — chosen
        expectations must never become a regression authority."""
        from orivellum.capabilities.workbench import _publish_version, _snapshot, verify_latest
        from orivellum.capabilities.workbench_proof import MANIFEST_FILENAME

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            with tempfile.TemporaryDirectory() as tmp2:
                src = Path(tmp2) / "out"
                src.mkdir()
                _error_workbook(src / "model.xlsx")
                forged = {
                    "format": 1,
                    "workbooks": {"model.xlsx": {"cases": []}},  # passes anything
                }
                (src / MANIFEST_FILENAME).write_text(json.dumps(forged))
                _publish_version(
                    db, cfg, p["id"], src, "import", _snapshot(src), {}, verdict="imported"
                )
            with self.assertRaises(ValueError) as ctx:
                verify_latest(db, cfg, p["id"])
            self.assertIn("no proven version", str(ctx.exception))

    def test_tampered_manifest_is_refused(self):
        from orivellum.capabilities.workbench import verify_latest, version_dir
        from orivellum.capabilities.workbench_proof import MANIFEST_FILENAME

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_proven_build(db, cfg, p["id"])
            mp = version_dir(cfg, p["id"], 1) / MANIFEST_FILENAME
            doc = json.loads(mp.read_text())
            doc["workbooks"]["model.xlsx"]["cases"] = []  # weaken the tests
            mp.write_text(json.dumps(doc))
            with self.assertRaises(ValueError) as ctx:
                verify_latest(db, cfg, p["id"])
            self.assertIn("digest", str(ctx.exception))

    def test_malformed_and_hostile_manifests_refused(self):
        from orivellum.capabilities.workbench_proof import validate_manifest_doc

        good_case = {"sheet": "DATA", "cell": "A1", "expected": 1}
        for bad in [
            None,
            [],
            {"format": 2, "workbooks": {"m.xlsx": {"cases": [good_case]}}},
            {"format": 1, "workbooks": {}},
            {"format": 1, "workbooks": {"../escape.xlsx": {"cases": [good_case]}}},
            {"format": 1, "workbooks": {"/abs.xlsx": {"cases": [good_case]}}},
            {"format": 1, "workbooks": {"m.txt": {"cases": [good_case]}}},
            {"format": 1, "workbooks": {"m.xlsx": {"cases": "nope"}}},
            {"format": 1, "workbooks": {"m.xlsx": {"cases": [{"sheet": "D"}]}}},
            {
                "format": 1,
                "workbooks": {"m.xlsx": {"cases": [{**good_case, "expected": float("nan")}]}},
            },
        ]:
            with self.assertRaises(ValueError, msg=repr(bad)):
                validate_manifest_doc(bad)
        validate_manifest_doc({"format": 1, "workbooks": {"sub/m.xlsx": {"cases": [good_case]}}})

    def test_repair_and_prove_ships_manifest(self):
        from orivellum.capabilities.workbench import repair_and_prove, verify_latest
        from orivellum.capabilities.workbench_proof import MANIFEST_FILENAME

        with tempfile.TemporaryDirectory() as tmp:
            _, db, cfg = _make_app(tmp)
            p = db.create_wb_project("Budget", "xlsx", "b")
            _publish_provable_import(db, cfg, p["id"])
            row = repair_and_prove(db, cfg, p["id"], 1)
            names = [f["name"] for f in json.loads(row["files_json"])]
            self.assertIn(MANIFEST_FILENAME, names)
            self.assertEqual(verify_latest(db, cfg, p["id"])["status"], "PASS")


class TestLatestProofStatus(unittest.TestCase):
    def test_status_mapping(self):
        from orivellum.capabilities.workbench import latest_proof_status

        proj = {"kind": "xlsx"}
        self.assertEqual(latest_proof_status(proj, [])[0], "n/a")
        self.assertEqual(latest_proof_status({"kind": "code"}, [{}])[0], "n/a")

        def v(checks):
            return [{"version_no": 1, "checks_json": json.dumps(checks)}]

        self.assertEqual(latest_proof_status(proj, v({}))[0], "unproven")
        self.assertEqual(
            latest_proof_status(proj, v({"proof": {"verdict": "proven", "workbooks": {}}}))[0],
            "proven",
        )
        status, detail = latest_proof_status(
            proj,
            v(
                {
                    "proof": {
                        "verdict": "failed",
                        "workbooks": {"a.xlsx": {"verdict": "failed", "problems": ["boom"]}},
                    }
                }
            ),
        )
        self.assertEqual(status, "failed")
        self.assertIn("boom", detail)


if __name__ == "__main__":
    unittest.main()
