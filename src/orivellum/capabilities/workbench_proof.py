"""Workbench ↔ Orivellum Runner bridge — six-gate workbook proving.

The Project Workbench used to accept an xlsx version when the workbook merely
*loaded* in openpyxl. The repo already owns a far stronger standard: the
Orivellum Runner's proof gates (recalculate every formula with the pure-Python
``formulas`` engine, values must match, no saved error cells, canonical OOXML
child order, surgery containment, clean double-load). This module runs that
exact gate suite in-process against a workbench output file.

Semantics mirror ``runner/jobs/xlsx.py::prove_unit``:

- Repairs first (OOXML reorder + refreshing stale/missing cached values from
  the recalculation), applied to a hidden candidate copy — the original is
  untouched until every gate passes, then atomically promoted. LLM-built
  workbooks almost never carry cached values, so the repair step is what makes
  an archived workbook open with real numbers instead of blanks.
- HONESTY RULE: an unavailable engine or an uncomputable function yields
  verdict ``unverified`` — reported as such, never as clean.

The runner stays a standalone harness (own requirements, own CLI); we import
only its two pure modules (engine + surgery: stdlib/openpyxl/formulas, no
runner config or store) via a lazy path insert.
"""

from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import pathlib
import re
import sys
import time
import zipfile
from collections import defaultdict
from typing import Any

logger = logging.getLogger("orivellum.workbench.proof")

MANIFEST_FILENAME = "workbook_tests.json"

# The proof recalculates every formula with the pure-Python `formulas` engine
# — twice (repair plan + candidate gates). Measured on this hardware:
# ~400 formulas ≈ 1.2 s, ~2,000 ≈ 5 s, ~6,000 ≈ 18 s (roughly linear), so the
# default ceiling of 20,000 caps a proof run at about a minute or two.
# Workbooks above it are honestly reported as `unverified` instead of holding
# the build claim hostage for the duration.
DEFAULT_FORMULA_CEILING = 20_000


def _formula_ceiling() -> int:
    try:
        return max(1, int(os.environ.get("ORIVELLUM_PROOF_FORMULA_CEILING", "")))
    except ValueError:
        return DEFAULT_FORMULA_CEILING


# Worksheet XML the preflight will stream before declaring a workbook too
# large to prove — a 20,000-formula sheet is a few MB, so 64 MB is generous.
_COUNT_XML_BUDGET = 64 * 1024 * 1024


class _OverLimit(Exception):
    """Internal: preflight counter passed the ceiling — stop parsing."""


def _feed_sheet_xml(fh, parser, scanned: int) -> int:
    """Stream one worksheet part into the parser; -1 when the cumulative
    scan budget is spent."""
    while True:
        chunk = fh.read(1 << 16)
        if not chunk:
            parser.Parse(b"", True)
            return scanned
        scanned += len(chunk)
        if scanned > _COUNT_XML_BUDGET:
            return -1
        parser.Parse(chunk, False)


def _count_formula_cells(path: pathlib.Path, limit: int) -> int:
    """Formula cells in the workbook, stopping at ``limit + 1``.

    Streams the worksheet XML through expat (a real XML parser, so
    namespace-prefixed ``<x:f>`` and any attribute whitespace count exactly
    like ``<f t="shared">``) instead of iterating coordinates — an inflated
    ``<dimension>`` or a sparse sheet with formulas at row one million
    serializes only its real cells, so the scan is bounded by actual file
    content, never by declared grid size. Counting by local name ``f`` can
    only ever over-count exotic non-SpreadsheetML elements — errs toward
    'too large', never toward running the engine. FAIL CLOSED: exhausting
    the XML scan budget returns ``limit + 1`` (too large to prove safely),
    never a stall. Returns 0 on unreadable files (the engine reports those
    honestly itself)."""
    import xml.parsers.expat

    state = {"n": 0}

    def _start(tag, _attrs):
        # expat qnames keep the raw prefix (e.g. 'x:f'); local name decides
        if tag.rpartition(":")[2] == "f":
            state["n"] += 1
            if state["n"] > limit:
                raise _OverLimit

    scanned = 0
    try:
        with zipfile.ZipFile(path) as z:
            for name in z.namelist():
                if not _SHEET_XML.match(name):
                    continue
                parser = xml.parsers.expat.ParserCreate()
                parser.StartElementHandler = _start
                with z.open(name) as fh:
                    scanned = _feed_sheet_xml(fh, parser, scanned)
                if scanned < 0:
                    return limit + 1
    except _OverLimit:
        return state["n"]
    except Exception:  # noqa: BLE001
        logger.debug("Formula count failed for %s", path, exc_info=True)
        return 0
    return state["n"]


GATE_NAMES = (
    "G1_recalc_covers_all",
    "G2_values_match",
    "G3_no_error_cells",
    "G4_ooxml_order",
    "G5_surgery_contained",
    "G6_loads_clean",
)

_SHEET_XML = re.compile(r"xl/worksheets/sheet\d+\.xml$")


def _runner_dir() -> pathlib.Path:
    override = os.environ.get("ORIVELLUM_RUNNER_DIR")
    if override:
        return pathlib.Path(override)
    # src/orivellum/capabilities/workbench_proof.py → repo root / orivellum-runner
    return pathlib.Path(__file__).resolve().parents[3] / "orivellum-runner"


_runner_cache: tuple | None = None


def _load_runner_modules():
    """Return (engine, surgery) or (None, None) when the harness is missing.

    Loads the two pure modules by absolute file path under private module
    names — never touches ``sys.path``, so an unrelated ``runner`` package
    elsewhere in the process can neither be shadowed nor picked up by
    mistake. Never raises — an absent harness must degrade to
    ``unverified``, not break a build.
    """
    global _runner_cache
    if _runner_cache is not None:
        return _runner_cache
    try:
        import importlib.util
        import types

        jobs = _runner_dir() / "runner" / "jobs"

        # Synthetic parent package so the modules' own relative imports
        # (e.g. structural_checks' ``from . import xlsx_surgery``) resolve —
        # still never touches ``sys.path``.
        pkg_name = "_orivellum_runner_jobs"
        pkg = types.ModuleType(pkg_name)
        pkg.__path__ = [str(jobs)]
        sys.modules[pkg_name] = pkg

        def _load(alias: str, path: pathlib.Path):
            spec = importlib.util.spec_from_file_location(alias, path)
            if spec is None or spec.loader is None:
                raise ImportError(f"cannot load {path}")
            mod = importlib.util.module_from_spec(spec)
            sys.modules[alias] = mod
            spec.loader.exec_module(mod)
            return mod

        # surgery first — engine's structural helpers reference it lazily,
        # but the functions we call (recalculate/compare/ERRVALS) do not.
        surgery = _load(f"{pkg_name}.xlsx_surgery", jobs / "xlsx_surgery.py")
        engine = _load(f"{pkg_name}.xlsx_engine", jobs / "xlsx_engine.py")
        _runner_cache = (engine, surgery)
    except Exception:  # noqa: BLE001
        logger.warning("Orivellum Runner proof modules unavailable", exc_info=True)
        _runner_cache = (None, None)
    return _runner_cache


def _unverified(error: str) -> dict:
    return {"verdict": "unverified", "error": error[:400], "gates": {}, "problems": [error[:400]]}


def prove_workbook(path: pathlib.Path, promote: bool = True) -> dict[str, Any]:
    """Repair → recalculate → six gates → atomic promotion, on ONE workbook.

    Returns a JSON-safe dict:
      {verdict: proven|failed|unverified, gates: {G1..G6: bool},
       failed_gates: [...], problems: [...], repairs: {reordered_parts,
       refreshed_cells}, recalc: {formulas_checked, agreed}}

    On ``proven`` with ``promote=True`` the file at *path* is atomically
    replaced by the repaired, fully-gated candidate. With ``promote=False``
    (imported versions stay verbatim) the original is never touched, and the
    verdict is honest about WHICH bytes passed: ``proven`` only when no
    repairs were needed (candidate == original); ``provable`` when the gates
    pass only after repairs the file itself never received.

    Every result carries ``limits`` (the formula ceiling in force) and
    ``elapsed_seconds``. A workbook above the ceiling is never recalculated —
    it returns ``unverified`` with a clear reason instead of stalling the
    build pipeline for minutes.
    """
    started = time.monotonic()
    result = _prove_workbook_inner(path, promote)
    result["limits"] = {"formula_ceiling": _formula_ceiling()}
    result["elapsed_seconds"] = round(time.monotonic() - started, 2)
    return result


def _prove_workbook_inner(path: pathlib.Path, promote: bool) -> dict[str, Any]:
    engine, surgery = _load_runner_modules()
    if engine is None or surgery is None:
        return _unverified("proving harness unavailable (orivellum-runner not found)")

    ceiling = _formula_ceiling()
    n = _count_formula_cells(path, ceiling)
    if n > ceiling:
        res = _unverified(
            f"workbook too large to prove safely: over {ceiling} formula cells "
            "or worksheet XML beyond the scan budget (raise "
            "ORIVELLUM_PROOF_FORMULA_CEILING to override); proving skipped so "
            "the build pipeline is never stalled"
        )
        res["formula_cells"] = n
        return res

    recalc = engine.recalculate(path)
    if not recalc["available"]:
        return _unverified(recalc["error"] or "recalculation unavailable")

    edits, reordered, refreshed = _plan_repairs(engine, surgery, path, recalc)
    return _run_gates(engine, surgery, path, edits, reordered, refreshed, promote)


def _plan_repairs(engine, surgery, path, recalc):
    """R1 (OOXML reorder) + R2 (refresh stale/missing caches) — the same
    repair plan as the runner's prove_unit. Returns (edits, reordered,
    refreshed_count)."""
    cmp0 = engine.compare(path, recalc)
    part_of = surgery.sheet_part_names(path)
    edits: dict[str, str] = {}
    reordered: list[str] = []
    refreshed = 0
    with zipfile.ZipFile(path) as z:
        raw = {
            n: z.read(n).decode("utf-8", errors="replace")
            for n in z.namelist()
            if _SHEET_XML.match(n)
        }
    for part, xml in raw.items():
        if surgery.sheet_order_violations(xml):
            fixed = surgery.reorder_sheet_xml(xml)
            if fixed != xml:  # reorder REFUSES on unknown content — G4 then fails
                raw[part] = fixed
                edits[part] = fixed
                reordered.append(part)
    by_sheet: dict[str, dict] = defaultdict(dict)
    for m in cmp0["mismatches"]:
        computed = m["computed_raw"]
        if isinstance(computed, str) and computed.startswith("#"):
            continue  # a formula that genuinely errors is not repairable
        by_sheet[m["sheet"]][m["cell"]] = computed
    for sheet, updates in by_sheet.items():
        part = part_of.get(sheet)
        if not part or part not in raw:
            continue
        new_xml, changed = surgery.refresh_cached(raw[part], updates)
        if changed:
            raw[part] = new_xml
            edits[part] = new_xml
            refreshed += len(changed)
    return edits, reordered, refreshed


def _scan_error_cells(engine, candidate) -> list[str]:
    """G3: every saved cell value that is an Excel error literal."""
    from openpyxl import load_workbook

    err_cells = []
    wv = load_workbook(candidate, data_only=True, read_only=True)
    try:
        for name in wv.sheetnames:
            for row in wv[name].iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() in engine.ERRVALS:
                        err_cells.append(f"{name}!{c.coordinate}")
    finally:
        wv.close()
    return err_cells


def _scan_sheet_order(surgery, candidate) -> list[str]:
    """G4: worksheet parts whose OOXML child order is non-canonical."""
    order_bad = []
    with zipfile.ZipFile(candidate) as z:
        for n in [x for x in z.namelist() if _SHEET_XML.match(x)]:
            if surgery.sheet_order_violations(z.read(n).decode("utf-8", errors="replace")):
                order_bad.append(n)
    return order_bad


def _passing_verdict(path, candidate, promote: bool, repaired: bool, problems: list) -> str:
    """All six gates passed — decide what that honestly certifies."""
    if promote:
        candidate.replace(path)  # atomic promotion — gates first, always
        return "proven"
    if not repaired:
        # candidate is byte-equivalent to the original — the file itself
        # passed every gate
        return "proven"
    # gates pass only after repairs the file never received; 'proven' would
    # certify bytes that were never archived
    problems.append(
        "passes all gates only after cache/order repairs; the file itself "
        "was kept verbatim and would show stale or blank values in Excel"
    )
    return "provable"


def _run_gates(engine, surgery, path, edits, reordered, refreshed, promote):
    """Apply the repair plan to a hidden candidate, run all six gates against
    it, and promote atomically only when every gate passes (and promotion is
    allowed). The original file is never touched on any other outcome."""
    import uuid

    from openpyxl import load_workbook

    candidate = path.with_name(f".candidate_{uuid.uuid4().hex[:8]}_{path.name}")
    gates: dict[str, bool] = {}
    problems: list[str] = []
    result: dict[str, Any] = {
        "verdict": "failed",
        "gates": gates,
        "problems": problems,
        "repairs": {"reordered_parts": len(reordered), "refreshed_cells": refreshed},
    }
    try:
        touched = surgery.apply(path, candidate, edits)

        recalc2 = engine.recalculate(candidate)
        if not recalc2["available"]:
            return _unverified(recalc2["error"] or "engine failed on repaired candidate")
        cmp2 = engine.compare(candidate, recalc2)
        gates["G1_recalc_covers_all"] = not cmp2["uncovered"]
        gates["G2_values_match"] = not cmp2["mismatches"]
        if cmp2["uncovered"]:
            problems.append(f"engine could not cover: {', '.join(cmp2['uncovered'][:5])}")
        for m in cmp2["mismatches"][:5]:
            problems.append(f"{m['ref']}: cached={m['cached']} computed={m['computed']}")

        err_cells = _scan_error_cells(engine, candidate)
        gates["G3_no_error_cells"] = not err_cells
        if err_cells:
            problems.append(f"error cells: {', '.join(err_cells[:5])}")

        order_bad = _scan_sheet_order(surgery, candidate)
        gates["G4_ooxml_order"] = not order_bad
        if order_bad:
            problems.append(f"OOXML order violated: {', '.join(order_bad[:3])}")

        diff = surgery.parts_diff(path, candidate)
        gates["G5_surgery_contained"] = (
            not diff["added"] and not diff["removed"] and set(diff["changed"]) <= set(touched)
        )
        try:
            load_workbook(candidate, data_only=False).close()
            load_workbook(candidate, data_only=True).close()
            gates["G6_loads_clean"] = True
        except Exception as exc:  # noqa: BLE001
            gates["G6_loads_clean"] = False
            problems.append(f"candidate fails to load: {str(exc)[:200]}")

        result["recalc"] = {
            "formulas_checked": len(cmp2.get("formula_cells", [])),
            "agreed": cmp2.get("agreed", 0),
        }
        if all(gates.values()):
            result["verdict"] = _passing_verdict(path, candidate, promote, bool(edits), problems)
            if result["verdict"] == "proven":
                # Regression manifest for the exact gated bytes: every formula
                # cell with its engine-computed expected value. Private key —
                # prove_outputs pops it before the result is persisted.
                result["_manifest"] = engine.build_test_manifest(path, recalc2, cmp2)
        else:
            result["failed_gates"] = sorted(k for k, v in gates.items() if not v)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Proof run crashed for %s", path)
        return _unverified(f"proof run crashed: {type(exc).__name__}: {exc}")
    finally:
        candidate.unlink(missing_ok=True)  # no gate suite, no candidate left
    return result


def prove_outputs(
    out_dir: pathlib.Path, workbooks: list[pathlib.Path], promote: bool = True
) -> dict[str, Any]:
    """Prove every workbook of a version; aggregate to one verdict.

    proven     — every workbook's ACTUAL bytes passed all six gates
    provable   — gates pass, but only after repairs a verbatim file never got
    failed     — at least one workbook failed a gate
    unverified — none failed, but at least one could not be recalculated

    When every workbook is proven AND promotion is allowed, a regression
    manifest (``workbook_tests.json``) is written into *out_dir* so it ships
    with the version's files: every formula cell with its engine-computed
    expected value, re-runnable against any future edit. Its sha256 is
    returned as ``manifest_sha256`` — PROVENANCE RULE: only a manifest whose
    bytes still match a digest recorded at proving time may ever be treated
    as a regression authority (a filename alone proves nothing; an import
    zip may carry an arbitrary file of the same name). Imports
    (promote=False) never get one — an imported version stays verbatim.
    """
    per_file: dict[str, dict] = {}
    manifests: dict[str, dict] = {}
    for p in workbooks:
        rel = str(p.relative_to(out_dir))
        res = prove_workbook(p, promote=promote)
        manifest = res.pop("_manifest", None)
        if manifest is not None:
            manifests[rel] = manifest
        per_file[rel] = res
    verdicts = {r["verdict"] for r in per_file.values()}
    for worst in ("failed", "unverified", "provable"):
        if worst in verdicts:
            return {"verdict": worst, "workbooks": per_file}
    if promote and manifests:
        doc = {
            "format": 1,
            "generated_by": "orivellum workbench six-gate proof",
            "workbooks": manifests,
        }
        # allow_nan=False: a non-finite expectation must fail loudly here,
        # never become non-standard JSON that a re-run parses differently.
        payload = json.dumps(doc, indent=1, sort_keys=True, allow_nan=False)
        (out_dir / MANIFEST_FILENAME).write_text(payload, encoding="utf-8")
        return {
            "verdict": "proven",
            "workbooks": per_file,
            "manifest_sha256": hashlib.sha256(payload.encode("utf-8")).hexdigest(),
        }
    return {"verdict": "proven", "workbooks": per_file}


def _safe_rel_xlsx(rel: str) -> bool:
    """True only for a normalized, relative, confinement-safe .xlsx path."""
    if not isinstance(rel, str) or not rel or "\\" in rel or ":" in rel or "\x00" in rel:
        return False
    p = pathlib.PurePosixPath(rel)
    return (
        not p.is_absolute()
        and p.suffix.lower() == ".xlsx"
        and all(part not in ("", ".", "..") for part in p.parts)
    )


def _json_scalar_ok(v) -> bool:
    if isinstance(v, float):
        return math.isfinite(v)
    return v is None or isinstance(v, (bool, int, str))


def validate_manifest_doc(doc) -> None:
    """Strict format-1 schema check — raises ValueError on anything off.

    A manifest is an executable expectation set; a malformed or hostile one
    must be REFUSED, never partially run: paths confined to relative .xlsx
    (no traversal, no absolutes), cases are {sheet, cell, expected} with
    finite JSON-scalar expectations.
    """
    if not isinstance(doc, dict) or doc.get("format") != 1:
        raise ValueError(f"not a format-1 {MANIFEST_FILENAME} document")
    wbs = doc.get("workbooks")
    if not isinstance(wbs, dict) or not wbs:
        raise ValueError("manifest lists no workbooks")
    for rel, manifest in wbs.items():
        if not _safe_rel_xlsx(rel):
            raise ValueError(f"unsafe workbook path in manifest: {rel!r}")
        if not isinstance(manifest, dict) or not isinstance(manifest.get("cases"), list):
            raise ValueError(f"malformed manifest entry for {rel}")
        for case in manifest["cases"]:
            if (
                not isinstance(case, dict)
                or not isinstance(case.get("sheet"), str)
                or not isinstance(case.get("cell"), str)
                or not _json_scalar_ok(case.get("expected"))
            ):
                raise ValueError(f"malformed test case in manifest entry for {rel}")


def run_saved_manifest(files_dir: pathlib.Path, doc: dict) -> dict[str, Any]:
    """Re-run a saved ``workbook_tests.json`` against the workbooks in
    *files_dir* — the regression check for a proven version's descendants.

    Returns {status: PASS|FAIL|UNAVAILABLE, workbooks: {rel: run_manifest
    result}}. HONESTY RULE: a missing workbook or unavailable engine is
    never a pass. Aggregate is FAIL if anything failed, else UNAVAILABLE if
    anything could not be recalculated, else PASS. Raises ValueError for a
    malformed or unsafe manifest document (never partially runs one).
    """
    validate_manifest_doc(doc)
    engine, _surgery = _load_runner_modules()
    if engine is None:
        return {
            "status": "UNAVAILABLE",
            "error": "proving harness unavailable (orivellum-runner not found)",
            "workbooks": {},
        }
    results: dict[str, dict] = {}
    root = files_dir.resolve()
    for rel, manifest in doc["workbooks"].items():
        target = (files_dir / rel).resolve()
        if root != target and root not in target.parents:
            # validate_manifest_doc already refuses traversal; belt over
            # suspenders for symlinked version dirs
            raise ValueError(f"manifest workbook path escapes the version files: {rel!r}")
        if not target.is_file():
            results[rel] = {
                "status": "FAIL",
                "error": "workbook missing from this version",
                "passed": 0,
                "failed": [],
                "total": len(manifest.get("cases") or []),
            }
            continue
        ceiling = _formula_ceiling()
        if _count_formula_cells(target, ceiling) > ceiling:
            # same guard as proving — a re-run recalculates too, and honesty
            # beats a stalled request: report UNAVAILABLE, never a pass
            results[rel] = {
                "status": "UNAVAILABLE",
                "error": f"workbook exceeds the {ceiling}-formula proving ceiling",
                "passed": 0,
                "failed": [],
                "total": len(manifest.get("cases") or []),
            }
            continue
        try:
            results[rel] = engine.run_manifest(target, manifest)
        except Exception as exc:  # noqa: BLE001 - explicit, never a silent pass
            logger.exception("Manifest re-run crashed for %s", target)
            results[rel] = {
                "status": "FAIL",
                "error": f"re-run crashed: {type(exc).__name__}: {exc}"[:400],
                "passed": 0,
                "failed": [],
                "total": len(manifest.get("cases") or []),
            }
    statuses = {r["status"] for r in results.values()}
    if not results:
        return {"status": "UNAVAILABLE", "error": "manifest lists no workbooks", "workbooks": {}}
    status = (
        "FAIL" if "FAIL" in statuses else ("UNAVAILABLE" if "UNAVAILABLE" in statuses else "PASS")
    )
    return {"status": status, "workbooks": results}
