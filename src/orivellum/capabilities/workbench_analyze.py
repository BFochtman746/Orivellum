"""Workbench analyze mode — review an existing project version and produce
a findings report as a new immutable version.

Deterministic checks come first (they never depend on a model):

- **xlsx**: load in both views, cached error values, broken cross-sheet
  references (formulas naming sheets that do not exist), volatile
  functions, external workbook links, merged cells, hidden sheets/rows/
  columns, README-sheet presence, and an independent formula
  recalculation with the pure-Python ``formulas`` engine comparing
  computed results to the cached values Excel saved.
- **code**: file inventory, ``.py`` syntax errors, ``.json`` validity,
  TODO/FIXME markers, test presence, dependency manifests.

The local LLM then writes a plain-language narrative on top of those
findings; if the model is unavailable the report says so explicitly and
still ships the deterministic findings — never a silent gap.

The report is published as ``ANALYSIS_REPORT.md`` in a new version whose
files are otherwise a byte-identical copy of the version analyzed, so the
"every version is the complete project state" invariant holds.
"""

from __future__ import annotations

import ast
import datetime
import json
import logging
import pathlib
import re
import shutil
import tempfile
from typing import Any

logger = logging.getLogger(__name__)

REPORT_FILENAME = "ANALYSIS_REPORT.md"

_ERROR_LITERALS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
_VOLATILE_RE = re.compile(r"\b(NOW|TODAY|RAND|RANDBETWEEN|OFFSET|INDIRECT)\s*\(", re.IGNORECASE)
# 'Quoted Sheet Name'!A1  and  BareName!A1 (quoted first so it wins)
_SHEET_REF_RE = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_.'])([A-Za-z_][A-Za-z0-9_.]*)!")
_MAX_LISTED = 20  # cap per-issue example lists in the report
_RECALC_MAX_FORMULAS = 20000
_RECALC_TOLERANCE = 1e-6

_CODE_EXTS = {
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".java",
    ".go",
    ".rs",
    ".c",
    ".h",
    ".cpp",
    ".cs",
    ".rb",
    ".php",
    ".sh",
    ".ps1",
}


# ── xlsx checks ───────────────────────────────────────────────────────────────


def analyze_workbook(path: pathlib.Path) -> dict[str, Any]:
    """Deterministic findings for one workbook. Never raises — a workbook
    that cannot be opened is itself a finding."""
    from openpyxl import load_workbook

    from orivellum.capabilities.workbench import check_xlsx_zip_safety

    out: dict[str, Any] = {"file": path.name}
    bomb = check_xlsx_zip_safety(path)
    if bomb:
        out["load_error"] = bomb
        return out
    try:
        wb = load_workbook(path, data_only=False)
        wbv = load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        out["load_error"] = str(exc)[:300]
        return out

    sheet_names = list(wb.sheetnames)
    known = {s.upper() for s in sheet_names}
    for dn in wb.defined_names:
        known.add(dn.upper())

    acc = {
        "cells": 0,
        "formulas": 0,
        "error_cells": [],
        "broken_refs": [],
        "volatile": [],
        "external": [],
    }
    merged = 0
    hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    hidden_rows = hidden_cols = 0

    for ws, wsv in zip(wb.worksheets, wbv.worksheets, strict=False):
        merged += len(ws.merged_cells.ranges)
        hidden_rows += sum(1 for d in ws.row_dimensions.values() if d.hidden)
        hidden_cols += sum(1 for d in ws.column_dimensions.values() if d.hidden)
        _scan_sheet_cells(ws, wsv, known, acc)
    wb.close()
    wbv.close()

    broken = sorted(set(acc["broken_refs"]))
    out.update(
        {
            "sheets": sheet_names,
            "cells": acc["cells"],
            "formulas": acc["formulas"],
            "error_cells": acc["error_cells"][:_MAX_LISTED],
            "error_cell_count": len(acc["error_cells"]),
            "broken_references": broken[:_MAX_LISTED],
            "broken_reference_count": len(broken),
            "volatile_functions": acc["volatile"][:_MAX_LISTED],
            "volatile_count": len(acc["volatile"]),
            "external_links": acc["external"][:_MAX_LISTED],
            "external_link_count": len(acc["external"]),
            "merged_ranges": merged,
            "hidden_sheets": hidden_sheets,
            "hidden_rows": hidden_rows,
            "hidden_cols": hidden_cols,
            "has_readme_sheet": any(s.strip().upper() == "README" for s in sheet_names),
        }
    )
    out["recalc"] = _recalc_workbook(path, acc["formulas"])
    return out


def _scan_sheet_cells(ws, wsv, known: set[str], acc: dict) -> None:
    """Scan one worksheet (formula view *ws*, value view *wsv*) into *acc*."""
    for row, rowv in zip(ws.iter_rows(), wsv.iter_rows(), strict=False):
        for c, cv in zip(row, rowv, strict=False):
            if c.value is None:
                continue
            acc["cells"] += 1
            loc = f"{ws.title}!{c.coordinate}"
            if isinstance(cv.value, str) and cv.value in _ERROR_LITERALS:
                acc["error_cells"].append(f"{loc} = {cv.value}")
            f = c.value
            if not (isinstance(f, str) and f.startswith("=")):
                continue
            acc["formulas"] += 1
            if _VOLATILE_RE.search(f):
                acc["volatile"].append(loc)
            if "[" in f:
                acc["external"].append(loc)
            for m in _SHEET_REF_RE.finditer(f):
                name = (m.group(1) or m.group(2) or "").strip()
                if name and name.upper() not in known and not name.upper().startswith("["):
                    acc["broken_refs"].append(f"{loc} references missing sheet '{name}'")


def _recalc_workbook(path: pathlib.Path, formula_count: int) -> dict[str, Any]:
    """Independently recalculate the workbook with the pure-Python
    ``formulas`` engine and compare against Excel's cached values.
    Unavailable/failed recalculation is reported, never hidden."""
    if formula_count == 0:
        return {"status": "skipped", "reason": "no formulas"}
    if formula_count > _RECALC_MAX_FORMULAS:
        return {"status": "skipped", "reason": f"too many formulas ({formula_count})"}
    try:
        import formulas as _formulas
    except Exception:  # noqa: BLE001  pragma: no cover - dependency is pinned
        return {"status": "unavailable", "reason": "recalculation engine not installed"}
    try:
        from openpyxl import load_workbook

        model = _formulas.ExcelModel().loads(str(path)).finish()
        solution = model.calculate()
        wbf = load_workbook(path, data_only=False)
        wbv = load_workbook(path, data_only=True)
        upper_to_sheet = {s.upper(): s for s in wbf.sheetnames}
        checked = mismatches = 0
        examples: list[str] = []
        for key, value in solution.items():
            pair = _cached_vs_computed(key, value, wbf, wbv, upper_to_sheet)
            if pair is None:
                continue
            label, cached, computed = pair
            checked += 1
            if abs(cached - computed) > max(_RECALC_TOLERANCE, abs(cached) * 1e-9):
                mismatches += 1
                if len(examples) < _MAX_LISTED:
                    examples.append(f"{label}: saved {cached!r} vs computed {computed!r}")
        wbf.close()
        wbv.close()
        return {
            "status": "ok",
            "numeric_formulas_checked": checked,
            "mismatches": mismatches,
            "mismatch_examples": examples,
        }
    except Exception as exc:  # noqa: BLE001
        return {"status": "engine_error", "reason": str(exc)[:300]}


_SOLUTION_KEY_RE = re.compile(r"'\[[^\]]+\]([^']+)'!([A-Z]+[0-9]+)$")


def _cached_vs_computed(
    key: str, value, wbf, wbv, upper_to_sheet: dict[str, str]
) -> tuple[str, float, float] | None:
    """Map one solver output cell back to the workbook and return
    ``(label, cached, computed)`` when both are comparable numbers."""
    m = _SOLUTION_KEY_RE.match(key)
    if not m:
        return None
    sheet = upper_to_sheet.get(m.group(1).upper())
    if sheet is None:
        return None
    coord = m.group(2)
    stored_formula = wbf[sheet][coord].value
    if not (isinstance(stored_formula, str) and stored_formula.startswith("=")):
        return None
    cached = wbv[sheet][coord].value
    try:
        computed = value.value[0, 0]
        if hasattr(computed, "item") and getattr(computed, "size", 1) == 1:
            computed = computed.item()
    except Exception:  # noqa: BLE001
        return None
    if not isinstance(cached, (int, float)) or not isinstance(computed, (int, float)):
        return None
    return f"{sheet}!{coord}", float(cached), float(computed)


# ── code checks ───────────────────────────────────────────────────────────────


def analyze_code_tree(root: pathlib.Path) -> dict[str, Any]:
    """Deterministic findings for a code project tree."""
    files = sorted(p for p in root.rglob("*") if p.is_file())
    by_ext: dict[str, int] = {}
    total_lines = 0
    syntax_errors: list[str] = []
    invalid_json: list[str] = []
    todo_count = 0
    largest: list[tuple[int, str]] = []
    dep_files: list[str] = []
    has_tests = False

    for p in files:
        rel = str(p.relative_to(root))
        ext = p.suffix.lower() or "(none)"
        by_ext[ext] = by_ext.get(ext, 0) + 1
        largest.append((p.stat().st_size, rel))
        low = rel.lower()
        if any(seg in low for seg in ("test_", "_test.", "/tests/", "tests/")):
            has_tests = True
        if p.name in (
            "requirements.txt",
            "pyproject.toml",
            "package.json",
            "Cargo.toml",
            "go.mod",
            "Gemfile",
            "pom.xml",
        ):
            dep_files.append(rel)
        if ext in _CODE_EXTS or ext in (".json", ".md", ".txt", ".yaml", ".yml", ".toml"):
            try:
                text = p.read_text(encoding="utf-8", errors="replace")
            except Exception:  # noqa: BLE001
                continue
            total_lines += text.count("\n") + 1
            todo_count += len(re.findall(r"\b(TODO|FIXME|XXX)\b", text))
            if ext == ".py":
                try:
                    ast.parse(text)
                except SyntaxError as exc:
                    syntax_errors.append(f"{rel}: line {exc.lineno}: {exc.msg}")
            elif ext == ".json":
                try:
                    json.loads(text)
                except Exception as exc:  # noqa: BLE001
                    invalid_json.append(f"{rel}: {str(exc)[:120]}")

    largest.sort(reverse=True)
    return {
        "file_count": len(files),
        "files_by_extension": dict(sorted(by_ext.items(), key=lambda kv: -kv[1])),
        "total_lines": total_lines,
        "python_syntax_errors": syntax_errors[:_MAX_LISTED],
        "invalid_json": invalid_json[:_MAX_LISTED],
        "todo_marker_count": todo_count,
        "has_tests": has_tests,
        "dependency_files": dep_files,
        "largest_files": [f"{name} ({size} bytes)" for size, name in largest[:5]],
    }


# ── findings → issue list ─────────────────────────────────────────────────────


def summarize_issues(kind: str, findings: dict[str, Any]) -> list[str]:
    """Flatten findings into the plain-language issue list stored in
    ``checks_json`` and shown at the top of the report."""
    if kind == "xlsx":
        issues: list[str] = []
        for w in findings.get("workbooks", []):
            issues.extend(_workbook_issues(w))
        return issues
    return _code_issues(findings)


def _workbook_issues(w: dict[str, Any]) -> list[str]:
    name = w["file"]
    if w.get("load_error"):
        return [f"{name}: cannot be opened ({w['load_error']})"]
    issues: list[str] = []
    if w.get("error_cell_count"):
        issues.append(f"{name}: {w['error_cell_count']} error cell(s) saved in the file")
    if w.get("broken_reference_count"):
        issues.append(
            f"{name}: {w['broken_reference_count']} formula reference(s) "
            "to sheets that do not exist"
        )
    recalc = w.get("recalc") or {}
    if recalc.get("status") == "ok" and recalc.get("mismatches"):
        issues.append(
            f"{name}: {recalc['mismatches']} formula(s) recompute to a "
            "different value than the file has saved"
        )
    if recalc.get("status") == "engine_error":
        issues.append(f"{name}: independent recalculation failed ({recalc['reason']})")
    if w.get("volatile_count"):
        issues.append(
            f"{name}: {w['volatile_count']} volatile function call(s) "
            "(NOW/TODAY/RAND/OFFSET/INDIRECT) — results change between opens"
        )
    if w.get("external_link_count"):
        issues.append(f"{name}: {w['external_link_count']} external workbook reference(s)")
    issues.extend(_workbook_hygiene_issues(w, name))
    return issues


def _workbook_hygiene_issues(w: dict[str, Any], name: str) -> list[str]:
    issues: list[str] = []
    if w.get("hidden_sheets"):
        issues.append(f"{name}: hidden sheet(s): {', '.join(w['hidden_sheets'])}")
    if w.get("hidden_rows") or w.get("hidden_cols"):
        issues.append(
            f"{name}: hidden rows/columns ({w.get('hidden_rows', 0)} row(s), "
            f"{w.get('hidden_cols', 0)} column(s))"
        )
    if not w.get("has_readme_sheet"):
        issues.append(f"{name}: no README sheet documenting the workbook")
    return issues


def _code_issues(findings: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    if findings.get("python_syntax_errors"):
        issues.append(f"{len(findings['python_syntax_errors'])} Python file(s) do not parse")
    if findings.get("invalid_json"):
        issues.append(f"{len(findings['invalid_json'])} JSON file(s) are invalid")
    if not findings.get("has_tests"):
        issues.append("no tests found in the project")
    if findings.get("todo_marker_count"):
        issues.append(f"{findings['todo_marker_count']} TODO/FIXME marker(s) in the sources")
    return issues


# ── report assembly ───────────────────────────────────────────────────────────


def _xlsx_findings_md(findings: dict[str, Any]) -> str:
    lines: list[str] = []
    for w in findings.get("workbooks", []):
        lines.append(f"### {w['file']}")
        if w.get("load_error"):
            lines.append(f"- ❌ cannot be opened: {w['load_error']}")
            continue
        lines.append(
            f"- {len(w['sheets'])} sheet(s), {w['cells']:,} filled cells, "
            f"{w['formulas']:,} formulas"
        )
        recalc = w.get("recalc") or {}
        if recalc.get("status") == "ok":
            verdictish = (
                "all match saved values"
                if not recalc.get("mismatches")
                else f"{recalc['mismatches']} MISMATCH(ES)"
            )
            lines.append(
                f"- independent recalculation: {recalc['numeric_formulas_checked']:,} "
                f"numeric formulas checked — {verdictish}"
            )
            lines.extend(f"    - {ex}" for ex in recalc.get("mismatch_examples", []))
        else:
            lines.append(
                f"- independent recalculation: {recalc.get('status')} "
                f"({recalc.get('reason', '')})".rstrip("( )")
            )
        for label, key, listing in (
            ("error cells saved in file", "error_cell_count", "error_cells"),
            ("references to missing sheets", "broken_reference_count", "broken_references"),
            ("volatile function calls", "volatile_count", "volatile_functions"),
            ("external workbook references", "external_link_count", "external_links"),
        ):
            n = w.get(key, 0)
            lines.append(f"- {label}: {n if n else 'none'}")
            if n:
                lines.extend(f"    - {item}" for item in w.get(listing, []))
        lines.append(f"- merged ranges: {w['merged_ranges']}")
        lines.append(
            f"- hidden content: sheets={w['hidden_sheets'] or 'none'}, "
            f"rows={w['hidden_rows']}, cols={w['hidden_cols']}"
        )
        lines.append(f"- README sheet present: {'yes' if w['has_readme_sheet'] else 'no'}")
        lines.append("")
    return "\n".join(lines)


def _code_findings_md(findings: dict[str, Any]) -> str:
    lines = [
        f"- {findings['file_count']} file(s), ~{findings['total_lines']:,} lines",
        "- by type: "
        + ", ".join(f"{ext} ×{n}" for ext, n in findings["files_by_extension"].items()),
        f"- dependency manifests: {', '.join(findings['dependency_files']) or 'none found'}",
        f"- tests present: {'yes' if findings['has_tests'] else 'no'}",
        f"- TODO/FIXME markers: {findings['todo_marker_count']}",
    ]
    if findings["python_syntax_errors"]:
        lines.append("- Python files that DO NOT PARSE:")
        lines.extend(f"    - {e}" for e in findings["python_syntax_errors"])
    if findings["invalid_json"]:
        lines.append("- invalid JSON files:")
        lines.extend(f"    - {e}" for e in findings["invalid_json"])
    lines.append("- largest files: " + ", ".join(findings["largest_files"]))
    return "\n".join(lines)


def _narrative(db, cfg, kind: str, brief: str, focus: str, findings: dict, inputs_desc: str) -> str:
    """Plain-language review from the local model. Failure is reported in
    the returned text, never silently dropped."""
    from orivellum.capabilities.llm import llm_call

    res = llm_call(
        [
            {
                "role": "system",
                "content": (
                    "You are a meticulous reviewer in a project workbench. You are given "
                    "a project's files and machine-verified findings. Write a review "
                    "report in plain language with these sections: What this project is "
                    "and how it works; Problems found (lead with the machine findings — "
                    "never contradict them); What is missing or weak; Recommended next "
                    "steps (concrete, ordered). Be specific; cite file and sheet names. "
                    "Do not invent findings the machine checks did not show unless you "
                    "can point to the exact file content that supports them."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"PROJECT KIND: {kind}\n\nPROJECT BRIEF:\n{brief}\n\n"
                    f"REVIEW FOCUS:\n{focus or 'full review'}\n\n"
                    f"MACHINE FINDINGS (JSON):\n{json.dumps(findings, indent=1)[:6000]}\n\n"
                    f"PROJECT FILES:\n{inputs_desc[:8000]}"
                ),
            },
        ],
        cfg=cfg,
        db=db,
        purpose="workbench.analyze",
        temperature=0.3,
        max_tokens=4000,
        timeout=180,
    )
    if res.ok and res.text:
        return res.text.strip()
    return (
        f"_AI review unavailable ({res.error or 'empty reply'}) — the automated "
        "findings above are complete and did not depend on the model._"
    )


def build_report(
    title: str,
    kind: str,
    version_no: int,
    focus: str,
    findings: dict[str, Any],
    issues: list[str],
    narrative: str,
) -> str:
    from orivellum.version import code_version

    now = datetime.datetime.now(datetime.UTC).isoformat(timespec="seconds")
    header = [
        f"# Analysis Report — {title}",
        "",
        f"Generated {now} · analyzed v{version_no} · engine {code_version()}",
        f"Focus: {focus or 'full review'}",
        "",
        "## Verdict",
        "",
    ]
    if issues:
        header.append(f"**{len(issues)} issue(s) found:**")
        header.extend(f"- {i}" for i in issues)
    else:
        header.append("**No issues found by the automated checks.**")
    header += ["", "## Automated findings", ""]
    body = _xlsx_findings_md(findings) if kind == "xlsx" else _code_findings_md(findings)
    return "\n".join(header) + body + "\n\n## AI review\n\n" + narrative + "\n"


# ── main entry point ──────────────────────────────────────────────────────────


def run_analysis(db, cfg, project_id: str, focus: str = "") -> None:
    """Analyze the latest version and publish the report as a new version.
    Runs on the background executor; the caller must already hold the
    build claim, which is released here in ``finally``."""
    from orivellum.capabilities.workbench import (
        _describe_inputs,
        _publish_version,
        _snapshot,
        _xlsx_files,
        version_dir,
    )

    proj = db.get_wb_project(project_id)
    if not proj or proj["status"] != "active":
        db.update_wb_project(project_id, building=0)
        return
    db.update_wb_project(project_id, last_error=None)
    try:
        versions = db.list_wb_versions(project_id)
        if not versions:
            raise RuntimeError("nothing to analyze — the project has no versions yet")
        prev_no = versions[-1]["version_no"]
        src = version_dir(cfg, project_id, prev_no)
        if not src.is_dir():
            raise FileNotFoundError(f"version v{prev_no} has no files on disk")

        with tempfile.TemporaryDirectory() as tmp:
            out = pathlib.Path(tmp) / "out"
            shutil.copytree(src, out)
            out_report = out / REPORT_FILENAME
            kind = proj["kind"]

            if kind == "xlsx":
                findings: dict[str, Any] = {
                    "workbooks": [
                        analyze_workbook(p) for p in _xlsx_files(out) if p.name != REPORT_FILENAME
                    ]
                }
                if not findings["workbooks"]:
                    raise RuntimeError("no .xlsx files in the latest version to analyze")
            else:
                findings = analyze_code_tree(out)

            issues = summarize_issues(kind, findings)
            inputs_desc = _describe_inputs(kind, out)
            narrative = _narrative(db, cfg, kind, proj["brief"], focus, findings, inputs_desc)
            out_report.write_text(
                build_report(proj["title"], kind, prev_no, focus, findings, issues, narrative),
                encoding="utf-8",
            )

            files = _snapshot(out)
            checks = {
                "analysis": {
                    "analyzed_version": prev_no,
                    "issue_count": len(issues),
                    "issues": issues[:_MAX_LISTED],
                }
            }
            row = _publish_version(
                db,
                cfg,
                project_id,
                out,
                f"Analyze: {focus or 'full review'}",
                files,
                checks,
                note=f"{len(issues)} issue(s) found" if issues else "no issues found",
                verdict="analyzed",
            )
        logger.info("Workbench %s analyzed v%d -> v%d", project_id, prev_no, row["version_no"])
    except Exception as exc:  # noqa: BLE001
        logger.exception("Workbench analysis failed for %s", project_id)
        db.update_wb_project(project_id, last_error=str(exc)[:500])
    finally:
        db.update_wb_project(project_id, building=0)
