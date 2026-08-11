"""Project Workbench — agentic build / edit / repair with version history.

The user describes a project (an Excel workbook or a small code project) in
plain words. The local LLM writes ONE Python build script; the script runs
in the same locked-down sandbox the Document Workshop uses (no network,
scrubbed environment, resource caps). The script reads the previous
version's files from ``inputs/`` and writes the COMPLETE new state of the
project into ``out/``.

Every successful, verified iteration becomes an immutable version under
``data/workbench/{project}/v{n}/``. Verification before a version is
accepted:

- **xlsx** projects: at least one ``.xlsx`` output; every workbook must
  load cleanly in openpyxl in BOTH value and formula view.
- **code** projects: at least one output file; every ``.py`` must parse
  (``ast.parse``) and every ``.json`` must parse.

A failed build or failed verification stores the error on the project row
and creates NO version — the last good version remains the truth.

Completing a project zips every version plus a ``manifest.json`` with
SHA-256 hashes into ``data/workbench/archives/`` and marks the project
archived. Archived projects are read-only. Archival re-hashes every file on
disk and refuses to produce an archive that does not match the recorded
version manifests.

Concurrency: all mutating operations (build, revert, archive, delete) must
first claim the project via ``db.claim_wb_build()`` — an atomic
conditional UPDATE on ``building`` — so two requests can never build from
the same predecessor or archive mid-build.

Threat model note: the sandbox is the same one the Document Workshop uses —
scrubbed environment, no inherited network config, POSIX resource caps. It
is a guard against accidental damage from generated code on the operator's
own machine, not a hostile-code security boundary; Orivellum is a
single-operator local system and the operator reviews what they run.
"""

from __future__ import annotations

import ast
import hashlib
import json
import logging
import pathlib
import secrets
import shutil
import subprocess
import sys
import tempfile
import zipfile
from typing import Any

logger = logging.getLogger(__name__)

KINDS = ("xlsx", "code")
_SCRIPT_TIMEOUT_S = 120
_MAX_FIX_RETRIES = 2
_MAX_OUTPUT_FILES = 200
_MAX_OUTPUT_BYTES = 25 * 1024 * 1024  # 25 MB per version
_CODE_CONTEXT_PER_FILE = 4000  # chars of each source file shown to the LLM
_CODE_CONTEXT_TOTAL = 24000


# ── Paths ─────────────────────────────────────────────────────────────────────


def _workbench_root(cfg) -> pathlib.Path:
    return pathlib.Path(getattr(cfg, "data_dir", "data")) / "workbench"


def project_dir(cfg, project_id: str) -> pathlib.Path:
    return _workbench_root(cfg) / project_id


def version_dir(cfg, project_id: str, version_no: int) -> pathlib.Path:
    return project_dir(cfg, project_id) / f"v{version_no}"


def archives_dir(cfg) -> pathlib.Path:
    return _workbench_root(cfg) / "archives"


# ── File helpers ──────────────────────────────────────────────────────────────


def _sha256(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _snapshot(dir_path: pathlib.Path) -> list[dict]:
    """Relative name + size + sha256 for every file under *dir_path*.

    Symlinks are rejected outright: a link smuggled into the output could
    make the published version serve bytes from outside the sandbox."""
    out = []
    for p in sorted(dir_path.rglob("*")):
        if p.is_symlink():
            raise ValueError(f"symlink in output is not allowed: {p.relative_to(dir_path)}")
        if p.is_file():
            out.append(
                {
                    "name": str(p.relative_to(dir_path)),
                    "size": p.stat().st_size,
                    "sha256": _sha256(p),
                }
            )
    return out


# ── Input description (what the LLM sees about the current version) ─────────


def _describe_inputs(kind: str, inputs: pathlib.Path) -> str:
    files = sorted(p for p in inputs.rglob("*") if p.is_file())
    if not files:
        return "(no existing files — this is the first build)"
    lines: list[str] = []
    if kind == "xlsx":
        from openpyxl import load_workbook

        for p in files:
            rel = p.relative_to(inputs)
            if p.suffix.lower() != ".xlsx":
                lines.append(f"- {rel} ({p.stat().st_size} bytes)")
                continue
            try:
                wb = load_workbook(p, read_only=True)
                for ws in wb.worksheets:
                    header = []
                    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                        header = [str(v) for v in row if v is not None][:12]
                    lines.append(
                        f"- {rel} :: sheet '{ws.title}' dims={ws.calculate_dimension()}"
                        f" headers={header}"
                    )
                wb.close()
            except Exception as exc:  # noqa: BLE001
                lines.append(f"- {rel} (unreadable: {exc})")
    else:
        budget = _CODE_CONTEXT_TOTAL
        for p in files:
            rel = p.relative_to(inputs)
            try:
                text = p.read_text(encoding="utf-8", errors="replace")
            except Exception:  # noqa: BLE001
                lines.append(f"- {rel} (binary, {p.stat().st_size} bytes)")
                continue
            take = min(_CODE_CONTEXT_PER_FILE, max(0, budget))
            budget -= take
            snippet = text[:take]
            suffix = "" if len(text) <= take else f"\n... [truncated, {len(text)} chars total]"
            lines.append(f"### {rel}\n```\n{snippet}{suffix}\n```")
    return "\n".join(lines)


# ── LLM script generation ─────────────────────────────────────────────────────

_XLSX_RULES = """Allowed imports: openpyxl and the Python standard library ONLY
(json, csv, math, datetime, statistics, ...). NO network, NO pip, NO
subprocesses or shelling out — the sandbox blocks launching processes.
Excel rules:
- Formulas are strings starting with '=' assigned to cells; use real cell
  references, never hardcoded results of other cells.
- Give every sheet a header row with bold font where it makes sense.
- Set sensible number formats (currency, dates, percentages).
- Never invent data the user did not ask for; leave input areas blank."""

_CODE_RULES = """Allowed imports in the BUILD SCRIPT: Python standard library ONLY.
The script's only job is to write the project's source files with open()/write.
The project files themselves may target any language the user asked for."""


def _system_prompt(kind: str) -> str:
    return (
        "You are the build engine of a project workbench. You write ONE complete "
        "Python script and nothing else — no prose, no markdown fences.\n"
        "The script runs in a sandbox with this contract:\n"
        "- Existing project files are in ./inputs/ (empty on first build).\n"
        "- Write the COMPLETE new state of the project into ./out/ — every file "
        "the project should contain after this change, not only changed files. "
        "Copy unchanged files from inputs/ to out/ (shutil.copy2 or re-create).\n"
        "- Print a one-line summary of what was built at the end.\n"
        + (_XLSX_RULES if kind == "xlsx" else _CODE_RULES)
    )


def _user_prompt(kind: str, brief: str, instruction: str, inputs_desc: str) -> str:
    return (
        f"PROJECT BRIEF:\n{brief}\n\n"
        f"CURRENT PROJECT FILES:\n{inputs_desc}\n\n"
        f"INSTRUCTION FOR THIS ITERATION:\n{instruction}\n\n"
        "Write the build script now."
    )


# ── Sandbox execution (reuses the Workshop's hardened sandbox pieces) ────────


def _run_build_script(script: str, workdir: pathlib.Path, cfg, db, request: str) -> dict:
    """Run *script* sandboxed with ./inputs and ./out available under
    *workdir*. Retries with LLM correction, same policy as the Workshop."""
    from orivellum.capabilities.llm import llm_call
    from orivellum.capabilities.workshop import (
        _SANDBOX_RUNNER,
        _clean_script,
        _sandbox_env,
        _sandbox_preexec,
    )

    current = script
    for attempt in range(_MAX_FIX_RETRIES + 1):
        script_path = workdir / "build_project.py"
        script_path.write_text(current, encoding="utf-8")
        runner_path = workdir / "_sandbox_runner.py"
        runner_path.write_text(_SANDBOX_RUNNER, encoding="utf-8")
        try:
            result = subprocess.run(
                [sys.executable, "-I", str(runner_path), str(script_path)],
                capture_output=True,
                text=True,
                timeout=_SCRIPT_TIMEOUT_S,
                cwd=str(workdir),
                env=_sandbox_env(str(workdir)),
                preexec_fn=_sandbox_preexec if sys.platform != "win32" else None,
            )
        except subprocess.TimeoutExpired:
            return {"ok": False, "error": f"Build script timed out ({_SCRIPT_TIMEOUT_S}s)"}
        stdout, stderr = result.stdout[:3000], result.stderr[:3000]
        if result.returncode == 0:
            return {"ok": True, "stdout": stdout, "script": current}
        if attempt >= _MAX_FIX_RETRIES:
            return {
                "ok": False,
                "error": f"Build failed after {attempt + 1} attempt(s):\n{stderr[-1000:]}",
                "stdout": stdout,
            }
        logger.warning("Workbench build attempt %d failed: %s", attempt + 1, stderr[:400])
        fix = llm_call(
            [
                {
                    "role": "system",
                    "content": "You are a Python debugging expert. A project build script failed. "
                    "Fix ONLY the errors shown. Return ONLY the corrected raw script.",
                },
                {
                    "role": "user",
                    "content": f"Request:\n{request[:500]}\n\nScript:\n```python\n{current}\n```\n\n"
                    f"Error:\n{stderr[-2000:]}",
                },
            ],
            cfg=cfg,
            db=db,
            purpose="workbench.fix",
            temperature=0.1,
            max_tokens=6000,
            timeout=90,
        )
        if not (fix.ok and fix.text):
            return {"ok": False, "error": f"Fix generation failed: {fix.error}"}
        current = _clean_script(fix.text)
    return {"ok": False, "error": "Max retries exceeded"}


# ── Code project tests ────────────────────────────────────────────────────────

_TEST_FILE = "project_tests.py"
_TEST_OUTPUT_CAP = 4000

_TESTGEN_SYSTEM = (
    "You write ONE Python test file and nothing else — no prose, no markdown "
    "fences. Rules:\n"
    "- Standard library ONLY, built on unittest. No pip, no network, no "
    "subprocess. Import only the project's own modules plus: unittest, os, "
    "pathlib, re, json, math, io, csv, string, datetime, collections, "
    "itertools, functools, textwrap, html, unicodedata, runpy, tempfile, "
    "shutil, statistics, decimal, fractions, random, time, typing. Never "
    "import sys, ctypes, gc, inspect, or __main__; never touch sys.modules, "
    "monkeypatch modules, or call exit()/os._exit().\n"
    "- The project's files sit in the SAME directory as the test file; import "
    "modules by file name (e.g. `import main` for main.py).\n"
    f"- Never import {_TEST_FILE} itself.\n"
    "- If a file is a script with top-level side effects rather than importable "
    "functions, run it with runpy.run_path and assert on its observable "
    "output or files.\n"
    "- The project may contain files in ANY language. Test non-Python files by "
    "reading them: parse JSON with json, check HTML with html.parser, and "
    "assert the requested functions, markup, or content exist (re / string "
    "checks). Every claim must be verified against the actual files — never "
    "write a test that passes without reading or running the project.\n"
    "- Write a handful of meaningful, deterministic assertions covering the "
    "behavior that was requested — no sleeps, no randomness, no network.\n"
    '- End with `if __name__ == "__main__": unittest.main()` so failures exit '
    "non-zero."
)


def _generate_tests(cfg, db, proj: dict, instruction: str, out_dir: pathlib.Path) -> str:
    from orivellum.capabilities.llm import llm_call
    from orivellum.capabilities.workshop import _clean_script

    gen = llm_call(
        [
            {"role": "system", "content": _TESTGEN_SYSTEM},
            {
                "role": "user",
                "content": f"PROJECT BRIEF:\n{proj['brief']}\n\n"
                f"WHAT THIS ITERATION WAS ASKED TO DO:\n{instruction}\n\n"
                f"PROJECT FILES:\n{_describe_inputs(proj['kind'], out_dir)}\n\n"
                "Write the test file now.",
            },
        ],
        cfg=cfg,
        db=db,
        purpose="workbench.testgen",
        temperature=0.1,
        max_tokens=4000,
        timeout=120,
    )
    if not (gen.ok and gen.text):
        raise RuntimeError(f"test generation failed: {gen.error or 'empty reply'}")
    code = _clean_script(gen.text)
    try:
        ast.parse(code)
    except SyntaxError as exc:
        raise RuntimeError(f"generated test file is not valid Python: {exc.msg}") from exc
    return code


_TEST_SUPERVISOR = '''\
"""Trusted test supervisor. Never imports or executes test/project code —
it screens the generated test file AND every project file statically (AST),
then runs the suite in a SEPARATE sandboxed process. Only this process ever
sees the auth token (read from stdin), so test or project code cannot forge
the token-authenticated result, and the screens reject every known way to
tamper with the harness in-process (module patching, sys.modules poisoning,
introspection escapes, or exiting 0 without running the suite)."""

import ast
import json
import os
import subprocess
import sys

# Test files may import unittest + a small stdlib toolbox + project modules.
_TEST_IMPORT_ALLOW = {
    "unittest", "os", "pathlib", "re", "json", "math", "io", "csv", "string",
    "datetime", "collections", "itertools", "functools", "textwrap", "html",
    "unicodedata", "runpy", "tempfile", "shutil", "statistics", "decimal",
    "fractions", "random", "time", "typing",
}
_BANNED_CALLS = (
    "setattr", "delattr", "globals", "vars", "exec", "eval", "compile",
    "exit", "quit", "__import__",
)


def _patch_and_exit_problems(tree):
    """Constructs that defeat certification in ANY file: patching an imported
    module's attributes, touching sys.modules, or hard process exits."""
    problems, imported = [], set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for a in node.names:
                imported.add(a.asname or a.name.split(".")[0])
        elif isinstance(node, (ast.Assign, ast.AugAssign)):
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            for t in targets:
                root = t
                while isinstance(root, ast.Attribute):
                    root = root.value
                if (
                    isinstance(t, ast.Attribute)
                    and isinstance(root, ast.Name)
                    and root.id in imported
                ):
                    problems.append("patches an imported module")
        elif isinstance(node, ast.Attribute):
            if node.attr == "_exit":
                problems.append("uses _exit")
            elif node.attr == "modules":
                problems.append("touches sys.modules")
        elif isinstance(node, ast.Raise):
            exc = node.exc
            name = getattr(exc, "id", None) or getattr(getattr(exc, "func", None), "id", None)
            if name in ("SystemExit", "BaseException", "KeyboardInterrupt"):
                problems.append("raises " + name)
    return problems


def _screen_test(tree, allowed):
    """Whitelist imports; reject tamper constructs; count real tests."""
    problems, tests, asserts = list(_patch_and_exit_problems(tree)), 0, 0
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for a in node.names:
                if a.name.split(".")[0] not in allowed:
                    problems.append("imports " + a.name)
        elif isinstance(node, ast.ImportFrom):
            if (node.module or "").split(".")[0] not in allowed:
                problems.append("imports " + (node.module or "?"))
        elif isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            if node.func.id in _BANNED_CALLS:
                problems.append("calls " + node.func.id)
        elif isinstance(node, ast.ClassDef):
            bases = [
                b.attr if isinstance(b, ast.Attribute) else getattr(b, "id", "")
                for b in node.bases
            ]
            if "TestCase" in bases:
                for item in node.body:
                    if isinstance(item, ast.FunctionDef) and item.name.startswith("test"):
                        tests += 1
        if isinstance(node, ast.Attribute) and node.attr.startswith("assert"):
            asserts += 1
    return problems, tests, asserts


def _screen_dir(test_path):
    """Screen every project .py the suite could import for harness tampering."""
    problems = []
    folder = os.path.dirname(os.path.abspath(test_path))
    for base, _dirs, files in os.walk(folder):
        for name in files:
            path = os.path.join(base, name)
            if not name.endswith(".py") or os.path.abspath(path) == os.path.abspath(test_path):
                continue
            try:
                with open(path, encoding="utf-8") as f:
                    tree = ast.parse(f.read())
            except (OSError, SyntaxError):
                continue  # unparseable files cannot be imported by the tests
            for prob in _patch_and_exit_problems(tree):
                problems.append(name + " " + prob)
    return problems


def main():
    test_path, result_path, runner_path = sys.argv[1], sys.argv[2], sys.argv[3]
    token = sys.stdin.readline().strip()
    report = {"token": token, "tests_run": 0, "ok": False}
    try:
        with open(test_path, encoding="utf-8") as f:
            tree = ast.parse(f.read())
        folder = os.path.dirname(os.path.abspath(test_path))
        stems = {n[:-3] for n in os.listdir(folder) if n.endswith(".py")}
        problems, tests, asserts = _screen_test(tree, _TEST_IMPORT_ALLOW | stems)
        problems += _screen_dir(test_path)
        if problems:
            report["error"] = "failed the safety screen: " + "; ".join(sorted(set(problems)))
        elif tests < 1 or asserts < 1:
            report["error"] = "test file defines no real tests with assertions"
        else:
            proc = subprocess.run(
                [sys.executable, "-I", runner_path, test_path, str(tests)],
                capture_output=True,
                text=True,
            )
            sys.stdout.write(proc.stdout or "")
            sys.stderr.write(proc.stderr or "")
            report["ok"] = proc.returncode == 0
            report["tests_run"] = tests if proc.returncode == 0 else 0
    except Exception as exc:  # explicit in the report, never a silent pass
        report["error"] = str(exc) or type(exc).__name__
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(report, f)
    sys.exit(0 if report["ok"] else 1)


main()
'''


def _test_runner_source() -> str:
    """The Workshop sandbox runner (network denial, scrubbed env; rlimits are
    inherited from the supervisor) with its execution tail swapped for a
    unittest harness. This process is UNTRUSTED — it never sees the token or
    result path; its only signal back to the supervisor is the exit code,
    which requires unittest to have genuinely run at least the statically
    counted number of tests and succeeded.

    Lessons baked in: import the real `unittest` before the project dir joins
    sys.path (a project unittest.py would shadow it), and reset sys.argv
    (unittest parses argv). `python -I` + importlib never adds the script dir
    to sys.path — insert it explicitly so sibling imports work."""
    from orivellum.capabilities.workshop import _SANDBOX_RUNNER

    marker = 'runpy.run_path(sys.argv[1], run_name="__main__")'
    if marker not in _SANDBOX_RUNNER:  # runner contract changed — fail loudly
        raise RuntimeError("sandbox runner no longer matches the test-runner injection point")
    harness = """\
import importlib.util
import os
import unittest

_target, _expected = sys.argv[1], int(sys.argv[2])
sys.argv = [_target]
sys.path.insert(0, os.path.dirname(os.path.abspath(_target)))
_spec = importlib.util.spec_from_file_location("project_tests_module", _target)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)
_result = unittest.TextTestRunner(stream=sys.stderr, verbosity=1).run(
    unittest.defaultTestLoader.loadTestsFromModule(_mod)
)
sys.exit(0 if (_result.wasSuccessful() and _result.testsRun >= _expected >= 1) else 1)
"""
    return _SANDBOX_RUNNER.replace(marker, harness)


def _run_project_tests(test_code: str, out_dir: pathlib.Path, workdir: pathlib.Path) -> dict:
    """Execute the generated test file in the build sandbox against an
    ISOLATED COPY of the built project — a test that mutates project files
    can only mutate the throwaway copy, so a pass always certifies the exact
    bytes that get published. A pass requires the trusted supervisor's
    token-authenticated result file: the supervisor screens the test file
    statically, runs it in a separate untrusted process, and is the only
    process that ever sees the token — exit codes and printed output alone
    can never certify a pass, and test code cannot forge the result."""
    from orivellum.capabilities.workshop import _sandbox_env, _sandbox_preexec

    test_dir = workdir / "_testrun"
    if test_dir.exists():
        shutil.rmtree(test_dir)
    shutil.copytree(out_dir, test_dir)
    test_path = test_dir / _TEST_FILE
    test_path.write_text(test_code, encoding="utf-8")
    runner_path = workdir / "_test_runner.py"
    runner_path.write_text(_test_runner_source(), encoding="utf-8")
    supervisor_path = workdir / "_test_supervisor.py"
    supervisor_path.write_text(_TEST_SUPERVISOR, encoding="utf-8")
    token = secrets.token_hex(16)
    result_path = workdir / "_test_result.json"
    result_path.unlink(missing_ok=True)
    try:
        result = subprocess.run(
            [
                sys.executable,
                "-I",
                str(supervisor_path),
                str(test_path),
                str(result_path),
                str(runner_path),
            ],
            input=token + "\n",
            capture_output=True,
            text=True,
            timeout=_SCRIPT_TIMEOUT_S,
            cwd=str(test_dir),
            env=_sandbox_env(str(workdir)),
            preexec_fn=_sandbox_preexec if sys.platform != "win32" else None,
        )
    except subprocess.TimeoutExpired:
        return {
            "passed": False,
            "output": f"tests timed out ({_SCRIPT_TIMEOUT_S}s)",
            "tests_run": 0,
        }
    output = ((result.stdout or "") + "\n" + (result.stderr or "")).strip()
    report = None
    try:
        data = json.loads(result_path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and data.get("token") == token:
            report = data
    except (OSError, ValueError):
        pass
    if report is None:
        return {
            "passed": False,
            "output": ("no trusted test result — the suite did not complete\n" + output)[
                -_TEST_OUTPUT_CAP:
            ],
            "tests_run": 0,
        }
    if report.get("error"):
        output = str(report["error"]) + "\n" + output
    tests_run = int(report.get("tests_run") or 0)
    passed = result.returncode == 0 and report.get("ok") is True and tests_run >= 1
    return {"passed": passed, "output": output[-_TEST_OUTPUT_CAP:], "tests_run": tests_run}


def _fix_script_for_tests(db, cfg, instruction, script, test_code, test_output) -> str:
    """Ask the LLM to repair the build script so its output passes the tests."""
    from orivellum.capabilities.llm import llm_call
    from orivellum.capabilities.workshop import _clean_script

    fix = llm_call(
        [
            {
                "role": "system",
                "content": "You are a Python debugging expert. A workbench build script "
                "produced project files that FAIL their tests. Fix the BUILD SCRIPT so "
                "the project it writes passes the tests. Do not weaken or game the "
                "tests. Return ONLY the corrected raw build script.",
            },
            {
                "role": "user",
                "content": f"Request:\n{instruction[:500]}\n\n"
                f"Build script:\n```python\n{script}\n```\n\n"
                f"Test file:\n```python\n{test_code}\n```\n\n"
                f"Test output:\n{test_output[-2000:]}",
            },
        ],
        cfg=cfg,
        db=db,
        purpose="workbench.testfix",
        temperature=0.1,
        max_tokens=8000,
        timeout=120,
    )
    if not (fix.ok and fix.text):
        raise RuntimeError(f"test-repair generation failed: {fix.error or 'empty reply'}")
    return _clean_script(fix.text)


def _rebuild_for_tests(db, cfg, kind, instruction, work, out, script) -> str:
    """Re-run a repaired build script into a fresh out/ and re-verify.
    Returns the script actually executed (may differ after syntax fixes)."""
    shutil.rmtree(out)
    out.mkdir()
    run = _run_build_script(script, work, cfg, db, instruction)
    if not run["ok"]:
        raise RuntimeError(f"rebuild after test failure failed: {run['error']}")
    ok, checks = _verify_output(kind, out)
    if not ok:
        raise RuntimeError(
            "verification failed after test repair: "
            + "; ".join(checks.get("problems") or [checks.get("error", "unknown")])
        )
    return run["script"]


def _pretest_code_build(db, cfg, proj, instruction, work, out, script) -> dict | None:
    """For non-xlsx builds: structural check first, then the generated test
    suite. Returns the tests record, or None for xlsx (proof harness covers
    those)."""
    if proj["kind"] == "xlsx":
        return None
    ok, checks = _verify_output(proj["kind"], out)
    if not ok:
        raise RuntimeError(
            "verification failed: "
            + "; ".join(checks.get("problems") or [checks.get("error", "unknown")])
        )
    return _test_code_project(db, cfg, proj, instruction, work, out, script)


def _test_code_project(db, cfg, proj, instruction, work, out, script) -> dict:
    """Generate a test file for the built code project, run it sandboxed, and
    feed failures back into the LLM repair loop (rebuild → re-verify → re-run
    the SAME tests). Returns the record for checks_json['tests']; raises when
    the tests still fail after retries — the version is then never published.

    Every code project gets tests, whatever language its files are in —
    Python modules are imported and exercised; non-Python files are verified
    by reading/parsing them. There is no untested path to a good verdict."""
    test_code = _generate_tests(cfg, db, proj, instruction, out)
    for attempt in range(_MAX_FIX_RETRIES + 1):
        res = _run_project_tests(test_code, out, work)
        if res["passed"]:
            # ship the passing test file with the version — the project bytes
            # it certified are untouched (tests ran on an isolated copy)
            (out / _TEST_FILE).write_text(test_code, encoding="utf-8")
            return {
                "passed": True,
                "output": res["output"],
                "attempts": attempt + 1,
                "tests_run": res["tests_run"],
                "test_file": _TEST_FILE,
            }
        if attempt >= _MAX_FIX_RETRIES:
            raise RuntimeError(
                f"project tests failed after {attempt + 1} run(s):\n{res['output'][-800:]}"
            )
        logger.warning("Workbench tests attempt %d failed: %s", attempt + 1, res["output"][:400])
        script = _fix_script_for_tests(db, cfg, instruction, script, test_code, res["output"])
        script = _rebuild_for_tests(db, cfg, proj["kind"], instruction, work, out, script)
    raise RuntimeError("Max test retries exceeded")  # pragma: no cover


# ── Verification ──────────────────────────────────────────────────────────────


def _verify_output(kind: str, out_dir: pathlib.Path, prove: bool = False) -> tuple[bool, dict]:
    """Structural verification, plus (for xlsx) the six-gate proof harness.

    ``prove=True`` runs every workbook through the Orivellum Runner gate
    suite (recalculation, value match, error scan, OOXML order, containment,
    clean load) and — when all gates pass — atomically promotes the repaired
    file, so cached values are real numbers, not blanks. Builds prove;
    imports don't (an imported v1 stays verbatim), but they may be proven
    without promotion by the caller.
    """
    files = [p for p in out_dir.rglob("*") if p.is_file()]
    checks: dict[str, Any] = {"file_count": len(files)}
    if not files:
        checks["error"] = "build produced no files"
        return False, checks
    if len(files) > _MAX_OUTPUT_FILES:
        checks["error"] = f"too many output files ({len(files)})"
        return False, checks
    total = sum(p.stat().st_size for p in files)
    checks["total_bytes"] = total
    if total > _MAX_OUTPUT_BYTES:
        checks["error"] = f"output too large ({total} bytes)"
        return False, checks

    problems: list[str] = []
    if kind == "xlsx":
        from openpyxl import load_workbook

        workbooks = [p for p in files if p.suffix.lower() == ".xlsx"]
        checks["workbooks"] = len(workbooks)
        if not workbooks:
            problems.append("no .xlsx file produced")
        for p in workbooks:
            rel = str(p.relative_to(out_dir))
            for data_only in (False, True):
                try:
                    load_workbook(p, data_only=data_only).close()
                except Exception as exc:  # noqa: BLE001
                    problems.append(
                        f"{rel} fails to load ({'values' if data_only else 'formulas'}): {exc}"
                    )
                    break
        if prove and workbooks and not problems:
            from orivellum.capabilities.workbench_proof import prove_outputs

            proof = prove_outputs(out_dir, workbooks)
            checks["proof"] = proof
            if proof["verdict"] == "failed":
                for rel, r in proof["workbooks"].items():
                    if r["verdict"] == "failed":
                        why = "; ".join(r.get("problems") or ["gates failed"])[:300]
                        problems.append(f"{rel} failed proof: {why}")
    else:
        for p in files:
            rel = str(p.relative_to(out_dir))
            if p.suffix == ".py":
                try:
                    ast.parse(p.read_text(encoding="utf-8", errors="replace"))
                except SyntaxError as exc:
                    problems.append(f"{rel}: syntax error line {exc.lineno}: {exc.msg}")
            elif p.suffix == ".json":
                try:
                    json.loads(p.read_text(encoding="utf-8", errors="replace"))
                except Exception as exc:  # noqa: BLE001
                    problems.append(f"{rel}: invalid JSON: {exc}")
    checks["problems"] = problems
    return not problems, checks


def _proof_verdict(checks: dict | None) -> str:
    """Version verdict from a checks dict: 'proven' / 'unverified' when the
    six-gate harness ran; 'tested' when the project's own generated tests
    passed; plain 'verified' otherwise."""
    proof = (checks or {}).get("proof")
    if proof:
        return {"proven": "proven", "unverified": "unverified"}.get(proof["verdict"], "verified")
    tests = (checks or {}).get("tests")
    if tests and tests.get("passed"):
        return "tested"
    return "verified"


class UnprovenError(ValueError):
    """Raised when archiving is refused because the latest version is not
    fully proven. Callers may retry with allow_unproven=True after an
    explicit user confirmation."""


def _xlsx_shas(version: dict) -> dict[str, str]:
    return {
        f["name"]: f["sha256"]
        for f in json.loads(version.get("files_json") or "[]")
        if f["name"].lower().endswith(".xlsx")
    }


def latest_proof_status(proj: dict, versions: list[dict]) -> tuple[str, str]:
    """(status, detail) for the LATEST version of an xlsx project.

    status: 'proven' | 'provable' | 'failed' | 'unverified' | 'unproven' | 'n/a'

    A latest version without its own proof (analysis reports and reverts
    copy files forward without re-gating) INHERITS the most recent earlier
    proof whose workbook set is byte-identical (same names, same sha256) —
    the proof certifies bytes, and identical bytes carry it.
    """
    if proj["kind"] != "xlsx" or not versions:
        return "n/a", ""
    latest = versions[-1]
    proof = json.loads(latest["checks_json"] or "{}").get("proof")
    if not proof:
        latest_shas = _xlsx_shas(latest)
        if latest_shas:
            for v in reversed(versions[:-1]):
                prior = json.loads(v["checks_json"] or "{}").get("proof")
                if prior and _xlsx_shas(v) == latest_shas:
                    proof = prior
                    break
    if not proof:
        return "unproven", (f"v{latest['version_no']} was never run through the proof gates")
    if proof["verdict"] == "proven":
        return "proven", ""
    detail = "; ".join(
        f"{name}: {'; '.join((r.get('problems') or ['gates failed'])[:2])}"
        for name, r in proof.get("workbooks", {}).items()
        if r["verdict"] != "proven"
    )[:400]
    return proof["verdict"], detail or proof.get("error", "")


# ── Main entry points ─────────────────────────────────────────────────────────


def run_build(db, cfg, project_id: str, instruction: str) -> None:
    """Build the next version of a project. Runs on the background executor.
    The caller must already hold the build claim (``db.claim_wb_build``);
    this function releases it in ``finally``. All outcomes (including
    failure) land on the project row."""
    from orivellum.capabilities.llm import llm_call
    from orivellum.capabilities.workshop import _clean_script

    proj = db.get_wb_project(project_id)
    if not proj or proj["status"] != "active":
        db.update_wb_project(project_id, building=0)
        return
    db.update_wb_project(project_id, last_error=None)
    try:
        versions = db.list_wb_versions(project_id)
        prev_no = versions[-1]["version_no"] if versions else 0

        with tempfile.TemporaryDirectory() as tmp:
            work = pathlib.Path(tmp)
            inputs = work / "inputs"
            out = work / "out"
            inputs.mkdir()
            out.mkdir()
            if prev_no:
                src = version_dir(cfg, project_id, prev_no)
                if src.is_dir():
                    shutil.copytree(src, inputs, dirs_exist_ok=True)

            desc = _describe_inputs(proj["kind"], inputs)
            gen = llm_call(
                [
                    {"role": "system", "content": _system_prompt(proj["kind"])},
                    {
                        "role": "user",
                        "content": _user_prompt(proj["kind"], proj["brief"], instruction, desc),
                    },
                ],
                cfg=cfg,
                db=db,
                purpose="workbench.build",
                temperature=0.2,
                max_tokens=8000,
                timeout=180,
            )
            if not (gen.ok and gen.text):
                raise RuntimeError(f"model call failed: {gen.error or 'empty reply'}")
            script = _clean_script(gen.text)

            run = _run_build_script(script, work, cfg, db, instruction)
            if not run["ok"]:
                raise RuntimeError(run["error"])

            # Code projects: generate + run the project's own tests in the
            # sandbox, feeding failures back to the LLM repair loop. Raises
            # when tests never pass — the version is never published.
            tests = _pretest_code_build(db, cfg, proj, instruction, work, out, run["script"])

            ok, checks = _verify_output(proj["kind"], out, prove=True)
            if not ok:
                raise RuntimeError(
                    "verification failed: "
                    + "; ".join(checks.get("problems") or [checks.get("error", "unknown")])
                )
            if tests is not None:
                checks["tests"] = tests

            # Accept: publish files FIRST (staging dir + atomic rename), and
            # only then commit the version row — a crash can leave an unused
            # staging dir behind, but never a version row without files.
            # Snapshot AFTER proving: promoted (cache-repaired) workbooks
            # must be the ones whose hashes the version records.
            files = _snapshot(out)
            note = (run.get("stdout") or "").strip()[:500]
            row = _publish_version(
                db,
                cfg,
                project_id,
                out,
                instruction,
                files,
                checks,
                note,
                verdict=_proof_verdict(checks),
            )
        logger.info("Workbench %s built v%d", project_id, row["version_no"])
    except Exception as exc:  # noqa: BLE001
        # Surfaced to the user on the project row — never swallowed.
        logger.exception("Workbench build failed for %s", project_id)
        db.update_wb_project(project_id, last_error=str(exc)[:500])
    finally:
        db.update_wb_project(project_id, building=0)


def _publish_version(
    db,
    cfg,
    project_id: str,
    src_dir: pathlib.Path,
    instruction: str,
    files: list[dict],
    checks: dict | None,
    note: str = "",
    verdict: str = "verified",
) -> dict:
    """Copy *src_dir* into the project as the next version: stage, insert the
    row, atomically rename staging → v{n}. If the rename fails the row is
    deleted again, so a verified row always has its files."""
    import uuid as _uuid_mod

    pdir = project_dir(cfg, project_id)
    pdir.mkdir(parents=True, exist_ok=True)
    staging = pdir / f".staging-{_uuid_mod.uuid4().hex}"
    try:
        shutil.copytree(src_dir, staging)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    try:
        row = db.create_wb_version(
            project_id, instruction, files, checks=checks, verdict=verdict, note=note
        )
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    dest = version_dir(cfg, project_id, row["version_no"])
    try:
        staging.replace(dest)
    except Exception:
        db.delete_wb_version(row["id"])
        shutil.rmtree(staging, ignore_errors=True)
        raise
    return row


_IMPORT_JUNK_PREFIXES = ("__MACOSX/",)
_IMPORT_JUNK_NAMES = {".DS_Store", "Thumbs.db"}
_XLSX_MAX_MEMBERS = 10_000
_XLSX_MAX_UNCOMPRESSED = 300 * 1024 * 1024


def check_xlsx_zip_safety(path: pathlib.Path) -> str | None:
    """Declared-size guard against OOXML decompression bombs: an .xlsx is
    itself a zip, so a small file can expand enormously when opened.
    Returns an error string, or ``None`` when safe (or not a zip at all —
    openpyxl then reports its own load error)."""
    try:
        with zipfile.ZipFile(path) as zf:
            infos = zf.infolist()
            if len(infos) > _XLSX_MAX_MEMBERS:
                return f"workbook zip has too many members ({len(infos)})"
            total = sum(i.file_size for i in infos)
            if total > _XLSX_MAX_UNCOMPRESSED:
                mb = total // (1024 * 1024)
                return f"workbook expands to {mb} MB uncompressed — refusing to open it"
    except (zipfile.BadZipFile, OSError):
        return None
    return None


def _detect_kind(names: list[str]) -> str:
    """xlsx if the upload contains a workbook and no source-code files,
    otherwise code."""
    from orivellum.capabilities.workbench_analyze import _CODE_EXTS

    exts = {pathlib.PurePosixPath(n).suffix.lower() for n in names}
    if ".xlsx" in exts and not (exts & _CODE_EXTS):
        return "xlsx"
    return "code"


def _safe_zip_member(info: zipfile.ZipInfo) -> pathlib.PurePosixPath | None:
    """Validate one zip entry. Returns its relative path, ``None`` for
    junk/directories, or raises for traversal attempts and symlinks."""
    raw = info.filename
    if info.is_dir() or raw.startswith(_IMPORT_JUNK_PREFIXES):
        return None
    parts = pathlib.PurePosixPath(raw.replace("\\", "/")).parts
    if not parts or parts[-1] in _IMPORT_JUNK_NAMES:
        return None
    if any(p in ("..", "") for p in parts) or parts[0].endswith(":") or raw.startswith("/"):
        raise ValueError(f"unsafe path in zip: {raw!r}")
    if (info.external_attr >> 16) & 0o170000 == 0o120000:
        raise ValueError(f"symlinks are not allowed in imports: {raw!r}")
    return pathlib.PurePosixPath(*parts)


def _extract_zip_member(zf, info, dest: pathlib.Path, total: int) -> int:
    """Stream one member to *dest*, enforcing the cumulative byte cap.
    Returns the new cumulative total."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    limit_mb = _MAX_OUTPUT_BYTES // (1024 * 1024)
    with zf.open(info) as src, open(dest, "wb") as dst:
        while True:
            chunk = src.read(1024 * 1024)
            if not chunk:
                return total
            total += len(chunk)
            if total > _MAX_OUTPUT_BYTES:
                raise ValueError(f"zip contents too large (limit {limit_mb} MB uncompressed)")
            dst.write(chunk)


def _extract_upload(upload_path: pathlib.Path, filename: str, stage: pathlib.Path) -> list[str]:
    """Stage the uploaded file (single .xlsx) or its zip contents.
    Rejects path traversal, symlinks, and anything over the version
    limits. Returns the staged relative file names."""
    suffix = pathlib.PurePosixPath(filename).suffix.lower()
    if suffix == ".xlsx":
        if upload_path.stat().st_size > _MAX_OUTPUT_BYTES:
            raise ValueError(f"file too large (limit {_MAX_OUTPUT_BYTES // (1024 * 1024)} MB)")
        shutil.copy2(upload_path, stage / pathlib.PurePosixPath(filename).name)
        return [pathlib.PurePosixPath(filename).name]
    if suffix != ".zip":
        raise ValueError("upload must be a .xlsx workbook or a .zip of project files")

    try:
        zf = zipfile.ZipFile(upload_path)
    except zipfile.BadZipFile as exc:
        raise ValueError("the uploaded file is not a valid zip") from exc
    names: list[str] = []
    total = 0
    stage_resolved = stage.resolve()
    with zf:
        for info in zf.infolist():
            rel = _safe_zip_member(info)
            if rel is None:
                continue
            dest = (stage / rel).resolve()
            if not dest.is_relative_to(stage_resolved):
                raise ValueError(f"unsafe path in zip: {info.filename!r}")
            if len(names) + 1 > _MAX_OUTPUT_FILES:
                raise ValueError(f"too many files in zip (limit {_MAX_OUTPUT_FILES})")
            total = _extract_zip_member(zf, info, dest, total)
            names.append(str(rel))
    if not names:
        raise ValueError("the zip contains no usable files")
    return names


def import_upload(
    db,
    cfg,
    title: str,
    brief: str,
    upload_path: pathlib.Path,
    filename: str,
    kind: str | None = None,
) -> dict:
    """Create a project whose v1 is the uploaded workbook / zip contents.
    No build runs — the imported files ARE the first version. Verification
    problems (e.g. a workbook that will not load) are recorded as warnings
    on the version, not rejected: reviewing broken files is a valid use."""
    with tempfile.TemporaryDirectory() as tmp:
        stage = pathlib.Path(tmp) / "stage"
        stage.mkdir()
        names = _extract_upload(upload_path, filename, stage)
        for staged in sorted(stage.rglob("*.xlsx")):
            err = check_xlsx_zip_safety(staged)
            if err:
                raise ValueError(f"{staged.name}: {err}")
        resolved_kind = kind if kind in KINDS else _detect_kind(names)
        _ok, raw_checks = _verify_output(resolved_kind, stage)
        checks = {
            "imported": True,
            "source_filename": filename,
            "source_sha256": _sha256(upload_path),
            "import_warnings": raw_checks.get("problems") or [],
            **({"error": raw_checks["error"]} if raw_checks.get("error") else {}),
        }
        if resolved_kind == "xlsx" and not raw_checks.get("error"):
            # Record the six-gate proof result WITHOUT promotion: an imported
            # v1 stays byte-for-byte verbatim, but its proof status is known.
            from orivellum.capabilities.workbench_proof import prove_outputs

            workbooks = [p for p in sorted(stage.rglob("*.xlsx")) if p.is_file()]
            if workbooks:
                checks["proof"] = prove_outputs(stage, workbooks, promote=False)
        if checks.get("error"):
            # Structural limits (count/bytes) are enforced even for imports.
            raise ValueError(checks["error"])
        proj = db.create_wb_project(title, resolved_kind, brief)
        db.claim_wb_build(proj["id"])  # fresh project — always succeeds
        try:
            files = _snapshot(stage)
            _publish_version(
                db,
                cfg,
                proj["id"],
                stage,
                f"Imported from {filename}",
                files,
                checks,
                note=f"{len(files)} file(s) imported",
                verdict="imported",
            )
        except Exception:
            # Remove both the DB rows AND any files staged under the project
            # directory so a failed import leaves nothing behind.
            db.delete_wb_project(proj["id"])
            shutil.rmtree(project_dir(cfg, proj["id"]), ignore_errors=True)
            raise
        finally:
            db.update_wb_project(proj["id"], building=0)
    return db.get_wb_project(proj["id"])


def auto_review_upload(db, cfg, file_path: pathlib.Path, filename: str) -> None:
    """Automatic workbook review for an .xlsx that entered the Library:
    imports it as a Workbench project and runs the analysis, so every
    uploaded workbook gets a findings report without being asked.
    Gated by the ``workbench_auto_review`` setting (default on).
    Best-effort — a failure is logged, never raised into the upload path."""
    try:
        enabled = (db.get_setting("workbench_auto_review", "true") or "true").strip().lower()
        if enabled not in ("true", "1", "yes"):
            return
        stem = pathlib.PurePosixPath(filename).stem
        proj = import_upload(
            db,
            cfg,
            title=f"Review: {stem}"[:200],
            brief=f"Automatic workbook review of Library upload {filename}.",
            upload_path=file_path,
            filename=filename,
            kind="xlsx",
        )
        if db.claim_wb_build(proj["id"]):
            from orivellum.capabilities.workbench_analyze import run_analysis

            run_analysis(db, cfg, proj["id"], "")
    except Exception:  # noqa: BLE001 - review is a bonus; the upload already succeeded
        logger.exception("automatic workbook review failed for %s", filename)


def revert_to(db, cfg, project_id: str, version_no: int) -> dict:
    """Copy an existing version's files forward as a NEW version (history is
    append-only; nothing is ever rewritten)."""
    src = version_dir(cfg, project_id, version_no)
    if not src.is_dir():
        raise FileNotFoundError(f"version v{version_no} has no files on disk")
    files = _snapshot(src)
    return _publish_version(
        db, cfg, project_id, src, f"Revert to v{version_no}", files, {"reverted_from": version_no}
    )


def repair_and_prove(db, cfg, project_id: str, version_no: int) -> dict:
    """Publish a repaired, fully-gated COPY of a version as the next version.

    Imported versions stay byte-for-byte verbatim, so a workbook with blank
    or stale formula caches can only ever reach verdict 'provable' — it
    passes the six gates only after repairs it never received. This takes a
    copy of the version's files, applies the runner's repairs, runs all six
    gates against the repaired copy, and — only when every gate passes —
    publishes it as a new 'proven' version. History stays append-only: the
    source version is never touched, and any other outcome publishes
    nothing and raises instead.
    """
    proj = db.get_wb_project(project_id)
    if not proj:
        raise FileNotFoundError("project not found")
    if proj["kind"] != "xlsx":
        raise ValueError("repair & prove applies to workbook (xlsx) projects only")
    src = version_dir(cfg, project_id, version_no)
    if not src.is_dir():
        raise FileNotFoundError(f"version v{version_no} has no files on disk")
    with tempfile.TemporaryDirectory() as tmp:
        stage = pathlib.Path(tmp) / "stage"
        shutil.copytree(src, stage)
        workbooks = [p for p in sorted(stage.rglob("*.xlsx")) if p.is_file()]
        if not workbooks:
            raise ValueError(f"v{version_no} contains no .xlsx workbook to prove")
        from orivellum.capabilities.workbench_proof import prove_outputs

        proof = prove_outputs(stage, workbooks, promote=True)
        if proof["verdict"] != "proven":
            details = "; ".join(
                f"{name}: "
                + "; ".join((r.get("problems") or [r.get("error") or "gates failed"])[:2])
                for name, r in proof["workbooks"].items()
                if r["verdict"] != "proven"
            )[:400]
            raise ValueError(
                f"repair could not certify v{version_no} ({proof['verdict']})"
                + (f": {details}" if details else "")
            )
        refreshed = sum(
            r.get("repairs", {}).get("refreshed_cells", 0) for r in proof["workbooks"].values()
        )
        reordered = sum(
            r.get("repairs", {}).get("reordered_parts", 0) for r in proof["workbooks"].values()
        )
        files = _snapshot(stage)
        checks = {"proof": proof, "repaired_from": version_no}
        return _publish_version(
            db,
            cfg,
            project_id,
            stage,
            f"Repair & prove v{version_no}",
            files,
            checks,
            note=(
                f"{refreshed} cached value(s) refreshed, {reordered} part(s) reordered; "
                "all six gates passed on the repaired copy"
            ),
            verdict=_proof_verdict(checks),
        )


def archive_project(db, cfg, project_id: str, allow_unproven: bool = False) -> str:
    """Zip every version + a hash manifest; mark the project archived.

    For xlsx projects the latest version must be fully proven (all six
    gates) unless the caller explicitly passes ``allow_unproven=True`` —
    an archive should mean "recalculated and certified", not "loaded once".
    """
    proj = db.get_wb_project(project_id)
    if not proj:
        raise FileNotFoundError("project not found")
    versions = db.list_wb_versions(project_id)
    if not versions:
        raise ValueError("nothing to archive — the project has no versions")

    status, detail = latest_proof_status(proj, versions)
    if status not in ("proven", "n/a") and not allow_unproven:
        why = {
            "failed": "the latest version FAILED the proof gates",
            "provable": "the latest workbook passes the gates only after repairs "
            "it never received (imported files stay verbatim)",
            "unverified": "the latest version could not be recalculated",
            "unproven": "the latest version was never proven",
        }[status]
        raise UnprovenError(
            f"refusing to archive: {why}"
            + (f" ({detail})" if detail else "")
            + ". Complete anyway to archive unproven."
        )

    archives_dir(cfg).mkdir(parents=True, exist_ok=True)
    safe_title = (
        "".join(c if c.isalnum() or c in "-_" else "-" for c in proj["title"])[:60] or "project"
    )
    zip_path = archives_dir(cfg) / f"{safe_title}_{project_id[:8]}.zip"

    from orivellum.version import code_version

    manifest = {
        "code_version": code_version(),
        "project": {
            "id": proj["id"],
            "title": proj["title"],
            "kind": proj["kind"],
            "brief": proj["brief"],
            "created_at": proj["created_at"],
        },
        "archived_at": __import__("datetime")
        .datetime.now(__import__("datetime").timezone.utc)
        .isoformat(),
        "versions": [
            {
                "version_no": v["version_no"],
                "instruction": v["instruction"],
                "verdict": v["verdict"],
                "created_at": v["created_at"],
                "files": json.loads(v["files_json"] or "[]"),
            }
            for v in versions
        ],
    }
    # Integrity gate: every recorded file must exist on disk and re-hash to
    # the value stored when the version was accepted. Never produce an
    # archive that silently disagrees with the version history.
    for v in versions:
        vdir = version_dir(cfg, project_id, v["version_no"])
        for f in json.loads(v["files_json"] or "[]"):
            p = vdir / f["name"]
            if not p.is_file():
                raise RuntimeError(f"v{v['version_no']}/{f['name']} is missing on disk")
            if _sha256(p) != f["sha256"]:
                raise RuntimeError(
                    f"v{v['version_no']}/{f['name']} does not match "
                    "its recorded hash — refusing to archive"
                )

    tmp_path = zip_path.with_suffix(".zip.part")
    with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("manifest.json", json.dumps(manifest, indent=1))
        for v in versions:
            vdir = version_dir(cfg, project_id, v["version_no"])
            if not vdir.is_dir():
                continue
            for p in sorted(vdir.rglob("*")):
                if p.is_file():
                    z.write(p, f"v{v['version_no']}/{p.relative_to(vdir)}")
    tmp_path.replace(zip_path)
    db.update_wb_project(project_id, status="archived", archive_path=str(zip_path))
    return str(zip_path)


def delete_project(db, cfg, project_id: str) -> None:
    """Remove the project rows and its working files. The archive zip (if the
    project was archived) is intentionally kept."""
    db.delete_wb_project(project_id)
    pdir = project_dir(cfg, project_id)
    if pdir.is_dir():
        shutil.rmtree(pdir, ignore_errors=True)


def make_version_zip(cfg, project_id: str, version_no: int) -> pathlib.Path:
    """Build (or reuse) a downloadable zip of one version's files."""
    vdir = version_dir(cfg, project_id, version_no)
    if not vdir.is_dir():
        raise FileNotFoundError(f"version v{version_no} has no files on disk")
    zpath = project_dir(cfg, project_id) / f"v{version_no}.zip"
    tmp = zpath.with_suffix(".zip.part")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(vdir.rglob("*")):
            if p.is_file():
                z.write(p, str(p.relative_to(vdir)))
    tmp.replace(zpath)
    return zpath
