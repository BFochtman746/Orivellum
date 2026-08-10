"""Text extraction from documents of various types.

Supports: PDF (pdfplumber), DOCX (python-docx), Excel (openpyxl),
CSV, plain text, Markdown, images (pytesseract OCR), and a
markitdown fallback for everything else.

Returns an ExtractionResult with full text, per-page segments, and
file metadata needed for chunking and knowledge harvest.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class PageSegment:
    """A logical page or section of text."""

    page: int
    text: str
    heading: str = ""


@dataclass
class ExtractionResult:
    kind: str
    full_text: str
    word_count: int
    pages: list[PageSegment] = field(default_factory=list)
    headings: list[str] = field(default_factory=list)
    meta: dict = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return bool(self.full_text.strip())


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def _extract_pdf(path: Path, db=None) -> ExtractionResult:
    # --- pdfplumber (primary: best for text-layer PDFs) ---
    try:
        import pdfplumber

        pages: list[PageSegment] = []
        headings: list[str] = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = (page.extract_text() or "").strip()
                if not text:
                    continue
                seg = PageSegment(page=i, text=text)
                first_line = text.splitlines()[0].strip() if text else ""
                if first_line and len(first_line) < 120:
                    seg.heading = first_line
                    headings.append(first_line)
                pages.append(seg)
        if pages:
            full = "\n\n".join(p.text for p in pages)
            return ExtractionResult(
                kind="pdf",
                full_text=full,
                word_count=len(full.split()),
                pages=pages,
                headings=headings,
            )
        logger.info("pdfplumber found no text in %s — trying pypdf", path.name)
    except Exception as exc:
        logger.warning("pdfplumber failed on %s: %s — trying pypdf", path.name, exc)

    # --- pypdf (secondary: handles more edge cases / newer PDFs) ---
    try:
        import pypdf

        reader = pypdf.PdfReader(str(path))
        pages2: list[PageSegment] = []
        headings2: list[str] = []
        for i, page in enumerate(reader.pages, 1):
            text = (page.extract_text() or "").strip()
            if not text:
                continue
            seg = PageSegment(page=i, text=text)
            first_line = text.splitlines()[0].strip() if text else ""
            if first_line and len(first_line) < 120:
                seg.heading = first_line
                headings2.append(first_line)
            pages2.append(seg)
        if pages2:
            full2 = "\n\n".join(p.text for p in pages2)
            return ExtractionResult(
                kind="pdf",
                full_text=full2,
                word_count=len(full2.split()),
                pages=pages2,
                headings=headings2,
            )
        logger.info("pypdf also found no text in %s — falling back to markitdown", path.name)
    except Exception as exc:
        logger.warning("pypdf failed on %s: %s — falling back to markitdown", path.name, exc)

    # --- VLM OCR (scanned / image-only pages — when vision model configured) ---
    vlm_result = _vlm_pdf_ocr(path, db=db)
    if vlm_result is not None and vlm_result.ok:
        return vlm_result

    # --- markitdown (final fallback: handles image-heavy / complex PDFs) ---
    return _extract_fallback(path, "pdf")


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------


def _extract_docx(path: Path) -> ExtractionResult:
    try:
        import docx as _docx  # python-docx
    except ImportError:
        return _extract_fallback(path, "docx")

    headings: list[str] = []
    blocks: list[str] = []
    try:
        doc = _docx.Document(str(path))

        # Preserve document order by iterating the XML body children
        from docx.oxml.ns import qn  # type: ignore

        body = doc.element.body
        for child in body.iterchildren():
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

            if tag == "p":
                # Paragraph
                para_text = "".join(n.text or "" for n in child.iter(qn("w:t"))).strip()
                if not para_text:
                    continue
                blocks.append(para_text)
                style_name = ""
                pPr = child.find(qn("w:pPr"))
                if pPr is not None:
                    pStyle = pPr.find(qn("w:pStyle"))
                    if pStyle is not None:
                        style_name = pStyle.get(qn("w:val"), "")
                if style_name.lower().startswith("heading"):
                    headings.append(para_text)

            elif tag == "tbl":
                # Table — extract all cell text as a TSV block
                rows_text: list[str] = []
                for tr in child.iter(qn("w:tr")):
                    cells = []
                    for tc in tr.iter(qn("w:tc")):
                        cell_text = "".join(n.text or "" for n in tc.iter(qn("w:t"))).strip()
                        cells.append(cell_text)
                    if any(cells):
                        rows_text.append("\t".join(cells))
                if rows_text:
                    blocks.append("[Table]\n" + "\n".join(rows_text))

    except Exception as exc:
        logger.warning("python-docx failed on %s: %s", path.name, exc)
        return _extract_fallback(path, "docx")

    full = "\n".join(blocks)
    # Group into ~40-block pages for the chunker
    chunk_size = 40
    pages = [
        PageSegment(page=i + 1, text="\n".join(blocks[i * chunk_size : (i + 1) * chunk_size]))
        for i in range(max(1, (len(blocks) + chunk_size - 1) // chunk_size))
    ]
    return ExtractionResult(
        kind="docx",
        full_text=full,
        word_count=len(full.split()),
        pages=pages,
        headings=headings,
    )


# ---------------------------------------------------------------------------
# Excel / CSV
# ---------------------------------------------------------------------------


def _extract_excel(path: Path) -> ExtractionResult:
    try:
        import openpyxl
    except ImportError:
        return _extract_fallback(path, "excel")

    parts: list[str] = []
    headings: list[str] = []
    pages: list[PageSegment] = []
    try:
        # data_only=True reads cached values; formula-only files (never opened in Excel)
        # have no cached values → every cell is None → empty.  We detect that and retry
        # with data_only=False so formulas themselves are visible as text.
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        has_any_value = any(
            cell.value is not None
            for sheet in wb.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        if not has_any_value:
            wb.close()
            logger.info("Excel %s has no cached values — retrying with data_only=False", path.name)
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)

        for idx, sheet in enumerate(wb.worksheets, 1):
            rows_text: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = "\t".join(cells).strip()
                if line and line.replace("\t", ""):
                    rows_text.append(line)
            if rows_text:
                # Cap at 5000 rows per sheet (was 500); include row count note if truncated
                cap = 5000
                truncated = len(rows_text) > cap
                sheet_text = (
                    f"[Sheet: {sheet.title}]"
                    + (f" ({len(rows_text)} rows — showing first {cap})" if truncated else "")
                    + "\n"
                    + "\n".join(rows_text[:cap])
                )
                parts.append(sheet_text)
                headings.append(sheet.title)
                pages.append(PageSegment(page=idx, text=sheet_text, heading=sheet.title))
    except Exception as exc:
        logger.warning("openpyxl failed on %s: %s", path.name, exc)
        return _extract_fallback(path, "excel")

    full = "\n\n".join(parts)
    return ExtractionResult(
        kind="excel",
        full_text=full,
        word_count=len(full.split()),
        pages=pages,
        headings=headings,
    )


def _extract_csv(path: Path) -> ExtractionResult:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = ["\t".join(r) for r in reader]
        # No artificial cap — extract everything; guard only at 500k rows
        cap = 500_000
        truncated = len(rows) > cap
        full = "\n".join(rows[:cap])
        if truncated:
            full += f"\n\n[Truncated: showing first {cap:,} of {len(rows):,} rows]"
        return ExtractionResult(
            kind="csv",
            full_text=full,
            word_count=len(full.split()),
            pages=[PageSegment(page=1, text=full)],
        )
    except Exception as exc:
        logger.warning("CSV extraction failed on %s: %s", path.name, exc)
        return ExtractionResult(kind="csv", full_text="", word_count=0)


# ---------------------------------------------------------------------------
# Plain text / Markdown
# ---------------------------------------------------------------------------


def _extract_text(path: Path, kind: str = "text") -> ExtractionResult:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        logger.warning("Text read failed on %s: %s", path.name, exc)
        return ExtractionResult(kind=kind, full_text="", word_count=0)

    # Split into ~100-line pages
    lines = text.splitlines()
    page_size = 100
    pages = [
        PageSegment(page=i + 1, text="\n".join(lines[i * page_size : (i + 1) * page_size]))
        for i in range(max(1, (len(lines) + page_size - 1) // page_size))
    ]

    # Extract markdown headings
    headings = [
        line.lstrip("#").strip() for line in lines if line.startswith("#") and len(line) < 120
    ]

    return ExtractionResult(
        kind=kind,
        full_text=text,
        word_count=len(text.split()),
        pages=pages,
        headings=headings,
    )


# ---------------------------------------------------------------------------
# Image (OCR) + VLM PDF OCR
# ---------------------------------------------------------------------------

_VLM_OCR_PROMPT = (
    "Transcribe all text from this document image exactly, "
    "preserving structure, headings, and layout. "
    "Return only the text, with no commentary."
)


def _vlm_pdf_ocr(path: Path, db=None) -> ExtractionResult | None:
    """Use the configured vision LLM to OCR a scanned (image-only) PDF.

    Renders each page with pdf2image at 150 DPI, encodes as JPEG base64,
    and POSTs to the vision model with a transcription prompt.
    Returns None when the vision model is not configured, pdf2image is
    unavailable, or the attempt fails entirely — so the caller falls through
    to markitdown cleanly.
    """
    try:
        from orivellum.configuration.config import load_config

        cfg = load_config()
        _vision_model: str = ""
        try:
            from orivellum.api._deps import get_db as _get_db

            _db = _get_db()
            _vision_model = _db.get_setting("vision_model", "") or cfg.serving.vision_model
        except Exception:
            _vision_model = cfg.serving.vision_model
        if not _vision_model:
            return None

        try:
            from pdf2image import convert_from_path  # type: ignore[import]
        except ImportError:
            logger.debug("pdf2image not installed — VLM PDF OCR unavailable")
            return None

        import base64
        import io

        from orivellum.capabilities.llm import llm_call

        pages_pil = convert_from_path(str(path), dpi=150)
        if not pages_pil:
            return None

        page_segs: list[PageSegment] = []
        page_headings: list[str] = []

        for i, page_img in enumerate(pages_pil, 1):
            try:
                if page_img.mode not in ("RGB", "L"):
                    page_img = page_img.convert("RGB")
                buf = io.BytesIO()
                page_img.save(buf, format="JPEG", quality=85)
                b64 = base64.b64encode(buf.getvalue()).decode()

                result = llm_call(
                    [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": _VLM_OCR_PROMPT},
                                {
                                    "type": "image_url",
                                    "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
                                },
                            ],
                        }
                    ],
                    base_url=cfg.serving.base_url,
                    model=_vision_model,
                    timeout=cfg.serving.extraction_timeout_sec,
                    purpose="extraction.ocr",
                    db=db,
                )
                text = (result.text or "").strip()
                if not text:
                    continue
                seg = PageSegment(page=i, text=text)
                first_line = text.splitlines()[0].strip()
                if first_line and len(first_line) < 120:
                    seg.heading = first_line
                    page_headings.append(first_line)
                page_segs.append(seg)
            except Exception as page_exc:
                logger.warning("VLM OCR failed on page %d of %s: %s", i, path.name, page_exc)

        if not page_segs:
            return None

        full = "\n\n".join(s.text for s in page_segs)
        logger.info(
            "VLM PDF OCR: %s — %d pages, %d words (model=%s)",
            path.name,
            len(page_segs),
            len(full.split()),
            _vision_model,
        )
        return ExtractionResult(
            kind="pdf",
            full_text=full,
            word_count=len(full.split()),
            pages=page_segs,
            headings=page_headings,
            meta={"ocr_engine": "vlm", "ocr_model": _vision_model},
        )
    except Exception as exc:
        logger.warning("VLM PDF OCR failed on %s: %s", path.name, exc)
    return None


def _probe_tesseract() -> None:
    """Ensure pytesseract can find the tesseract binary.

    On NixOS/Replit the binary lands in the nix store but may not be
    on the process PATH.  We probe common locations and set
    pytesseract.pytesseract.tesseract_cmd explicitly when needed.
    """
    import shutil
    import subprocess as _sp

    import pytesseract as _pt

    if shutil.which("tesseract"):
        return  # already on PATH

    import sys as _sys

    if _sys.platform != "win32":
        # On Unix/NixOS ask the login shell — it has a broader PATH
        try:
            result = _sp.run(
                ["bash", "-lc", "which tesseract"],
                capture_output=True,
                text=True,
                timeout=5,
            )
            candidate = result.stdout.strip()
            if candidate and Path(candidate).is_file():
                _pt.pytesseract.tesseract_cmd = candidate
                return
        except Exception:
            pass

        # Known nix store prefix pattern — walk /nix/store at depth-1 only
        nix_store = Path("/nix/store")
        if nix_store.exists():
            for pkg_dir in nix_store.iterdir():
                if "tesseract" in pkg_dir.name:
                    candidate = pkg_dir / "bin" / "tesseract"
                    if candidate.is_file():
                        _pt.pytesseract.tesseract_cmd = str(candidate)
                        return
    else:
        # Common Windows install location from the UB-Mannheim installer
        win_default = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
        if win_default.is_file():
            _pt.pytesseract.tesseract_cmd = str(win_default)
            return


def _extract_image_vision(path: Path, db=None) -> ExtractionResult | None:
    """Use the configured vision LLM to describe image content.

    Returns an ExtractionResult when the vision model is configured and
    responds successfully, or None so the caller falls through to Tesseract.
    """
    try:
        from orivellum.configuration.config import load_config

        cfg = load_config()
        # Resolve vision model: DB setting overrides YAML config.
        # Import here (lazy) to avoid circular import at module load.
        _vision_model: str = ""
        try:
            from orivellum.api._deps import get_db as _get_db

            _db = _get_db()
            _vision_model = _db.get_setting("vision_model", "") or cfg.serving.vision_model
        except Exception:
            _vision_model = cfg.serving.vision_model
        if not _vision_model:
            return None

        import base64
        import io

        from PIL import Image as _PIL

        from orivellum.capabilities.llm import llm_call

        img = _PIL.open(path)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        b64 = base64.b64encode(buf.getvalue()).decode()

        result = llm_call(
            [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Describe everything in this image in detail. "
                                "Include all visible text, numbers, labels, "
                                "diagrams, charts, tables, and visual elements. "
                                "Be thorough — this description will be used for "
                                "search and knowledge extraction."
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
                        },
                    ],
                }
            ],
            base_url=cfg.serving.base_url,
            model=_vision_model,
            timeout=cfg.serving.extraction_timeout_sec,
            purpose="extraction.llm",
            db=db,
        )
        text = result.text
        if text and text.strip():
            logger.info("Vision model described %s (%d words)", path.name, len(text.split()))
            return ExtractionResult(
                kind="image",
                full_text=text,
                word_count=len(text.split()),
                pages=[PageSegment(page=1, text=text)],
            )
    except Exception as exc:
        logger.warning("Vision extraction failed on %s: %s", path.name, exc)
    return None


def _extract_image(path: Path, db=None) -> ExtractionResult:
    # --- vision LLM (primary when configured) ---
    vision_result = _extract_image_vision(path, db=db)
    if vision_result is not None:
        return vision_result

    # --- pytesseract (fallback: requires tesseract binary) ---
    try:
        import pytesseract
        from PIL import Image

        _probe_tesseract()
        img = Image.open(path)
        # Pre-process: convert to RGB so tesseract handles all modes
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        text = pytesseract.image_to_string(img, config="--psm 6")
        if text.strip():
            return ExtractionResult(
                kind="image",
                full_text=text,
                word_count=len(text.split()),
                pages=[PageSegment(page=1, text=text)],
            )
        logger.info("pytesseract returned no text for %s", path.name)
    except pytesseract.TesseractNotFoundError:
        logger.warning("tesseract not found — skipping OCR for %s", path.name)
    except Exception as exc:
        logger.warning("pytesseract OCR failed on %s: %s", path.name, exc)

    # --- markitdown fallback (handles some image formats) ---
    result = _extract_fallback(path, "image")
    if result.ok:
        return result

    # No text recovered — mark explicitly so callers can surface it to the user
    return ExtractionResult(kind="no_text", full_text="", word_count=0)


# ---------------------------------------------------------------------------
# Markitdown fallback
# ---------------------------------------------------------------------------


def _extract_fallback(path: Path, kind: str) -> ExtractionResult:
    try:
        from markitdown import MarkItDown

        md = MarkItDown()
        result = md.convert(str(path))
        text = result.text_content or ""
        return ExtractionResult(
            kind=kind,
            full_text=text,
            word_count=len(text.split()),
            pages=[PageSegment(page=1, text=text)],
        )
    except Exception as exc:
        logger.warning("markitdown fallback failed on %s: %s", path.name, exc)
        return ExtractionResult(kind=kind, full_text="", word_count=0)


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------


def _extract_html(path: Path) -> ExtractionResult:
    """Strip tags and extract readable text from an HTML file."""
    from html.parser import HTMLParser

    class _Stripper(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self._skip = False
            self._parts: list[str] = []

        def handle_starttag(self, tag: str, attrs: object) -> None:
            if tag in ("script", "style", "head", "nav", "footer"):
                self._skip = True

        def handle_endtag(self, tag: str) -> None:
            if tag in ("script", "style", "head", "nav", "footer"):
                self._skip = False
            if tag in ("p", "br", "div", "li", "h1", "h2", "h3", "h4", "h5", "h6"):
                self._parts.append("\n")

        def handle_data(self, data: str) -> None:
            if not self._skip and data.strip():
                self._parts.append(data)

    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
        stripper = _Stripper()
        stripper.feed(raw)
        full_text = " ".join(stripper._parts).strip()
        pages = [PageSegment(page=1, text=full_text)] if full_text else []
        return ExtractionResult(
            kind="html",
            full_text=full_text,
            word_count=len(full_text.split()),
            pages=pages,
        )
    except Exception as exc:
        logger.warning("HTML extraction failed for %s: %s", path.name, exc)
        return ExtractionResult(kind="html", full_text="", word_count=0)


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


def _extract_json(path: Path) -> ExtractionResult:
    """Format JSON as indented text for indexing."""
    import json

    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        full_text = json.dumps(data, indent=2, ensure_ascii=False)
    except Exception:
        # Malformed JSON — fall back to plain-text read
        full_text = path.read_text(encoding="utf-8", errors="replace")
    pages = [PageSegment(page=1, text=full_text)] if full_text.strip() else []
    return ExtractionResult(
        kind="json",
        full_text=full_text,
        word_count=len(full_text.split()),
        pages=pages,
    )


# ---------------------------------------------------------------------------
# ZIP archive
# ---------------------------------------------------------------------------


def _extract_zip(path: Path) -> ExtractionResult:
    """Extract text from every supported file inside a ZIP archive.

    Each member's content is concatenated under a ``=== filename ===`` header
    so the whole archive becomes one searchable document.

    Hardening:
    - Skips macOS metadata (``__MACOSX/`` entries and ``._*`` dotfiles)
    - Avoids basename collisions by prefixing each extracted file with its index
    - Records per-member failure reasons in the returned meta dict
    - Falls back to markitdown for formats not in the known list
    """
    import tempfile
    import zipfile

    # Extension → kind map (subset of _KIND_MAP; resolved at call-time)
    _EXT_KIND = {
        ".pdf": "pdf",
        ".docx": "docx",
        ".doc": "docx",
        ".xlsx": "excel",
        ".xls": "excel",
        ".csv": "csv",
        ".txt": "text",
        ".md": "markdown",
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".html": "html",
        ".htm": "html",
        ".json": "json",
        ".py": "code",
        ".js": "code",
        ".ts": "code",
        ".pptx": "pptx",
        ".ppt": "pptx",
        ".eml": "email",
        ".msg": "email",
    }

    all_text: list[str] = []
    all_pages: list[PageSegment] = []
    all_headings: list[str] = []
    # Collect per-member results for informative error messages
    member_results: list[dict] = []  # {name, status, reason}

    try:
        with zipfile.ZipFile(path, "r") as zf:
            members = [n for n in zf.namelist() if not n.endswith("/")]
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp = Path(tmpdir)
                for idx, name in enumerate(members):
                    basename = Path(name).name

                    # ── Skip macOS metadata ────────────────────────────────────
                    if name.startswith("__MACOSX/") or basename.startswith("._"):
                        continue

                    ext = Path(name).suffix.lower()
                    kind = _EXT_KIND.get(ext)
                    if not kind:
                        member_results.append(
                            {
                                "name": name,
                                "status": "skipped",
                                "reason": f"unsupported type ({ext or 'no extension'})",
                            }
                        )
                        continue

                    # ── Collision-safe extraction: prefix with index ────────────
                    safe_name = f"{idx:04d}_{basename}"
                    member_path = tmp / safe_name
                    try:
                        member_path.write_bytes(zf.read(name))
                    except Exception as exc:
                        logger.warning("ZIP: could not read member %s: %s", name, exc)
                        member_results.append(
                            {"name": name, "status": "error", "reason": f"read error: {exc}"}
                        )
                        continue

                    handler = _DISPATCH.get(kind, lambda p: _extract_fallback(p, kind))
                    try:
                        sub = handler(member_path)  # type: ignore[call-arg]
                    except Exception as exc:
                        logger.warning("ZIP: extraction failed for %s: %s", name, exc)
                        member_results.append(
                            {"name": name, "status": "error", "reason": f"extraction error: {exc}"}
                        )
                        continue

                    if not sub.full_text.strip():
                        member_results.append(
                            {"name": name, "status": "empty", "reason": "no readable text found"}
                        )
                        continue

                    header = f"=== {name} ==="
                    all_text.append(f"{header}\n{sub.full_text}")
                    all_headings.append(header)
                    offset = len(all_pages)
                    for seg in sub.pages:
                        all_pages.append(
                            PageSegment(
                                page=offset + seg.page,
                                text=seg.text,
                                heading=name,
                            )
                        )
                    member_results.append(
                        {"name": name, "status": "ok", "reason": f"{sub.word_count} words"}
                    )

    except zipfile.BadZipFile as exc:
        logger.error("Bad ZIP file %s: %s", path.name, exc)
        return ExtractionResult(
            kind="zip", full_text="", word_count=0, meta={"error": f"Invalid ZIP file: {exc}"}
        )

    full_text = "\n\n".join(all_text)
    ok_count = sum(1 for r in member_results if r["status"] == "ok")
    fail_count = sum(1 for r in member_results if r["status"] in ("error", "empty"))
    skip_count = sum(1 for r in member_results if r["status"] == "skipped")

    meta: dict = {
        "zip_members": member_results,
        "zip_summary": f"{ok_count} extracted, {fail_count} empty/failed, {skip_count} skipped",
    }

    if not full_text.strip():
        # Build a human-readable explanation from member_results
        lines = []
        if not member_results:
            lines.append("ZIP archive is empty or contains no files.")
        else:
            tried = [r for r in member_results if r["status"] != "skipped"]
            if not tried:
                exts = list({Path(r["name"]).suffix.lower() for r in member_results})
                lines.append(
                    f"ZIP contains {len(member_results)} file(s) but none have a "
                    f"supported format. Found: {', '.join(exts) or 'unknown'}"
                )
            else:
                lines.append(
                    f"ZIP contains {len(member_results)} file(s); none produced readable text:"
                )
                for r in tried[:8]:
                    lines.append(f"  • {Path(r['name']).name}: {r['reason']}")
                if len(tried) > 8:
                    lines.append(f"  … and {len(tried) - 8} more")
        meta["user_message"] = " ".join(lines)

    return ExtractionResult(
        kind="zip",
        full_text=full_text,
        word_count=len(full_text.split()),
        pages=all_pages,
        headings=all_headings,
        meta=meta,
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# faster-whisper — lazy singleton (loaded on first transcription call)
# ---------------------------------------------------------------------------
# Use threading.Lock to guard against concurrent first-call races.
# _fw_instance is:
#   None  — not yet attempted
#   False — attempted and failed (package missing or load error)
#   WhisperModel — successfully loaded

import threading as _threading

_fw_lock = _threading.Lock()
_fw_instance: object = None  # WhisperModel | False | None
_fw_loaded_size: str = ""  # model size string for the loaded singleton
_fw_requested_size: str = ""  # the REQUEST that produced the current state
_fw_fallback_reason: str | None = None  # why a smaller model was substituted

# Model sizes that need substantial RAM (weights + CTranslate2 workspace).
# Anything here falls back to "base" when the machine can't hold it —
# a wrong-but-working transcription beats an OOM-killed server.
_FW_HEAVY_SIZES = {"medium", "large-v2", "large-v3", "large-v3-turbo", "distil-large-v3"}
_FW_HEAVY_MIN_AVAILABLE_BYTES = 6 * 1024**3  # 6 GB free RAM required for heavy tiers
_FW_FALLBACK_SIZE = "base"

# Valid sizes for the runtime setting (UI/API validation).
FW_ALLOWED_SIZES = (
    "tiny",
    "base",
    "small",
    "medium",
    "large-v3",
    "large-v3-turbo",
    "distil-large-v3",
)


def _resolve_asr_local_model(db=None, cfg_default: str = "large-v3-turbo") -> str:
    """Effective local ASR model size: DB setting overrides config.yaml."""
    if db is not None:
        try:
            override = (db.get_setting("asr_local_model", "") or "").strip()
            if override in FW_ALLOWED_SIZES:
                return override
        except Exception:
            pass
    return cfg_default or "large-v3-turbo"


def _fw_effective_size(model_size: str) -> tuple[str, str | None]:
    """Apply the low-memory guard; returns (size_to_load, fallback_reason)."""
    if model_size not in _FW_HEAVY_SIZES:
        return model_size, None
    try:
        import psutil  # type: ignore[import]

        available = psutil.virtual_memory().available
        if available < _FW_HEAVY_MIN_AVAILABLE_BYTES:
            reason = (
                f"only {available / 1024**3:.1f} GB RAM available "
                f"(< {_FW_HEAVY_MIN_AVAILABLE_BYTES / 1024**3:.0f} GB needed for "
                f"'{model_size}') — using '{_FW_FALLBACK_SIZE}'"
            )
            return _FW_FALLBACK_SIZE, reason
    except Exception:
        pass  # psutil missing / probe failed: trust the requested size
    return model_size, None


def _fw_snapshot_for(model_size: str) -> tuple[object, str, str | None] | None:
    """Return (model, loaded_size, fallback_reason) when the current singleton
    state satisfies *model_size*, or None when a (re)load is needed.

    A request is satisfied when it is the SAME request that produced the
    current state (even if a fallback model was substituted — never re-attempt
    the heavy load on every call) or when the loaded model is exactly the
    requested size.
    """
    inst, loaded, requested, reason = (
        _fw_instance,
        _fw_loaded_size,
        _fw_requested_size,
        _fw_fallback_reason,
    )
    if inst is False and requested == model_size:
        return (False, "", None)  # this exact request already failed
    if inst is not None and inst is not False:
        if loaded == model_size:
            # Exact match — no fallback happened *for this request*.
            return (inst, loaded, None)
        if requested == model_size:
            # Same request as before — reuse the fallback that was chosen.
            return (inst, loaded, reason)
    return None


def _get_faster_whisper_snapshot(
    model_size: str = "large-v3-turbo",
) -> tuple[object | None, str, str | None]:
    """Return (model | None, loaded_size, fallback_reason) atomically.

    The singleton reloads when a DIFFERENT size is requested (settings change),
    substitutes "base" on low-memory machines, and retries with "base" when a
    heavy model fails to load. Both fallbacks are cached per request, so a
    failing heavy model is never re-downloaded/re-loaded on every call.
    Model is None when the package is not installed or no model could load.

    Note: during a rare settings-driven reload, an in-flight transcription may
    keep the OLD model alive via its local reference until it finishes — a
    transient double-residency we accept because reloads are user-initiated.
    """
    global _fw_instance, _fw_loaded_size, _fw_requested_size, _fw_fallback_reason

    snap = _fw_snapshot_for(model_size)
    if snap is not None:
        model, loaded, reason = snap
        return (None if model is False else model, loaded, reason)

    with _fw_lock:
        snap = _fw_snapshot_for(model_size)
        if snap is not None:
            model, loaded, reason = snap
            return (None if model is False else model, loaded, reason)
        try:
            from faster_whisper import WhisperModel  # type: ignore[import]
        except ImportError:
            logger.info("faster-whisper not installed — local ASR fallback unavailable")
            _fw_requested_size = model_size
            _fw_instance = False
            return (None, "", None)

        target, mem_reason = _fw_effective_size(model_size)
        if mem_reason:
            logger.warning("faster-whisper low-memory fallback: %s", mem_reason)

        _fw_instance = None  # drop any previously loaded model before reloading
        for attempt_size in dict.fromkeys([target, _FW_FALLBACK_SIZE]):
            try:
                logger.info(
                    "Loading faster-whisper model '%s' (int8 — first-run "
                    "download may take a moment)…",
                    attempt_size,
                )
                model = WhisperModel(attempt_size, device="auto", compute_type="int8")
                if attempt_size == model_size:
                    reason = None
                elif attempt_size == target and mem_reason:
                    reason = mem_reason
                else:
                    reason = f"'{target}' failed to load — using '{attempt_size}'"
                # Order matters for the lock-free fast path: set metadata
                # first, the instance last.
                _fw_loaded_size = attempt_size
                _fw_requested_size = model_size
                _fw_fallback_reason = reason
                _fw_instance = model
                logger.info("faster-whisper model '%s' ready.", attempt_size)
                return (model, attempt_size, reason)
            except Exception as exc:
                logger.warning("faster-whisper failed to load model '%s': %s", attempt_size, exc)
        _fw_loaded_size = ""
        _fw_requested_size = model_size  # remember which REQUEST failed
        _fw_fallback_reason = None
        _fw_instance = False
    return (None, "", None)


def _get_faster_whisper(model_size: str = "large-v3-turbo"):
    """Back-compat wrapper: return just the cached WhisperModel (or None)."""
    return _get_faster_whisper_snapshot(model_size)[0]


def _is_faster_whisper_loaded() -> bool:
    """True only when the WhisperModel is actually in memory and ready to use."""
    return _fw_instance is not None and _fw_instance is not False


def faster_whisper_status() -> dict:
    """Snapshot for status endpoints: loaded size + any fallback reason."""
    loaded = _is_faster_whisper_loaded()
    return {
        "loaded": loaded,
        "loaded_size": _fw_loaded_size if loaded else None,
        "fallback_reason": _fw_fallback_reason,
    }


# Caps keep document meta a reasonable size even for hours-long recordings.
_FW_MAX_SEGMENTS_META = 2000
_FW_MAX_WORDS_META = 6000


def _transcribe_faster_whisper(
    path: Path, model_size: str = "large-v3-turbo"
) -> ExtractionResult | None:
    """Transcribe *path* locally using faster-whisper.

    Captures per-segment AND per-word timestamps into the result meta
    (enables read-along highlighting later). Returns None when the package
    is absent or transcription fails so the caller can fall through to the
    metadata-only result.
    """
    model, loaded_size, fallback_reason = _get_faster_whisper_snapshot(model_size)
    if model is None:
        return None
    try:
        segments, info = model.transcribe(str(path), word_timestamps=True)
        seg_meta: list[dict] = []
        word_meta: list[dict] = []
        texts: list[str] = []
        words_truncated = False
        for seg in segments:  # generator — consumed once
            seg_text = seg.text.strip()
            if seg_text:
                texts.append(seg_text)
            if len(seg_meta) < _FW_MAX_SEGMENTS_META:
                seg_meta.append(
                    {
                        "start": round(float(seg.start), 2),
                        "end": round(float(seg.end), 2),
                        "text": seg_text,
                    }
                )
            for w in seg.words or []:
                if len(word_meta) >= _FW_MAX_WORDS_META:
                    words_truncated = True
                    break
                word_meta.append(
                    {
                        "start": round(float(w.start), 2),
                        "end": round(float(w.end), 2),
                        "word": w.word.strip(),
                    }
                )
        text = " ".join(texts).strip()
        if not text:
            logger.info("faster-whisper returned no text for %s", path.name)
            return None
        full = f"[Audio transcript: {path.name}]\n\n{text}"
        logger.info(
            "faster-whisper transcription OK: %d words from %s (model=%s)",
            len(text.split()),
            path.name,
            loaded_size or model_size,
        )
        meta: dict = {
            "transcription": "faster_whisper",
            "model_size": loaded_size or model_size,  # what actually ran
            "model_requested": model_size,
            "source": str(path.name),
            "language": getattr(info, "language", None),
            "duration": round(float(getattr(info, "duration", 0.0)), 2),
            "segments": seg_meta,
            "words": word_meta,
        }
        if words_truncated:
            meta["words_truncated"] = True
        if fallback_reason:
            meta["model_fallback_reason"] = fallback_reason
        return ExtractionResult(
            kind="audio",
            full_text=full,
            word_count=len(text.split()),
            pages=[PageSegment(page=1, text=text)],
            meta=meta,
        )
    except Exception as exc:
        logger.warning("faster-whisper transcription failed for %s: %s", path.name, exc)
        return None


# ---------------------------------------------------------------------------
# Audio extraction (AI server → faster-whisper → metadata-only)
# ---------------------------------------------------------------------------


def _extract_audio(path: Path, db=None) -> ExtractionResult:
    """Transcribe an audio file.

    Priority order:
      1. AI server /v1/audio/transcriptions (OpenAI-compatible Whisper API)
      2. faster-whisper local (CTranslate2 — ~4× faster than vanilla Whisper)
      3. Metadata-only placeholder so the file is still searchable/downloadable

    ``db`` (optional) is threaded from the public extract() entry point.
    """
    import json as _json
    import mimetypes as _mt
    import urllib.request as _urlr

    base_url: str = ""
    asr_model: str = "whisper-1"
    asr_local_model: str = "large-v3-turbo"
    try:
        # Prefer the request-scoped config when running inside the FastAPI server.
        # Fall back to load_config() when called from a standalone script or test.
        try:
            from orivellum.api._deps import get_config as _get_cfg

            _cfg_obj = _get_cfg()
            base_url = _cfg_obj.serving.base_url.rstrip("/")
            asr_model = _cfg_obj.serving.asr_model
            asr_local_model = getattr(_cfg_obj.serving, "asr_local_model", "large-v3-turbo")
        except Exception:
            from orivellum.configuration.config import load_config as _load_cfg

            _cfg_obj = _load_cfg()
            base_url = _cfg_obj.serving.base_url.rstrip("/")
            asr_model = getattr(_cfg_obj.serving, "asr_model", "whisper-1")
            asr_local_model = getattr(_cfg_obj.serving, "asr_local_model", "large-v3-turbo")
    except Exception:
        pass
    # DB setting overrides config.yaml (same precedence as chat model overrides).
    asr_local_model = _resolve_asr_local_model(db, asr_local_model)

    def _metadata_only(msg: str) -> ExtractionResult:
        text = (
            f"Audio file: {path.name}\n"
            f"Note: {msg}\n\n"
            "To enable transcription: (1) ensure your AI server exposes "
            "/v1/audio/transcriptions, or (2) install faster-whisper "
            "(`uv add faster-whisper`) for fully offline transcription."
        )
        return ExtractionResult(
            kind="audio",
            full_text=text,
            word_count=len(text.split()),
            meta={"transcription": None, "reason": msg},
        )

    # ── 0. Optional audio enhancement (DeepFilterNet3) ───────────────────────
    # Pre-process the audio before Whisper to remove background noise, room
    # reverb, and non-stationary interference.  Best-effort — any failure
    # silently falls back to the original file so transcription always runs.
    import shutil as _sh_enh
    import tempfile as _tf_enh  # noqa: E401

    _enh_tmp: str | None = None
    transcribe_path = path
    try:
        if db is not None and db.get_setting("audio_enhance_enabled", "false").lower() == "true":
            from orivellum.capabilities.enhancement import (  # noqa: PLC0415
                enhance_audio as _enhance_audio,
            )
            from orivellum.capabilities.enhancement import (
                is_available as _dfn_available,
            )

            if _dfn_available():
                _enh_tmp = _tf_enh.mkdtemp(prefix="orivellum_dfn_")
                _ep = _enhance_audio(path, output_dir=Path(_enh_tmp))
                if _ep != path:
                    transcribe_path = _ep
                    logger.info(
                        "DeepFilterNet3 applied for %s; transcribing enhanced audio",
                        path.name,
                    )
    except Exception as _enh_err:
        logger.debug("Audio enhancement skipped: %s", _enh_err)

    try:
        # ── 1. AI server (Lemonade / any OpenAI-compatible Whisper endpoint) ─
        ai_server_exc: str = ""
        if base_url:
            try:
                mime_type, _ = _mt.guess_type(path.name)
                mime_type = mime_type or "application/octet-stream"

                import uuid as _uuid

                boundary = f"---{_uuid.uuid4().hex}"
                filename = path.name  # keep original filename in the multipart header

                with open(transcribe_path, "rb") as fh:
                    file_bytes = fh.read()

                body_parts: list[bytes] = []
                body_parts.append(
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
                    f"Content-Type: {mime_type}\r\n\r\n".encode()
                    + file_bytes
                    + b"\r\n"
                )
                body_parts.append(
                    f"--{boundary}\r\n"
                    'Content-Disposition: form-data; name="model"\r\n\r\n'
                    f"{asr_model}\r\n".encode()
                )
                file_and_model = b"".join(body_parts)

                def _ai_request(verbose: bool) -> dict:
                    extra = b""
                    if verbose:
                        # verbose_json returns timestamps; OpenAI-compatible
                        # servers need BOTH granularities requested explicitly
                        # or word-level timing is omitted.
                        extra = (
                            f"--{boundary}\r\n"
                            'Content-Disposition: form-data; name="response_format"\r\n\r\n'
                            "verbose_json\r\n"
                            f"--{boundary}\r\n"
                            'Content-Disposition: form-data; name="timestamp_granularities[]"\r\n\r\n'
                            "segment\r\n"
                            f"--{boundary}\r\n"
                            'Content-Disposition: form-data; name="timestamp_granularities[]"\r\n\r\n'
                            "word\r\n".encode()
                        )
                    body = file_and_model + extra + f"--{boundary}--\r\n".encode()
                    req = _urlr.Request(
                        f"{base_url}/audio/transcriptions",
                        data=body,
                        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
                        method="POST",
                    )
                    with _urlr.urlopen(req, timeout=120) as resp:
                        return _json.loads(resp.read().decode("utf-8"))

                # Try with timestamps first; some servers reject unknown form
                # fields, so retry once in plain mode before giving up on the
                # AI server entirely.
                try:
                    data = _ai_request(verbose=True)
                except Exception:
                    data = _ai_request(verbose=False)
                transcript = (data.get("text") or "").strip()
                if transcript:
                    logger.info(
                        "AI server transcription OK: %d words from %s",
                        len(transcript.split()),
                        path.name,
                    )
                    meta: dict = {
                        "transcription": "ai_server",
                        "asr_model": asr_model,
                        "source": str(path.name),
                        "enhanced": transcribe_path != path,
                    }
                    raw_segs = data.get("segments")
                    if isinstance(raw_segs, list) and raw_segs:
                        segs = []
                        for s in raw_segs[:_FW_MAX_SEGMENTS_META]:
                            try:
                                segs.append(
                                    {
                                        "start": round(float(s.get("start", 0.0)), 2),
                                        "end": round(float(s.get("end", 0.0)), 2),
                                        "text": str(s.get("text", "")).strip(),
                                    }
                                )
                            except Exception:
                                continue
                        if segs:
                            meta["segments"] = segs
                    raw_words = data.get("words")
                    if isinstance(raw_words, list) and raw_words:
                        words = []
                        for w in raw_words[:_FW_MAX_WORDS_META]:
                            try:
                                words.append(
                                    {
                                        "start": round(float(w.get("start", 0.0)), 2),
                                        "end": round(float(w.get("end", 0.0)), 2),
                                        "word": str(w.get("word", "")).strip(),
                                    }
                                )
                            except Exception:
                                continue
                        if words:
                            meta["words"] = words
                    if data.get("language"):
                        meta["language"] = data.get("language")
                    if data.get("duration") is not None:
                        try:
                            meta["duration"] = round(float(data["duration"]), 2)
                        except Exception:
                            pass
                    return ExtractionResult(
                        kind="audio",
                        full_text=f"[Audio transcript: {path.name}]\n\n{transcript}",
                        word_count=len(transcript.split()),
                        pages=[PageSegment(page=1, text=transcript)],
                        meta=meta,
                    )
                ai_server_exc = "AI server returned an empty transcription"
                logger.warning("AI server transcription: empty response for %s", path.name)
            except Exception as exc:
                ai_server_exc = str(exc)
                logger.info(
                    "AI server transcription unavailable for %s (%s) — trying faster-whisper",
                    path.name,
                    exc,
                )

        # ── 2. faster-whisper local ───────────────────────────────────────────
        fw_result = _transcribe_faster_whisper(transcribe_path, asr_local_model)
        if fw_result is not None:
            return fw_result

        # ── 3. Metadata-only ─────────────────────────────────────────────────
        reason = ai_server_exc or "AI server URL not configured"
        if not _is_faster_whisper_loaded():
            reason += " | faster-whisper not installed"
        return _metadata_only(reason)

    finally:
        # Always clean up the temporary enhanced audio file.
        if _enh_tmp:
            _sh_enh.rmtree(_enh_tmp, ignore_errors=True)


def _extract_email(path: Path) -> ExtractionResult:
    """Extract text from an .eml (RFC 2822) or Outlook .msg email file.

    For .eml files, Python's stdlib ``email`` module parses headers and body
    with no extra dependencies.  For .msg files the optional ``extract_msg``
    package is tried; if absent the file is read as raw bytes and any readable
    ASCII/UTF-8 content is returned.

    Attachments within the message are noted in metadata but not recursively
    extracted (the pipeline can re-process them if they are saved separately).
    """
    import email as _email_mod
    import email.policy as _ep

    suffix = path.suffix.lower()
    attachments: list[str] = []

    # ── .eml — stdlib ─────────────────────────────────────────────────────────
    if suffix == ".eml":
        try:
            raw = path.read_bytes()
            msg = _email_mod.message_from_bytes(raw, policy=_ep.default)
            subject = str(msg.get("Subject", "")).strip()
            from_hdr = str(msg.get("From", "")).strip()
            to_hdr = str(msg.get("To", "")).strip()
            date_hdr = str(msg.get("Date", "")).strip()

            body_parts: list[str] = []
            for part in msg.walk():
                ct = part.get_content_type()
                cd = str(part.get("Content-Disposition", ""))
                if "attachment" in cd:
                    fn = part.get_filename()
                    if fn:
                        attachments.append(fn)
                    continue
                if ct == "text/plain":
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        payload = part.get_payload(decode=True)
                        if payload:
                            body_parts.append(payload.decode("utf-8", errors="replace"))
                elif ct == "text/html" and not body_parts:
                    # Use HTML body only if no plain text was found
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass

            header_block = "\n".join(
                filter(
                    None,
                    [
                        f"Subject: {subject}" if subject else "",
                        f"From: {from_hdr}" if from_hdr else "",
                        f"To: {to_hdr}" if to_hdr else "",
                        f"Date: {date_hdr}" if date_hdr else "",
                    ],
                )
            )
            body_text = "\n\n".join(body_parts).strip()
            full_text = f"{header_block}\n\n{body_text}".strip() if body_text else header_block
            if attachments:
                full_text += f"\n\n[Attachments: {', '.join(attachments)}]"

            return ExtractionResult(
                kind="email",
                full_text=full_text,
                word_count=len(full_text.split()),
                pages=[PageSegment(page=1, text=full_text)],
                meta={
                    "subject": subject,
                    "from": from_hdr,
                    "to": to_hdr,
                    "date": date_hdr,
                    "attachments": attachments,
                },
            )
        except Exception as exc:
            logger.warning("EML extraction failed for %s: %s", path.name, exc)
            return ExtractionResult(kind="email", full_text="", word_count=0)

    # ── .msg — Outlook compound document ─────────────────────────────────────
    if suffix == ".msg":
        try:
            import extract_msg as _em  # type: ignore[import]

            with _em.Message(str(path)) as m:
                subject = (m.subject or "").strip()
                from_hdr = (m.sender or "").strip()
                to_hdr = (m.to or "").strip()
                date_hdr = str(m.date or "").strip()
                body_text = (m.body or "").strip()
                for att in m.attachments or []:
                    name = getattr(att, "longFilename", None) or getattr(
                        att, "shortFilename", "attachment"
                    )
                    if name:
                        attachments.append(name)
            header_block = "\n".join(
                filter(
                    None,
                    [
                        f"Subject: {subject}" if subject else "",
                        f"From: {from_hdr}" if from_hdr else "",
                        f"To: {to_hdr}" if to_hdr else "",
                        f"Date: {date_hdr}" if date_hdr else "",
                    ],
                )
            )
            full_text = f"{header_block}\n\n{body_text}".strip() if body_text else header_block
            if attachments:
                full_text += f"\n\n[Attachments: {', '.join(attachments)}]"
            return ExtractionResult(
                kind="email",
                full_text=full_text,
                word_count=len(full_text.split()),
                pages=[PageSegment(page=1, text=full_text)],
                meta={
                    "subject": subject,
                    "from": from_hdr,
                    "to": to_hdr,
                    "date": date_hdr,
                    "attachments": attachments,
                },
            )
        except ImportError:
            logger.info("extract_msg not installed — falling back to raw read for %s", path.name)
        except Exception as exc:
            logger.warning("MSG extraction failed for %s: %s", path.name, exc)

        # Raw fallback — grab any readable text from the compound doc
        try:
            raw = path.read_bytes()
            text = raw.decode("utf-8", errors="ignore").strip()
            if len(text) > 50:
                return ExtractionResult(
                    kind="email",
                    full_text=text,
                    word_count=len(text.split()),
                    pages=[PageSegment(page=1, text=text)],
                    meta={"parse_method": "raw_fallback"},
                )
        except Exception:
            pass
        return ExtractionResult(kind="email", full_text="", word_count=0)

    return ExtractionResult(kind="email", full_text="", word_count=0)


_DISPATCH: dict[str, object] = {
    "pdf": _extract_pdf,
    "docx": _extract_docx,
    "excel": _extract_excel,
    "csv": _extract_csv,
    "text": lambda p: _extract_text(p, "text"),
    "markdown": lambda p: _extract_text(p, "markdown"),
    "html": _extract_html,
    "json": _extract_json,
    "zip": _extract_zip,
    "image": _extract_image,
    "audio": _extract_audio,
    "email": _extract_email,
}


def extract(path: str | Path, kind: str, db=None) -> ExtractionResult:
    """Extract text from *path* using the handler for *kind*.

    Falls back to markitdown for unknown kinds.  ``db`` (when supplied) is
    threaded to handlers that make LLM calls (image vision) so telemetry is
    recorded.
    """
    path = Path(path)
    handler = _DISPATCH.get(kind, lambda p: _extract_fallback(p, kind))
    try:
        if kind == "image":
            result = _extract_image(path, db=db)
        elif kind == "audio":
            result = _extract_audio(path, db=db)
        elif kind == "pdf":
            result = _extract_pdf(path, db=db)
        else:
            result = handler(path)  # type: ignore[call-arg]
    except Exception as exc:
        logger.error("Extraction error on %s (%s): %s", path.name, kind, exc)
        result = ExtractionResult(kind=kind, full_text="", word_count=0)
    logger.info("Extracted %s: %d words, %d pages", path.name, result.word_count, len(result.pages))
    return result
