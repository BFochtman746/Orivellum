"""Tax / expense package action.

Gathers all documents for a given year (optionally filtered by Work),
generates an Excel summary sheet with metadata, and bundles everything
into a downloadable zip.
"""

from __future__ import annotations

import logging
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from orivellum.capabilities.actions import ActionBase, _now

if TYPE_CHECKING:
    from orivellum.configuration.config import OrivellumConfig
    from orivellum.database.db import OrivellumDB

logger = logging.getLogger("orivellum.actions.tax_package")

_EXPENSE_KEYWORDS = {
    "receipt",
    "invoice",
    "expense",
    "payment",
    "bill",
    "purchase",
    "refund",
    "transaction",
    "charge",
    "fee",
    "cost",
    "price",
    "amount",
    "tax",
    "vat",
    "gst",
    "statement",
    "reimbursement",
}


class TaxPackageAction(ActionBase):
    name = "tax_package"
    description = (
        "Gather filed receipts and expense documents for a year, generate an Excel "
        "summary sheet with dates and descriptions, and bundle everything into a "
        "downloadable zip."
    )
    category = "finance"
    input_schema = {
        "type": "object",
        "properties": {
            "year": {"type": "integer", "description": "Calendar year, e.g. 2024"},
            "work_id": {"type": "string", "description": "Optional: limit to one Work"},
        },
        "required": ["year"],
    }

    def confirm_message(self, inputs: dict) -> str:
        year = inputs.get("year", datetime.now(UTC).year)
        work_part = ""
        if inputs.get("work_id"):
            work_part = " linked to this Work"
        return (
            f"Gather all expense and receipt documents from **{year}**{work_part}, "
            f"generate an Excel summary sheet with titles, dates, and descriptions, "
            f"and bundle the source files into a zip archive you can download."
        )

    def _execute_impl(self, inputs: dict, db: OrivellumDB, cfg: OrivellumConfig) -> dict:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        year: int = int(inputs.get("year") or datetime.now(UTC).year)
        work_id: str | None = inputs.get("work_id")
        data_dir = Path(cfg.data_dir)

        # ── Query documents matching year range ──
        year_start = f"{year}-01-01"
        year_end = f"{year}-12-31"

        with db._lock:
            if work_id:
                rows = db._conn.execute(
                    """SELECT d.id, d.title, d.kind, d.source, d.content_path,
                              d.readiness, o.created_at, d.extracted_text
                       FROM documents d
                       JOIN objects o ON o.id = d.id
                       WHERE d.work_id=?
                         AND substr(o.created_at,1,10) BETWEEN ? AND ?
                       ORDER BY o.created_at DESC""",
                    (work_id, year_start, year_end),
                ).fetchall()
            else:
                rows = db._conn.execute(
                    """SELECT d.id, d.title, d.kind, d.source, d.content_path,
                              d.readiness, o.created_at, d.extracted_text
                       FROM documents d
                       JOIN objects o ON o.id = d.id
                       WHERE substr(o.created_at,1,10) BETWEEN ? AND ?
                       ORDER BY o.created_at DESC""",
                    (year_start, year_end),
                ).fetchall()

        docs = [dict(r) for r in rows]

        # Prioritise documents that look like expense/receipt docs
        def _is_expense(doc: dict) -> bool:
            text = (doc.get("title") or "") + " " + (doc.get("extracted_text") or "")[:500]
            return any(kw in text.lower() for kw in _EXPENSE_KEYWORDS)

        expense_docs = [d for d in docs if _is_expense(d)]
        other_docs = [d for d in docs if not _is_expense(d)]
        all_docs = expense_docs + other_docs  # expense docs first

        # ── Build Excel summary ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year} Expense Summary"

        hdr_fill = PatternFill("solid", fgColor="1E293B")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        title_font = Font(bold=True, size=14)

        ws["A1"] = f"{year} Expense & Receipt Summary"
        ws["A1"].font = title_font
        ws["A2"] = f"Generated {_now()[:10]}"
        ws["A2"].font = Font(size=9, color="94A3B8")

        headers = ["#", "Title", "Kind", "Date", "Source / Description", "Readiness", "Document ID"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, doc in enumerate(all_docs, 1):
            date_str = (doc.get("created_at") or "")[:10]
            src = (doc.get("source") or "").split("/")[-1]
            ws.append(
                [
                    row_idx,
                    (doc.get("title") or "")[:80],
                    doc.get("kind", ""),
                    date_str,
                    src[:60],
                    doc.get("readiness", ""),
                    doc.get("id", ""),
                ]
            )

        # Stats row
        ws.append([])
        ws.append(["Total", len(all_docs), "", "", "", "", ""])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

        # Column widths
        for col, width in enumerate([5, 50, 10, 14, 40, 12, 38], 1):
            from openpyxl.utils import get_column_letter

            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Write files to zip ──
        out_dir = data_dir / "outputs" / "generate" / f"tax_{year}"
        out_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now(UTC).strftime("%Y%m%d_%H%M")
        xlsx_name = f"tax_summary_{year}_{ts}.xlsx"
        xlsx_path = out_dir / xlsx_name
        wb.save(str(xlsx_path))

        zip_name = f"tax_package_{year}_{ts}.zip"
        zip_path = out_dir / zip_name

        with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(str(xlsx_path), xlsx_name)
            for doc in all_docs:
                cp = doc.get("content_path")
                if not cp:
                    continue
                src_path = data_dir / cp
                if src_path.is_file():
                    arc_name = f"docs/{doc.get('kind', 'file')}/{src_path.name}"
                    try:
                        zf.write(str(src_path), arc_name)
                    except Exception as exc:
                        logger.warning("Could not add %s to zip: %s", src_path.name, exc)

        # ── Register output ──
        from orivellum.capabilities.generate import _register_output

        summary_text = "\n".join(
            f"{d.get('title', '')}: {(d.get('extracted_text') or '')[:200]}" for d in all_docs[:50]
        )
        doc_id = _register_output(
            zip_path,
            work_id or None,
            db,
            cfg,
            "zip",
            f"Tax Package {year}",
            summary_text,
        )

        rel_path = str(zip_path.relative_to(data_dir))
        return {
            "output_path": rel_path,
            "output_label": zip_name,
            "output_doc_id": doc_id,
            "document_count": len(all_docs),
            "expense_count": len(expense_docs),
            "summary": f"Tax package {year} — {len(all_docs)} documents, {len(expense_docs)} expense-matched",
        }
