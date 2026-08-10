"""Workbook build + acceptance gates for the PDF→Excel protocol v2.1.

Builds the protocol's workbook architecture (doc 06) with openpyxl:
README, Page_Register, per-page data sheets, Narrative, Exceptions,
Checks (real cross-sheet control formulas), Changelog — with the standard
color convention (yellow = source, green = formula, blue = metadata,
dark blue = headers).

Acceptance gates (deterministic, blocking — a failed gate publishes
nothing):

- G-completeness (O-4): page register rows == source page count, and every
  page with content is represented by a data or narrative row.
- G-recalc (O-2): the exported file reloads in openpyxl (value + formula
  view), contains zero stored error cells, and recalculates cleanly in an
  independent engine (``formulas``) with every cross-sheet reference
  resolving.
"""

from __future__ import annotations

import datetime
import logging
import pathlib
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover
    from orivellum.capabilities.pdf_excel import ExceptionRow, PageExtract

logger = logging.getLogger(__name__)

# Protocol color convention (doc 06)
_HEADER_FILL = "1F4E78"  # dark blue
_SOURCE_FILL = "FFF2CC"  # yellow — source values
_FORMULA_FILL = "E2EFDA"  # green — formula output
_META_FILL = "DDEBF7"  # blue — metadata / provenance

_ERROR_STRINGS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
# Hard ceilings — exceeding them raises (fail closed). The protocol forbids
# silent truncation: a partial workbook must never publish as "verified".
_MAX_TABLE_ROWS = 5000
_MAX_NARRATIVE_CHARS = 100_000


def _style_header(ws, row: int, n_cols: int) -> None:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font


def _fill_row(ws, row: int, n_cols: int, color: str) -> None:
    from openpyxl.styles import PatternFill

    fill = PatternFill("solid", fgColor=color)
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = fill


def _sheet_readme(wb, source_name: str, manifest: dict) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ("PDF→Excel Protocol", f"v{manifest['protocol_version']}"),
        ("Source file", source_name),
        ("Source SHA-256", manifest["source_sha256"]),
        ("Source size (bytes)", manifest["source_size"]),
        ("Page count", manifest["page_count"]),
        ("Source type", manifest["source_type"]),
        ("Encrypted", "yes" if manifest["encrypted"] else "no"),
        ("Risk class", manifest["risk"]["class"]),
        ("QA tier", manifest["risk"]["qa_tier"]),
        ("Generated (UTC)", datetime.datetime.now(datetime.UTC).isoformat(timespec="seconds")),
        ("Extraction", "Dual-channel (pdfplumber + pypdf); disagreements in Exceptions"),
        ("Abstention rule", "An exception row always beats a guessed value (O-3)"),
    ]
    ws.append(("Field", "Value"))
    _style_header(ws, 1, 2)
    for r in rows:
        ws.append(r)
        _fill_row(ws, ws.max_row, 2, _META_FILL)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70


def _page_status(p: PageExtract) -> str:
    if p.tables:
        return "extracted"
    if p.text_a.strip():
        return "narrative"
    return "empty"


def _sheet_page_register(wb, pages: list[PageExtract], exceptions: list[ExceptionRow]) -> None:
    ws = wb.create_sheet("Page_Register")
    headers = (
        "PDF_Page",
        "Content_Type",
        "Orientation",
        "Text_Layer_Useful",
        "Tables_Found",
        "Exception_Count",
        "Extraction_Status",
    )
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    exc_by_page: dict[int, int] = {}
    for e in exceptions:
        exc_by_page[e.pdf_page] = exc_by_page.get(e.pdf_page, 0) + 1
    for p in pages:
        status = _page_status(p)
        ws.append(
            (
                p.page,
                "table" if p.tables else ("narrative" if p.text_a.strip() else "blank"),
                "landscape" if p.width > p.height else "portrait",
                "yes" if len(p.text_a.strip()) > 20 else "no",
                len(p.tables),
                exc_by_page.get(p.page, 0),
                status,
            )
        )
        _fill_row(ws, ws.max_row, len(headers), _META_FILL)


def _sheet_page_tables(wb, p: PageExtract) -> None:
    from openpyxl.styles import PatternFill

    source_fill = PatternFill("solid", fgColor=_SOURCE_FILL)
    ws = wb.create_sheet(f"P{p.page:03d}")
    record = 1
    for t_idx, table in enumerate(p.tables, start=1):
        if len(table) > _MAX_TABLE_ROWS:
            raise ValueError(
                f"page {p.page} table {t_idx} has {len(table)} rows — the "
                f"transcription limit is {_MAX_TABLE_ROWS}; refusing to truncate"
            )
        width = max(len(r) for r in table)
        ws.append((f"Table {t_idx} — PDF page {p.page}",))
        _style_header(ws, ws.max_row, width + 3)
        ws.append(("Record_ID", "PDF_Page", "Source_Order", *[""] * width))
        _style_header(ws, ws.max_row, 3)
        for row in table:
            ws.append((f"R{p.page:03d}-{record:04d}", p.page, record, *row))
            _fill_row(ws, ws.max_row, 3, _META_FILL)
            for c in range(4, 4 + len(row)):
                ws.cell(row=ws.max_row, column=c).fill = source_fill
            record += 1
        ws.append(())


def _sheet_narrative(wb, pages: list[PageExtract]) -> None:
    rows = []
    for p in pages:
        text = p.text_a.strip()
        if not text or p.tables:
            continue
        if len(text) > _MAX_NARRATIVE_CHARS:
            raise ValueError(
                f"page {p.page} narrative is {len(text)} characters — the "
                f"transcription limit is {_MAX_NARRATIVE_CHARS}; refusing to truncate"
            )
        rows.append((p.page, text))
    if not rows:
        return
    ws = wb.create_sheet("Narrative")
    ws.append(("PDF_Page", "Text"))
    _style_header(ws, 1, 2)
    for r in rows:
        ws.append(r)
        _fill_row(ws, ws.max_row, 2, _SOURCE_FILL)
    ws.column_dimensions["B"].width = 110


def _sheet_exceptions(wb, exceptions: list[ExceptionRow]) -> None:
    ws = wb.create_sheet("Exceptions")
    headers = (
        "Exception_ID",
        "PDF_Page",
        "Exception_Type",
        "Description",
        "Channel_A_Value",
        "Channel_B_Value",
        "Disposition",
        "Status",
    )
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for e in exceptions:
        ws.append(
            (e.exception_id, e.pdf_page, e.exception_type, e.description,
             e.value_a, e.value_b, "", "open")
        )
        _fill_row(ws, ws.max_row, len(headers), _META_FILL)
    ws.column_dimensions["D"].width = 70


def _sheet_checks(wb, manifest: dict, exceptions: list[ExceptionRow]) -> None:
    """Control layer with REAL cross-sheet formulas — the recalculation
    gate (O-2) exists precisely so a renamed sheet breaks these loudly."""
    ws = wb.create_sheet("Checks")
    headers = ("Control_ID", "Description", "Observed", "Expected", "Status")
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    n = manifest["page_count"]
    checks = [
        ("C1", "Pages registered", f"=COUNTA(Page_Register!A2:A{n + 1})", n),
        (
            "C2",
            "Pages with extracted tables",
            f'=COUNTIF(Page_Register!G2:G{n + 1},"extracted")',
            manifest["table_pages"],
        ),
        (
            "C3",
            "Exceptions logged",
            f"=COUNTA(Exceptions!A2:A{len(exceptions) + 1})" if exceptions else "=0",
            len(exceptions),
        ),
    ]
    for cid, desc, observed, expected in checks:
        r = ws.max_row + 1
        ws.append((cid, desc, observed, expected, f'=IF(C{r}=D{r},"PASS","FAIL")'))
        _fill_row(ws, r, len(headers), _FORMULA_FILL)
    r = ws.max_row + 1
    open_formula = f'=COUNTIF(Exceptions!H2:H{len(exceptions) + 1},"open")' if exceptions else "=0"
    ws.append(("C4", "Open exceptions (resolve before release)", open_formula, 0,
               f'=IF(C{r}=0,"PASS","REVIEW")'))
    _fill_row(ws, r, len(headers), _FORMULA_FILL)
    ws.column_dimensions["B"].width = 44


def _sheet_changelog(wb, source_name: str) -> None:
    ws = wb.create_sheet("Changelog")
    ws.append(("Version", "Date_UTC", "Change"))
    _style_header(ws, 1, 3)
    ws.append(
        (
            "v1",
            datetime.datetime.now(datetime.UTC).date().isoformat(),
            f"Initial dual-channel transcription of {source_name}",
        )
    )
    _fill_row(ws, 2, 3, _META_FILL)


def build_workbook(
    out_path: pathlib.Path,
    source_name: str,
    manifest: dict,
    pages: list[PageExtract],
    exceptions: list[ExceptionRow],
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    _sheet_readme(wb, source_name, manifest)
    _sheet_page_register(wb, pages, exceptions)
    for p in pages:
        if p.tables:
            _sheet_page_tables(wb, p)
    _sheet_narrative(wb, pages)
    _sheet_exceptions(wb, exceptions)
    _sheet_checks(wb, manifest, exceptions)
    _sheet_changelog(wb, source_name)
    wb.save(out_path)


# ── Acceptance gates ──────────────────────────────────────────────────────────


def _fail(detail: str) -> dict:
    return {"name": "page_completeness", "ok": False, "detail": detail}


def _check_table_rows(wb, totals: dict) -> str | None:
    """Every data sheet must carry exactly the row count extraction saw."""
    for page_str, expected_rows in (totals.get("table_rows") or {}).items():
        sheet = f"P{int(page_str):03d}"
        if sheet not in wb.sheetnames:
            return f"data sheet {sheet} missing"
        prefix = f"R{int(page_str):03d}-"
        found = sum(
            1
            for row in wb[sheet].iter_rows(values_only=True)
            if row and isinstance(row[0], str) and row[0].startswith(prefix)
        )
        if found != expected_rows:
            return f"{sheet} carries {found} row(s), extraction saw {expected_rows}"
    return None


def _check_narrative(wb, totals: dict) -> str | None:
    """Every narrative page's full text must be present, uncut."""
    expected = {int(k): v for k, v in (totals.get("narrative_chars") or {}).items()}
    if not expected:
        return None
    if "Narrative" not in wb.sheetnames:
        return "Narrative sheet missing"
    found: dict[int, int] = {}
    for row in wb["Narrative"].iter_rows(min_row=2, values_only=True):
        if row and isinstance(row[0], int):
            found[row[0]] = len(str(row[1] or ""))
    for page, chars in expected.items():
        if found.get(page) != chars:
            return (
                f"narrative for page {page} carries {found.get(page, 0)} "
                f"character(s), extraction saw {chars}"
            )
    return None


def _gate_completeness(xlsx: pathlib.Path, manifest: dict) -> dict:
    """O-4: register rows == source page count; every table page has a data
    sheet; per-page row counts and narrative lengths match extraction
    exactly (no silent truncation can pass)."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True)
    try:
        if "Page_Register" not in wb.sheetnames:
            return _fail("Page_Register missing")
        ws = wb["Page_Register"]
        registered = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
        expected = manifest["page_count"]
        data_sheets = sum(1 for s in wb.sheetnames if s.startswith("P") and s[1:].isdigit())
        if registered != expected:
            return _fail(f"register has {registered} page(s), source has {expected}")
        if data_sheets != manifest["table_pages"]:
            return _fail(
                f"{data_sheets} data sheet(s) for {manifest['table_pages']} table page(s)"
            )
        totals = manifest.get("content_totals") or {}
        problem = _check_table_rows(wb, totals) or _check_narrative(wb, totals)
        if problem:
            return _fail(problem)
        return {
            "name": "page_completeness",
            "ok": True,
            "detail": f"{registered}/{expected} pages certified, content totals match",
        }
    finally:
        wb.close()


def _stored_error_cells(xlsx: pathlib.Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                count += sum(1 for v in row if isinstance(v, str) and v in _ERROR_STRINGS)
        return count
    finally:
        wb.close()


def _gate_recalc(xlsx: pathlib.Path) -> dict:
    """O-2: reload + independent recalculation, zero error cells, every
    cross-sheet reference resolves."""
    stored_errors = _stored_error_cells(xlsx)
    if stored_errors:
        return {"name": "recalc", "ok": False, "detail": f"{stored_errors} stored error cell(s)"}
    try:
        import formulas

        model = formulas.ExcelModel().loads(str(xlsx)).finish()
        solution = model.calculate()
    except Exception as exc:  # noqa: BLE001 - unresolved refs raise here
        return {"name": "recalc", "ok": False, "detail": f"independent recalculation failed: {exc}"}
    bad = 0
    for value in solution.values():
        text = str(getattr(value, "value", value))
        if any(err in text for err in _ERROR_STRINGS):
            bad += 1
    if bad:
        return {"name": "recalc", "ok": False, "detail": f"{bad} formula(s) recalculate to errors"}
    return {"name": "recalc", "ok": True, "detail": "recalculates cleanly, all references resolve"}


def run_acceptance_gates(xlsx: pathlib.Path, manifest: dict) -> dict:
    gates = [_gate_completeness(xlsx, manifest), _gate_recalc(xlsx)]
    return {"passed": all(g["ok"] for g in gates), "gates": gates}
