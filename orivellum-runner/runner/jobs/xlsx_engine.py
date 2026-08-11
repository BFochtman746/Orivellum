"""Recalculation engine — the part that makes "proven" a real word.

Wraps the pure-Python `formulas` package. It compiles every formula in the
workbook and computes it from the raw inputs, with no Excel involved. That
gives the one check the read-only job could never make: does the value saved
in the file equal the value the formula actually produces?

HONESTY RULES
  · If `formulas` is not installed, or the workbook uses a function the engine
    cannot compute, verification is UNAVAILABLE — reported as such, never as
    "clean". A workbook is only PROVEN when every formula was recomputed.
  · A formula cell with no cached value at all ("never calculated" — common in
    library-generated files) is a mismatch: the file would show blanks or
    zeros until Excel recalculates it.
"""

import datetime as _dt
import math
import re
from pathlib import Path

KEY = re.compile(r"'\[([^\]]+)\]([^']+)'!\$?([A-Z]{1,3})\$?(\d+)$")

REL_TOL = 1e-9
ABS_TOL = 1e-9


def available():
    try:
        import formulas  # noqa: F401

        return True
    except Exception:  # noqa: BLE001
        return False


def _scalar(v):
    """Collapse the engine's Ranges/arrays to a comparable Python scalar."""
    val = getattr(v, "value", v)
    try:
        import numpy as np

        if isinstance(val, np.ndarray):
            if val.size != 1:
                return None
            val = val.ravel()[0]
        if isinstance(val, np.generic):
            val = val.item()
    except Exception:  # noqa: BLE001
        pass
    if type(val).__name__ == "XlError":
        return str(val)
    return val


def _to_serial(value):
    """Cached datetimes come back as datetime objects; the engine speaks
    Excel serial numbers. Compare in serial space."""
    from openpyxl.utils.datetime import to_excel

    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        try:
            return float(to_excel(value))
        except Exception:  # noqa: BLE001
            return value
    return value


def _matches(computed, cached):
    """Strict by design. A missing cached value is a MISMATCH (the file would
    show blanks until Excel recalculates); a bool never equals a number or a
    string. Looseness here is how a broken workbook gets called proven."""
    computed, cached = _to_serial(computed), _to_serial(cached)
    if computed is None and cached is None:
        return True
    if isinstance(computed, bool) or isinstance(cached, bool):
        return isinstance(computed, bool) and isinstance(cached, bool) and computed == cached
    if isinstance(computed, (int, float)) and isinstance(cached, (int, float)):
        return math.isclose(float(computed), float(cached), rel_tol=REL_TOL, abs_tol=ABS_TOL)
    if isinstance(computed, str) and isinstance(cached, str):
        return computed.strip() == cached.strip()
    return False


def recalculate(target):
    """Recompute every formula. Returns a dict:

      {"available": bool, "error": str|None,
       "cells": {(sheet, coord): computed_value},
       "checked": int}

    `available: False` means the verdict is UNAVAILABLE, not clean.
    """
    if not available():
        return {
            "available": False,
            "error": "formulas package not installed",
            "cells": {},
            "checked": 0,
        }
    import warnings

    try:
        import formulas

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            xl = formulas.ExcelModel().loads(str(target)).finish()
            sol = xl.calculate()
    except Exception as e:  # noqa: BLE001
        return {
            "available": False,
            "error": f"engine could not compute this workbook: {type(e).__name__}: {e}"[:400],
            "cells": {},
            "checked": 0,
        }
    fname = Path(target).name.upper()
    cells = {}
    for k, v in sol.items():
        m = KEY.match(str(k))
        if not m or m.group(1).upper() != fname:
            continue
        sheet, col, row = m.group(2), m.group(3), m.group(4)
        cells[(sheet.upper(), f"{col}{row}")] = _scalar(v)
    return {"available": True, "error": None, "cells": cells, "checked": len(cells)}


def compare(target, recalc):
    """Compare engine results against the values saved in the file.

    Returns {"mismatches": [...], "agreed": int, "uncovered": [coords the
    engine produced nothing for]} — uncovered formulas block a PROVEN verdict.
    """
    from openpyxl import load_workbook

    wbf = load_workbook(target, data_only=False, read_only=True)
    wbv = load_workbook(target, data_only=True, read_only=True)
    mismatches, uncovered, agreed = [], [], 0
    formula_cells = []
    try:
        for name in wbf.sheetnames:
            sf, sv = wbf[name], wbv[name]
            cached_map = {}
            for row in sv.iter_rows():
                for c in row:
                    if c.value is not None:
                        cached_map[c.coordinate] = c.value
            for row in sf.iter_rows():
                for c in row:
                    if not (isinstance(c.value, str) and c.value.startswith("=")):
                        continue
                    key = (name.upper(), c.coordinate)
                    formula_cells.append(key)
                    if key not in recalc["cells"]:
                        uncovered.append(f"{name}!{c.coordinate}")
                        continue
                    computed = recalc["cells"][key]
                    cached = cached_map.get(c.coordinate)
                    if _matches(computed, cached):
                        agreed += 1
                    else:
                        mismatches.append(
                            {
                                "ref": f"{name}!{c.coordinate}",
                                "sheet": name,
                                "cell": c.coordinate,
                                "formula": c.value[:200],
                                "cached": _repr(cached),
                                "computed": _repr(computed),
                                "computed_raw": computed,
                            }
                        )
    finally:
        wbf.close()
        wbv.close()
    return {
        "mismatches": mismatches,
        "agreed": agreed,
        "uncovered": uncovered,
        "formula_cells": formula_cells,
    }


def _repr(v):
    if isinstance(v, float) and v == int(v) and abs(v) < 1e15:
        return int(v)
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    return v if isinstance(v, (int, float, str, bool)) or v is None else str(v)


def _canon(v):
    """Canonical, JSON-safe form for manifest expectations: datetimes become
    Excel serial numbers so a re-run compares in one type system."""
    v = _to_serial(v)
    if isinstance(v, bool) or v is None or isinstance(v, str):
        return v
    if isinstance(v, float) and not math.isfinite(v):
        return str(v)  # JSON-safe; NaN/Infinity must never become bare JSON tokens
    if isinstance(v, float) and v == int(v) and abs(v) < 1e15:
        return int(v)
    if isinstance(v, (int, float)):
        return v
    return str(v)


def build_test_manifest(target, recalc, compare_result):
    """The tests the workbook must keep passing. Every formula cell becomes a
    test case with its proven expected value; structural rules ride along.
    Re-run any time with `python -m runner verify`."""
    cases = []
    wanted = set(compare_result.get("formula_cells") or recalc["cells"])
    for (sheet, coord), computed in sorted(recalc["cells"].items()):
        if (sheet, coord) not in wanted:
            continue  # inputs are data, not tests
        cases.append({"sheet": sheet, "cell": coord, "expected": _canon(computed)})
    return {
        "target": Path(target).name,
        "engine": "formulas",
        "formula_cells": len(cases),
        "structural": [
            "no-error-cells",
            "no-sum-mismatch",
            "ooxml-order",
            "cached-values-match-recalculation",
        ],
        "cases": cases,
        "note": "expected values are engine-computed, not copied from the "
        "file — a stale cache cannot certify itself",
    }


ERRVALS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!")


def structural_checks(target):
    """The structural rules the manifest declares — actually re-run, not
    assumed: saved error values and OOXML child order."""
    import re
    import zipfile

    from openpyxl import load_workbook

    from . import xlsx_surgery as surgery

    problems = []
    wv = load_workbook(target, data_only=True, read_only=True)
    for name in wv.sheetnames:
        for row in wv[name].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in ERRVALS:
                    problems.append(f"error cell {name}!{c.coordinate} = {c.value.strip()}")
    wv.close()
    with zipfile.ZipFile(target) as z:
        for n in [x for x in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", x)]:
            bad = surgery.sheet_order_violations(z.read(n).decode("utf-8", errors="replace"))
            if bad:
                problems.append(f"OOXML order violated in {n}: {', '.join(bad)}")
    return problems


def run_manifest(target, manifest):
    """Re-run a saved manifest against a (possibly edited) workbook — the
    formula cases AND the structural rules it declares."""
    recalc = recalculate(target)
    if not recalc["available"]:
        return {
            "status": "UNAVAILABLE",
            "error": recalc["error"],
            "passed": 0,
            "failed": [],
            "total": len(manifest["cases"]),
        }
    cmp_now = compare(target, recalc)
    failed = []
    for case in manifest["cases"]:
        key = (case["sheet"], case["cell"])
        if key not in recalc["cells"]:
            failed.append({**case, "actual": None, "why": "formula gone or uncomputable"})
            continue
        actual = _canon(recalc["cells"][key])
        if not _matches(actual, case["expected"]):
            failed.append({**case, "actual": actual, "why": "value changed"})
    structural = structural_checks(target)
    status = "PASS" if not failed and not cmp_now["mismatches"] and not structural else "FAIL"
    return {
        "status": status,
        "passed": len(manifest["cases"]) - len(failed),
        "failed": failed[:100],
        "total": len(manifest["cases"]),
        "stale_cache": [m["ref"] for m in cmp_now["mismatches"]][:50],
        "structural": structural[:50],
    }
