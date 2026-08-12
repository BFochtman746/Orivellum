"""XLSX job — take a workbook apart, repair it, TEST it, and prove it.

THE RULES, VERSION 2. The old read-only doctrine is retired by request. What
replaces it is stricter, not looser:

  1. The runner MAY write a workbook — but only by XML surgery inside the zip.
     A library round-trip on a deliverable is still forbidden: openpyxl's
     writer discards external-link caches, VBA, and parts it does not model.
     `parts_diff` proves, byte for byte, that surgery touched only the parts
     it claimed.
  2. Nothing ships unproven. Every formula is recomputed by a real
     recalculation engine (the pure-Python `formulas` package) and compared
     against the value saved in the file. The output workbook is emitted ONLY
     when the full gate suite passes; otherwise there is no output file and
     the report says exactly which gate failed.
  3. The run BUILDS TESTS: a manifest of every formula cell with its
     engine-computed expected value plus the structural rules. Re-run it any
     time with `python -m runner verify` — after your own edits, on another
     machine, before shipping.
  4. If the engine is unavailable or cannot compute the workbook, the verdict
     is UNVERIFIED, stated in bold. Absence of a recalculation is never
     reported as cleanliness.

WHAT SURGERY MAY DO (mechanical, re-checkable operations only)
  · reorder <worksheet> children into the OOXML sequence iOS Excel enforces
  · refresh a formula cell's stale cached value with the recomputed one, or
    insert values where the file shipped never-calculated
Semantic edits — changing a formula, extending a range — remain proposals with
cell addresses. A machine that silently rewrites your formulas is not a
verifier, it is a liability.

PROOF GATES (all must pass before a workbook is returned)
  G1 recalculation ran and covered every formula cell
  G2 every computed value equals the saved value (after repair)
  G3 zero error values (#REF! etc.) saved in the file
  G4 OOXML child order clean in every sheet part
  G5 surgery byte-diff limited to declared sheet parts
  G6 output loads cleanly (both formula and value passes)
Doctrine findings (volatile / iOS-dynamic functions) do not block the gates —
they are design rules, not correctness — but the certificate lists them and
the verdict is downgraded to PROVEN WITH WARNINGS.
"""

import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from .. import llm, shield, store
from . import xlsx_engine as engine
from . import xlsx_formula as fx
from . import xlsx_graph as depgraph
from . import xlsx_surgery as surgery

# Formula parsing is TOKENIZED (openpyxl.formula.tokenizer via xlsx_formula) —
# the old regexes produced phantom references from string literals and were
# blind to table references and defined names. An audit tool that emits false
# findings stops being read; the tokenizer removes that entire class.
ERRVALS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!")

# Aggregations whose ranges silently mis-total over merged cells (only the
# top-left cell of a merge holds the value; the rest read as empty).
AGG_FUNCS = {"SUM", "AVERAGE", "AVERAGEA", "COUNT", "COUNTA", "MAX", "MIN", "MEDIAN", "PRODUCT", "SUBTOTAL", "SUMPRODUCT"}

# Text that LOOKS like a date, sitting in a column of real date serials.
TEXT_DATE = re.compile(r"^\s*(?:\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{2,4})\s*$")

VOLATILE = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT", "CELL", "INFO", "AREAS"}
# Ruled out for iOS-bound workbooks by doctrine
DYNAMIC = {
    "FILTER",
    "XLOOKUP",
    "XMATCH",
    "LET",
    "UNIQUE",
    "SORT",
    "SORTBY",
    "SEQUENCE",
    "VSTACK",
    "HSTACK",
    "TOCOL",
    "TOROW",
    "TAKE",
    "DROP",
    "CHOOSECOLS",
    "CHOOSEROWS",
    "TEXTSPLIT",
    "BYROW",
    "BYCOL",
    "LAMBDA",
    "REDUCE",
    "SCAN",
    "MAP",
}
SHEET_ORDER = surgery.SHEET_ORDER

# ---------------------------------------------------------------- load cache
# The old job loaded the FULL workbook twice per sheet unit — O(sheets × 2)
# full parses. One shared pair per (path, mtime) makes a 40-sheet workbook
# roughly 40× cheaper to analyse.
_WB_CACHE = {}


def _load(target):
    from openpyxl import load_workbook

    p = Path(target)
    key = (str(p), p.stat().st_mtime_ns)
    if key not in _WB_CACHE:
        _drop_cache()  # only ever one workbook in flight; close first
        _WB_CACHE[key] = (
            load_workbook(p, data_only=False, keep_vba=p.suffix.lower() == ".xlsm"),
            load_workbook(p, data_only=True),
        )
    return _WB_CACHE[key]


def _drop_cache():
    for wbf, wbv in _WB_CACHE.values():
        try:
            wbf.close()
            wbv.close()
        except Exception:
            pass  # noqa: BLE001
    _WB_CACHE.clear()
    _GRAPH_CACHE.clear()


# One dependency graph per (path, mtime) — shared by every sheet unit and the
# workbook unit so tracing and cycle detection never re-parse the formulas.
_GRAPH_CACHE = {}


def _graph(target):
    p = Path(target)
    key = (str(p), p.stat().st_mtime_ns)
    if key not in _GRAPH_CACHE:
        wbf, wbv = _load(target)
        _GRAPH_CACHE.clear()
        g = depgraph.WorkbookGraph.build(wbf, wbv)
        # workbook-wide error map, for tracing an error chain to its root
        g.error_cells = {}
        for ws in wbv.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() in ERRVALS:
                        g.error_cells[f"{ws.title}!{c.coordinate}"] = c.value.strip()
        _GRAPH_CACHE[key] = g
    return _GRAPH_CACHE[key]


def _cols_to_num(col):
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def normalize(formula, row, col):
    """Turn A1 references into offsets so sibling formulas can be compared.
    =B2+C2 in row 2 and =B3+C3 in row 3 normalise to the same shape.
    Tokenized: cell-like text inside string literals is never rewritten."""
    return fx.normalize(formula or "", row, col)


def plan(target, run_dir):
    """One unit per SHEET, one workbook unit (raw OOXML), then one PROVE unit
    that repairs, recalculates, gates, and emits — always last."""
    from openpyxl import load_workbook

    p = Path(target)
    wbf = load_workbook(p, read_only=True, data_only=False)
    sheets = wbf.sheetnames
    meta = {
        "sheets": sheets,
        "vba": p.suffix.lower() == ".xlsm",
        "engine": "formulas" if engine.available() else "UNAVAILABLE",
    }
    wbf.close()
    units = [{"kind": "sheet", "ref": s, "payload": {"sheet": s, "target": str(p)}} for s in sheets]
    units.append({"kind": "workbook", "ref": "(workbook)", "payload": {"target": str(p)}})
    units.append({"kind": "prove", "ref": "(prove)", "payload": {"target": str(p)}})
    unavailable = [] if engine.available() else ["formulas (recalculation engine)"]
    return {
        "root": str(p),
        "units": units,
        "meta": meta,
        "unavailable": unavailable,
        "note": "writes only a PROVEN copy under runs/<id>/ — the input file "
        "is never modified in place",
    }


def sheet_unit(run_id, payload):
    name = payload["sheet"]
    target = payload["target"]
    wbf, wbv = _load(target)
    graph = _graph(target)
    sf, sv = wbf[name], wbv[name]
    formulas, blocks, funcs = {}, defaultdict(list), Counter()
    skeletons = defaultdict(list)  # structural identity for anchor checks
    errors, hardcoded, ext_refs = [], [], []
    inputs = 0

    for row in sf.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                formulas[c.coordinate] = v
                facts = graph.facts.get(f"{name}!{c.coordinate}") or fx.analyze(
                    v,
                    name,
                    graph.names,
                    graph.tables,
                    cell=(c.column, c.row),
                    sheets=wbf.sheetnames,
                )
                for fn in facts["functions"]:
                    funcs[fn] += 1
                if facts["external"]:
                    ext_refs.append(c.coordinate)
                # numeric literals used as operands — tokenized, so numbers in
                # string text or array constants never count as "magic"
                lits = [x for x in facts["literals"] if float(x) not in (0, 1, 2, 100)]
                if lits:
                    hardcoded.append((c.coordinate, lits[:3]))
                shape = normalize(v, c.row, c.column)
                blocks[("row", c.row, shape)].append(c.coordinate)
                blocks[("col", c.column, shape)].append(c.coordinate)
                sig = fx.anchor_signature(v)
                if sig:
                    skel = fx.skeleton(v)
                    skeletons[("row", c.row, skel)].append((c.coordinate, sig))
                    skeletons[("col", c.column, skel)].append((c.coordinate, sig))
            else:
                inputs += 1
    # cached values: errors and SUM consistency
    for row in sv.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in ERRVALS:
                errors.append((c.coordinate, c.value.strip()))

    # row/col-consistency: a shape with a single member alongside a dominant
    # shape is a suspect edited cell
    suspects, seen_suspect = [], set()
    groups = defaultdict(list)
    for (axis, idx, shape), coords in blocks.items():
        groups[(axis, idx)].append((shape, coords))
    for (axis, idx), shapes in groups.items():
        if len(shapes) < 2:
            continue
        shapes.sort(key=lambda s: -len(s[1]))
        dom_coords = shapes[0][1]
        if len(dom_coords) < 3:
            continue
        for shape, coords in shapes[1:]:
            if len(coords) == 1 and coords[0] not in seen_suspect:
                seen_suspect.add(coords[0])
                suspects.append((coords[0], len(dom_coords), axis))

    # anchor drift: structurally identical formulas in one row/col whose
    # $-anchoring differs — the classic fill-without-anchoring slip
    anchor_drift, seen_drift = [], set()
    for (axis, _idx, _skel), members in skeletons.items():
        if len(members) < 3:
            continue
        by_sig = Counter(sig for _, sig in members)
        if len(by_sig) < 2:
            continue
        dom_sig, dom_n = by_sig.most_common(1)[0]
        if dom_n < 3:
            continue
        for coord, sig in members:
            if sig != dom_sig and by_sig[sig] == 1 and coord not in seen_drift:
                seen_drift.add(coord)
                anchor_drift.append((coord, dom_n, axis))

    # merged cells inside an aggregated range: only the top-left of a merge
    # holds the value, the rest read as empty — a silent wrong total
    merged_in_range, seen_merge = [], set()
    merged_ranges = list(getattr(sf, "merged_cells", None) and sf.merged_cells.ranges or [])
    if merged_ranges:
        for coord, f in formulas.items():
            facts = graph.facts.get(f"{name}!{coord}")
            if not facts or not (facts["functions"] & AGG_FUNCS):
                continue
            for asheet, c1, r1, c2, r2 in facts["areas"]:
                if str(asheet).upper() != name.upper():
                    continue
                if (c2 - c1 + 1) * (r2 - r1 + 1) < 2:
                    continue
                for m in merged_ranges:
                    # any merge overlapping the area beyond its own top-left
                    if m.min_col > c2 or m.max_col < c1 or m.min_row > r2 or m.max_row < r1:
                        continue
                    span = (m.max_col - m.min_col + 1) * (m.max_row - m.min_row + 1)
                    if span > 1 and coord not in seen_merge:
                        seen_merge.add(coord)
                        merged_in_range.append((coord, f, str(m)))
                        break

    # date-serial vs text-date mixing within one column: half the column
    # compares as numbers, half as strings — sorts and lookups silently split
    date_mix = []
    col_kinds = defaultdict(lambda: [0, 0, None, None])  # [dates, textdates, ex_date, ex_text]
    import datetime as _dt

    for row in sv.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, (_dt.datetime, _dt.date)):
                k = col_kinds[c.column_letter]
                k[0] += 1
                k[2] = k[2] or c.coordinate
            elif isinstance(v, str) and TEXT_DATE.match(v):
                k = col_kinds[c.column_letter]
                k[1] += 1
                k[3] = k[3] or c.coordinate
    for col, (n_date, n_text, ex_date, ex_text) in sorted(col_kinds.items()):
        if n_date >= 2 and n_text >= 1:
            date_mix.append((col, n_date, n_text, ex_date, ex_text))

    # SUM consistency using cached values
    sum_checks = []
    for coord, f in formulas.items():
        m = re.fullmatch(r"=\s*SUM\(\s*([A-Z]{1,3}\d+)\s*:\s*([A-Z]{1,3}\d+)\s*\)\s*", f, re.I)
        if not m:
            continue
        try:
            cached = sv[coord].value
            if not isinstance(cached, (int, float)):
                continue
            total = 0.0
            seen = 0
            for rr in sv[f"{m.group(1)}:{m.group(2)}"]:
                for cc in rr:
                    if isinstance(cc.value, (int, float)):
                        total += cc.value
                        seen += 1
            if seen and abs(total - cached) > max(0.01, abs(cached) * 1e-6):
                sum_checks.append(
                    {
                        "cell": coord,
                        "formula": f,
                        "cached": cached,
                        "computed": round(total, 4),
                        "cells": seen,
                    }
                )
        except Exception:
            continue

    short_ranges = []
    for coord, f in formulas.items():
        m = re.fullmatch(
            r"=\s*(SUM|AVERAGE|COUNT|MAX|MIN)\((\$?[A-Z]{1,3})(\$?\d+):(\$?[A-Z]{1,3})(\$?\d+)\)\s*",
            f,
            re.I,
        )
        if not m:
            continue
        c1, _r1, c2, r2 = (
            m.group(2).replace("$", ""),
            int(m.group(3).replace("$", "")),
            m.group(4).replace("$", ""),
            int(m.group(5).replace("$", "")),
        )
        if c1 != c2:
            continue  # vertical ranges only
        nxt = r2 + 1
        try:
            below = sv[f"{c1}{nxt}"].value
            in_own_col = sf[coord].column == _cols_to_num(c1)
            if isinstance(below, (int, float)) and not (in_own_col and nxt == sf[coord].row):
                short_ranges.append((coord, f, f"{c1}{nxt}", below))
        except Exception:
            pass
    for coord, f, nextcell, val in short_ranges[:60]:
        store.add_finding(
            run_id,
            "HIGH",
            "XL-SHORTRANGE",
            f"{name}!{coord}",
            f"Range stops one row short of live data ({nextcell} = {val})",
            detail=f"{f} — the cell immediately below the range holds a number",
            source="range-extent",
            fix="Confirm the range is meant to exclude it. A total that stops "
            "short is self-consistent, so no value check will ever catch it. "
            "Semantic fix — stays a proposal, never auto-applied.",
        )

    upper = {k.upper() for k in funcs}
    vol = sorted(upper & VOLATILE)
    dyn = sorted(upper & DYNAMIC)

    # findings
    for coord, err in errors[:200]:
        key = f"{name}!{coord}"
        chain = graph.trace_error_root(key, getattr(graph, "error_cells", {}))
        trace = graph.trace_precedents(key, depth=3)
        detail_parts = []
        if len(chain) > 1:
            detail_parts.append(
                "error chain: " + " ← ".join(chain) + f" (root: {chain[-1]})"
            )
        if trace:
            detail_parts.append("precedents: " + " ".join(trace))
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-ERRCELL",
            key,
            f"Error value baked into the file: {err}",
            detail=" | ".join(detail_parts)[:800],
            source="cached-value",
            fix=(
                f"Fix the root precedent ({chain[-1]})."
                if len(chain) > 1
                else "Fix the precedent."
            )
            + " The prove pass refreshes the cache only "
            "when the recalculated cell no longer errors.",
        )
    for s in sum_checks[:100]:
        trace = graph.trace_precedents(f"{name}!{s['cell']}", depth=2)
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-SUMMISMATCH",
            f"{name}!{s['cell']}",
            f"SUM says {s['cached']} but its own range totals {s['computed']}",
            detail=f"{s['formula']} over {s['cells']} numeric cells"
            + (f" | precedents: {' '.join(trace)}"[:400] if trace else ""),
            source="value-check",
            fix="Stale cache or a typed-over cell. The prove pass recomputes "
            "it and repairs the cached value by XML surgery.",
        )
    for coord, n, axis in suspects[:100]:
        store.add_finding(
            run_id,
            "HIGH",
            "XL-INCONSISTENT",
            f"{name}!{coord}",
            f"Lone formula shape in a {axis} of identical formulas",
            detail=f"{n} sibling cells along that {axis} share one shape; this cell differs",
            source="shape-analysis",
            fix="Confirm it is deliberate. Semantic — never auto-applied.",
        )
    for coord, n, axis in anchor_drift[:60]:
        store.add_finding(
            run_id,
            "MEDIUM",
            "XL-ANCHOR-DRIFT",
            f"{name}!{coord}",
            f"$-anchoring differs from {n} structurally identical formulas in the same {axis}",
            detail=f"formula: {formulas.get(coord, '')[:150]}",
            source="shape-analysis",
            fix="A fill that should have used (or dropped) $ anchors. Confirm "
            "which target the formula is meant to track. Semantic — never auto-applied.",
        )
    for coord, f, merge in merged_in_range[:60]:
        store.add_finding(
            run_id,
            "HIGH",
            "XL-MERGED-RANGE",
            f"{name}!{coord}",
            f"Aggregation over a range containing merged cells ({merge})",
            detail=f"{f[:150]} — only the top-left cell of a merge holds a value; "
            "the rest read as empty",
            source="graph-analysis",
            fix="Unmerge the cells inside the summed range, or exclude the "
            "merged block. A silent wrong total that no value check catches.",
        )
    for col, n_date, n_text, ex_date, ex_text in date_mix[:40]:
        store.add_finding(
            run_id,
            "MEDIUM",
            "XL-DATE-MIX",
            f"{name}!{col}:{col}",
            f"Column {col} mixes {n_date} real date(s) with {n_text} text date(s)",
            detail=f"date serial at {ex_date}, text at {ex_text} — sorts, lookups "
            "and comparisons treat the two as different types",
            source="type-analysis",
            fix="Convert the text dates to real dates (Data → Text to Columns, "
            "or DATEVALUE). Semantic — never auto-applied.",
        )
    if vol:
        store.add_finding(
            run_id,
            "HIGH",
            "XL-VOLATILE",
            name,
            f"Volatile functions present: {', '.join(vol)}",
            source="doctrine",
            fix="Zero-volatile doctrine — CHOOSE-based defined names replace "
            "INDIRECT. Downgrades the verdict to PROVEN WITH WARNINGS.",
        )
    if dyn:
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-IOS-DYNAMIC",
            name,
            f"Dynamic array functions present: {', '.join(dyn)}",
            source="doctrine",
            fix="Ruled out for iOS-bound workbooks. Use INDEX/MATCH.",
        )
    for coord in ext_refs[:40]:
        store.add_finding(
            run_id,
            "HIGH",
            "XL-EXTLINK",
            f"{name}!{coord}",
            "Formula references another workbook",
            source="analysis",
            fix="The cached value is the only thing holding that data. "
            "Copy the values in, or keep the source file with it.",
        )
    for coord, lits in hardcoded[:60]:
        store.add_finding(
            run_id,
            "MEDIUM",
            "XL-MAGIC",
            f"{name}!{coord}",
            f"Hardcoded number(s) inside a formula: {', '.join(lits)}",
            source="analysis",
            fix="Move each to a labelled assumption cell and reference it.",
        )
    if not formulas and inputs:
        store.add_finding(
            run_id,
            "INFO",
            "XL-DATAONLY",
            name,
            f"No formulas — {inputs} static cells (a data sheet)",
            source="analysis",
        )

    digest = {
        "sheet": name,
        "formulas": len(formulas),
        "static_cells": inputs,
        "error_cells": len(errors),
        "sum_mismatches": len(sum_checks),
        "inconsistent": len(suspects),
        "anchor_drift": len(anchor_drift),
        "merged_in_range": len(merged_in_range),
        "date_mix_columns": len(date_mix),
        "volatile": vol,
        "dynamic": dyn,
        "external_refs": len(ext_refs),
        "hardcoded": len(hardcoded),
        "short_ranges": len(short_ranges),
        "top_functions": funcs.most_common(8),
    }

    labels = [
        str(c.value)[:80]
        for r in sv.iter_rows(max_row=40)
        for c in r
        if isinstance(c.value, str) and len(str(c.value)) > 3
    ][:60]
    hits = shield.screen(" \n".join(labels), where=name)
    for h in hits:
        store.add_finding(
            run_id,
            "HIGH",
            "INJECT-CELL",
            name,
            f"Injection-shaped text in cell content: {h['kind']}",
            detail=h["match"],
            source="shield",
        )
    out = llm.as_json(
        llm.chat(
            "You describe ONE spreadsheet sheet for an engineer auditing a workbook. "
            "Reply as JSON: purpose (one sentence), role (input|calculation|output|reference), "
            "risks (max 3). Use only the evidence given. Say so if you cannot tell.",
            shield.wrap(
                f"Sheet: {name}\nFormula count: {len(formulas)}\n"
                f"Functions: {dict(funcs.most_common(10))}\n"
                f"Sample labels: {labels[:25]}",
                name,
            ),
            max_tokens=300,
        )
    )
    if out:
        digest.update({k: out.get(k) for k in ("purpose", "role", "risks")})
        digest["by"] = "model"
    else:
        digest["by"] = "structure-only"
    return digest


def _graph_audit(run_id, target):
    """Workbook-level dependency-graph checks. One graph, four audits:
    cycles, orphan formulas, unread numeric inputs, defined-name hygiene."""
    graph = _graph(target)
    _wbf, wbv = _load(target)
    digest = {"graph": graph.stats()}

    # circular references — detected as SCCs, every member named. The count
    # is ALWAYS complete; only the per-cycle findings are capped, and any
    # truncation is disclosed explicitly — never silently undercounted.
    cycles = graph.cycles()
    digest["circular"] = len(cycles)
    _CYCLE_FINDING_CAP = 20
    if len(cycles) > _CYCLE_FINDING_CAP:
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-CIRCULAR",
            "workbook",
            f"Circular references — {len(cycles)} independent cycles found",
            detail=(
                f"Only the first {_CYCLE_FINDING_CAP} cycles are itemized below; "
                f"{len(cycles) - _CYCLE_FINDING_CAP} more exist. Affected cells "
                "(first of each remaining cycle): "
                + ", ".join(c["members"][0] for c in cycles[_CYCLE_FINDING_CAP:])
            )[:800],
            source="graph-analysis",
            fix="Every cycle must be broken; the total count above is complete "
            "even though only the first cycles are itemized.",
        )
    for cyc in cycles[:_CYCLE_FINDING_CAP]:
        members, loop_cells = cyc["members"], cyc["loop"]
        loop = " → ".join(loop_cells + [loop_cells[0]])
        detail = f"cycle: {loop}"
        if len(members) > len(loop_cells):
            detail += f"; all {len(members)} affected cells: " + ", ".join(members)
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-CIRCULAR",
            members[0],
            f"Circular reference — {len(members)} cell(s) depend on themselves",
            detail=detail[:800],
            source="graph-analysis",
            fix="Break the loop: one of these formulas must take its input "
            "from a cell outside the cycle. With iterative calculation off, "
            "Excel shows 0; with it on, a plausible wrong number — forever.",
        )

    # honesty note: dynamic references and unresolvable names mean the graph
    # is a lower bound — say so instead of pretending completeness
    partial_bits = []
    if graph.computed_ref_cells:
        digest["computed_ref_cells"] = graph.computed_ref_cells[:20]
        partial_bits.append(
            f"{len(graph.computed_ref_cells)} formula(s) compute references at "
            "run time (INDIRECT/OFFSET): " + ", ".join(graph.computed_ref_cells[:10])
        )
    if graph.unresolved_names:
        digest["unresolved_names"] = sorted(graph.unresolved_names)[:20]
        partial_bits.append(
            f"{len(graph.unresolved_names)} name(s)/ref(s) with no resolvable "
            "cell destination (constants, formula-defined names): "
            + ", ".join(sorted(graph.unresolved_names)[:10])
        )
    if partial_bits:
        store.add_finding(
            run_id,
            "INFO",
            "XL-GRAPH-PARTIAL",
            "(workbook)",
            "The dependency graph is a lower bound — some references are not "
            "statically resolvable",
            detail=" | ".join(partial_bits)[:800],
            source="graph-analysis",
            fix="Cycle and orphan results cover only statically visible edges. "
            "Replace INDIRECT with CHOOSE-based names to make the graph complete.",
        )

    # orphans: formulas nothing depends on (terminal outputs land here too —
    # an observation to review, not a defect) and numeric inputs nothing reads
    orphans = graph.orphan_formulas()
    digest["orphan_formulas"] = len(orphans)
    unread = graph.unread_inputs(wbv)
    total_unread = sum(len(v) for v in unread.values())
    digest["unread_inputs"] = total_unread
    if total_unread:
        sample = [c for cells in unread.values() for c in cells][:12]
        store.add_finding(
            run_id,
            "INFO",
            "XL-UNREAD-INPUT",
            "(workbook)",
            f"{total_unread} numeric input cell(s) that no formula reads",
            detail="e.g. " + ", ".join(sample),
            source="graph-analysis",
            fix="Dead data, or a range that should include them and does not "
            "— cross-check against any XL-SHORTRANGE findings.",
        )

    # defined-name hygiene
    orphan_names, shadow_names = graph.name_audit()
    if orphan_names:
        store.add_finding(
            run_id,
            "LOW",
            "XL-NAME-ORPHAN",
            "(workbook)",
            f"{len(orphan_names)} defined name(s) never referenced by any formula",
            detail=", ".join(
                f"{n} (sheet: {s})" if s else n for n, s in orphan_names[:15]
            ),
            source="graph-analysis",
            fix="Delete unused names, or wire them in — a stale name that "
            "points at moved data is a latent wrong reference.",
        )
    for n, s in shadow_names[:15]:
        store.add_finding(
            run_id,
            "MEDIUM",
            "XL-NAME-SHADOW",
            f"{s}!{n}",
            f"Sheet-scoped name '{n}' shadows a workbook-scoped name",
            detail=f"formulas on '{s}' resolve {n} differently from every other sheet",
            source="graph-analysis",
            fix="Rename one of the two. Same spelling, different target, "
            "per-sheet resolution — a classic silent divergence.",
        )
    return digest


def workbook_unit(run_id, payload):
    """Raw OOXML checks (element ordering is invisible to openpyxl) plus the
    workbook-level graph audit: circular references, orphans, name hygiene."""
    target = payload["target"]
    digest = {"ooxml": {}, "order_violations": []}
    digest.update(_graph_audit(run_id, target))
    try:
        with zipfile.ZipFile(target) as z:
            names = z.namelist()
            digest["ooxml"]["parts"] = len(names)
            digest["ooxml"]["has_vba"] = any(n.endswith("vbaProject.bin") for n in names)
            digest["ooxml"]["external_links"] = sum(1 for n in names if "externalLink" in n)
            for n in [x for x in names if re.match(r"xl/worksheets/sheet\d+\.xml$", x)]:
                xml = z.read(n).decode("utf-8", errors="replace")
                bad = surgery.sheet_order_violations(xml)
                if bad:
                    digest["order_violations"].append({"part": n, "elements": bad})
                    store.add_finding(
                        run_id,
                        "CRITICAL",
                        "XL-OOXML-ORDER",
                        n,
                        f"OOXML child order violated: {', '.join(bad)}",
                        detail="Required order: sheetData → mergeCells → conditionalFormatting "
                        "→ dataValidations → pageMargins",
                        source="raw-xml",
                        fix="iOS Excel Mobile enforces this ordering strictly. The prove "
                        "pass repairs it by XML surgery and re-checks the result.",
                    )
            if digest["ooxml"]["external_links"]:
                store.add_finding(
                    run_id,
                    "HIGH",
                    "XL-EXTPARTS",
                    "(workbook)",
                    f"{digest['ooxml']['external_links']} external link part(s) present",
                    source="raw-xml",
                    fix="Surgery copies these parts through byte-identical; a library "
                    "round-trip would strip their cached values. Never round-trip.",
                )
            if digest["ooxml"]["has_vba"]:
                store.add_finding(
                    run_id,
                    "MEDIUM",
                    "XL-VBA",
                    "(workbook)",
                    "Workbook contains a VBA project",
                    source="raw-xml",
                    fix="Macros are code. They need the same review as the rest; surgery "
                    "preserves the project byte-for-byte.",
                )
    except Exception as e:  # noqa: BLE001
        digest["ooxml"]["error"] = str(e)[:200]
    return digest


# ------------------------------------------------------------------ PROVE
def prove_unit(run_id, payload):
    """Repair → recalculate → gate → emit. The unit that earns the word
    'proven' — or explains exactly why it cannot."""
    from openpyxl import load_workbook

    from ..config import CFG

    target = Path(payload["target"])
    _drop_cache()
    run_dir = Path(CFG.runs_dir) / str(run_id)
    run_dir.mkdir(parents=True, exist_ok=True)
    gates = {}
    digest = {"verdict": "UNVERIFIED", "gates": gates, "repairs": [], "output": None}

    recalc = engine.recalculate(target)
    if not recalc["available"]:
        digest["engine_error"] = recalc["error"]
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-UNVERIFIED",
            "(workbook)",
            "Recalculation unavailable — the workbook CANNOT be proven",
            detail=recalc["error"] or "",
            source="engine",
            fix="Install the `formulas` package, or remove the function the "
            "engine cannot compute. No proof, no shipped workbook.",
        )
        return digest

    # ---- plan repairs -----------------------------------------------------
    cmp0 = engine.compare(target, recalc)
    part_of = surgery.sheet_part_names(target)
    edits, repairs = {}, []
    with zipfile.ZipFile(target) as z:
        raw = {
            n: z.read(n).decode("utf-8", errors="replace")
            for n in z.namelist()
            if re.match(r"xl/worksheets/sheet\d+\.xml$", n)
        }

    # R1: OOXML child order, per part. reorder_sheet_xml REFUSES (returns the
    # input) on content it cannot restructure safely — the violation then
    # stands, gate G4 fails, and the report says exactly that.
    for part, xml in raw.items():
        bad = surgery.sheet_order_violations(xml)
        if bad:
            fixed = surgery.reorder_sheet_xml(xml)
            if fixed == xml:
                repairs.append(
                    {
                        "op": "reorder-refused",
                        "part": part,
                        "elements": bad,
                        "why": "unknown top-level content — surgery "
                        "never restructures what it cannot prove",
                    }
                )
            else:
                raw[part] = fixed
                edits[part] = fixed
                repairs.append({"op": "reorder", "part": part, "elements": bad})

    # R2: stale / missing cached values, from the recalculation
    by_sheet = defaultdict(dict)
    for m in cmp0["mismatches"]:
        computed = m["computed_raw"]
        if isinstance(computed, str) and computed.startswith("#"):
            continue  # a formula that genuinely errors is not repairable
        by_sheet[m["sheet"]][m["cell"]] = computed
    for sheet, updates in by_sheet.items():
        part = part_of.get(sheet)
        if not part or part not in raw:
            continue
        new_xml, changed = surgery.refresh_cached(raw[part], updates)
        if changed:
            raw[part] = new_xml
            edits[part] = new_xml
            repairs.append({"op": "refresh-cache", "part": part, "sheet": sheet, "cells": changed})

    digest["repairs"] = repairs
    # The candidate NEVER carries the PROVEN name until every gate has passed.
    # A crash mid-gate must not leave a file that looks certified.
    out_path = run_dir / f"PROVEN_{target.name}"
    candidate = run_dir / f".candidate_{target.name}"
    out_path.unlink(missing_ok=True)  # stale output from a prior run
    passed = False
    try:
        touched = surgery.apply(target, candidate, edits)

        # ---- gate suite, run against the CANDIDATE ------------------------
        recalc2 = engine.recalculate(candidate)
        cmp2 = (
            engine.compare(candidate, recalc2)
            if recalc2["available"]
            else {"mismatches": [], "agreed": 0, "uncovered": ["engine failed on output"]}
        )
        gates["G1_recalc_covers_all"] = recalc2["available"] and not cmp2["uncovered"]
        gates["G2_values_match"] = recalc2["available"] and not cmp2["mismatches"]

        err_cells = []
        wv = load_workbook(candidate, data_only=True, read_only=True)
        for name in wv.sheetnames:
            for row in wv[name].iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() in ERRVALS:
                        err_cells.append(f"{name}!{c.coordinate}")
        wv.close()
        gates["G3_no_error_cells"] = not err_cells

        order_bad = []
        with zipfile.ZipFile(candidate) as z:
            for n in [x for x in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", x)]:
                if surgery.sheet_order_violations(z.read(n).decode("utf-8", errors="replace")):
                    order_bad.append(n)
        gates["G4_ooxml_order"] = not order_bad

        diff = surgery.parts_diff(target, candidate)
        gates["G5_surgery_contained"] = (
            not diff["added"] and not diff["removed"] and set(diff["changed"]) <= set(touched)
        )
        try:
            load_workbook(candidate, data_only=False).close()
            load_workbook(candidate, data_only=True).close()
            gates["G6_loads_clean"] = True
        except Exception as e:  # noqa: BLE001
            gates["G6_loads_clean"] = False
            digest["load_error"] = str(e)[:200]

        passed = all(gates.values())
        if passed:
            candidate.replace(out_path)  # atomic promotion, gates first
    finally:
        candidate.unlink(missing_ok=True)  # no gate suite, no candidate left

    f = store.findings(run_id)
    doctrine = sorted({x["code"] for x in f if x["code"] in ("XL-VOLATILE", "XL-IOS-DYNAMIC")})
    if passed:
        digest["verdict"] = "PROVEN WITH WARNINGS" if doctrine else "PROVEN"
        digest["doctrine_warnings"] = doctrine
        digest["output"] = str(out_path)
        manifest = engine.build_test_manifest(out_path, recalc2, cmp2)
        mpath = run_dir / "workbook_tests.json"
        mpath.write_text(json.dumps(manifest, indent=1))
        digest["tests"] = {"path": str(mpath), "cases": manifest["formula_cells"]}
    else:
        digest["verdict"] = "FAILED PROOF"
        digest["failed_gates"] = sorted(k for k, v in gates.items() if not v)
        detail = []
        if cmp2.get("uncovered"):
            detail.append(f"uncovered: {cmp2['uncovered'][:5]}")
        if cmp2.get("mismatches"):
            try:
                g = _graph(target)
                traced = []
                for m in cmp2["mismatches"][:5]:
                    lines = g.trace_precedents(m["ref"], depth=2, width=4)
                    traced.append(
                        f"{m['ref']} cached={m['cached']} computed={m['computed']}"
                        + (f" [{' '.join(lines)}]" if lines else "")
                    )
                detail.append("mismatches: " + "; ".join(traced))
            except Exception:  # noqa: BLE001 — tracing never blocks the verdict
                detail.append(
                    "mismatches: "
                    + "; ".join(
                        f"{m['ref']} cached={m['cached']} computed={m['computed']}"
                        for m in cmp2["mismatches"][:5]
                    )
                )
        if err_cells:
            detail.append(f"error cells: {err_cells[:5]}")
        if order_bad:
            detail.append(f"order: {order_bad}")
        store.add_finding(
            run_id,
            "CRITICAL",
            "XL-PROOF-FAILED",
            "(workbook)",
            f"Proof failed at: {', '.join(digest['failed_gates'])}",
            detail=" | ".join(detail)[:800],
            source="prove",
            fix="Fix the listed cells (formulas that genuinely error cannot be "
            "cache-repaired) and rerun. No workbook is returned unproven.",
        )
    digest["recalc"] = {
        "formulas_checked": len(cmp2.get("formula_cells", [])) or recalc2.get("checked", 0),
        "agreed": cmp2.get("agreed", 0),
        "repaired_cells": sum(
            len(r.get("cells", [])) for r in repairs if r["op"] == "refresh-cache"
        ),
    }
    return digest


def unit_worker(run_id, unit):
    if unit["kind"] == "sheet":
        return sheet_unit(run_id, unit["payload"])
    if unit["kind"] == "prove":
        return prove_unit(run_id, unit["payload"])
    return workbook_unit(run_id, unit["payload"])


def final_pass(run_id):
    _drop_cache()
    ds = store.digests(run_id)
    sheets = [d["digest"] for d in ds if d["kind"] == "sheet"]
    proofs = [d["digest"] for d in ds if d["kind"] == "prove"]
    tot = lambda k: sum(s.get(k) or 0 for s in sheets)
    rows = [
        "| Sheet | Formulas | Static | Errors | SUM mismatch | Short range | Odd formula | Volatile |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for s in sheets:
        rows.append(
            f"| {s['sheet']} | {s['formulas']} | {s['static_cells']} | "
            f"{s['error_cells']} | {s['sum_mismatches']} | "
            f"{s.get('short_ranges', 0)} | {s['inconsistent']} | "
            f"{', '.join(s['volatile']) or '—'} |"
        )
    wbs = [d["digest"] for d in ds if d["kind"] == "workbook"]
    verdict = []
    if wbs and wbs[0].get("circular"):
        verdict.append(
            f"**{wbs[0]['circular']} circular reference chain(s)** — the one defect "
            "that produces a plausible wrong number forever."
        )
    if tot("error_cells"):
        verdict.append(f"**{tot('error_cells')} error cells** are saved in the file.")
    if tot("sum_mismatches"):
        verdict.append(
            f"**{tot('sum_mismatches')} totals disagree with their own ranges** — this is the "
            "one that reports wrong numbers without ever looking broken."
        )
    if tot("short_ranges"):
        verdict.append(f"{tot('short_ranges')} total(s) stop one row short of live data.")
    if tot("inconsistent"):
        verdict.append(f"{tot('inconsistent')} lone formulas sit inside rows of identical ones.")
    if tot("merged_in_range"):
        verdict.append(
            f"{tot('merged_in_range')} aggregation(s) run over merged cells — silent under-totals."
        )
    if tot("anchor_drift"):
        verdict.append(f"{tot('anchor_drift')} formula(s) anchored differently from their siblings.")
    if tot("date_mix_columns"):
        verdict.append(f"{tot('date_mix_columns')} column(s) mix real dates with text dates.")
    if not verdict:
        verdict.append("No error cells, no total mismatches, no odd formulas found.")

    graph_lines = []
    if wbs:
        gs = wbs[0].get("graph") or {}
        if gs:
            graph_lines.append(
                f"- {gs.get('formula_cells', 0)} formula cells, "
                f"{gs.get('edges', 0)} dependency edges, "
                f"{gs.get('referenced_cells', 0)} cells read by formulas"
            )
            graph_lines.append(
                f"- {gs.get('defined_names', 0)} defined name(s), "
                f"{gs.get('tables', 0)} table(s) resolved into the graph"
            )
            graph_lines.append(
                f"- {wbs[0].get('circular', 0)} circular chain(s), "
                f"{wbs[0].get('orphan_formulas', 0)} formula(s) nothing depends on, "
                f"{wbs[0].get('unread_inputs', 0)} numeric input(s) no formula reads"
            )
            if gs.get("computed_ref_cells"):
                graph_lines.append(
                    f"- {gs['computed_ref_cells']} cell(s) compute references at run "
                    "time (INDIRECT/OFFSET) — the graph is a lower bound there"
                )

    proof_lines = []
    if proofs:
        p = proofs[0]
        proof_lines.append(f"**Verdict: {p.get('verdict', 'UNVERIFIED')}**")
        for g, ok in (p.get("gates") or {}).items():
            proof_lines.append(f"- {'PASS' if ok else 'FAIL'} — {g}")
        rc = p.get("recalc") or {}
        if rc:
            proof_lines.append(
                f"- {rc.get('formulas_checked', 0)} formulas recomputed, "
                f"{rc.get('agreed', 0)} agree with the saved values, "
                f"{rc.get('repaired_cells', 0)} cached values repaired by surgery"
            )
        for r in p.get("repairs", []):
            what = (
                f"reordered {', '.join(r['elements'])}"
                if r["op"] == "reorder"
                else f"refreshed {len(r['cells'])} cached value(s) on {r.get('sheet', '?')}"
            )
            proof_lines.append(f"- surgery: {what} in {r['part']}")
        if p.get("output"):
            proof_lines.append(f"- proven workbook: `{p['output']}`")
            proof_lines.append(
                f"- test manifest: `{(p.get('tests') or {}).get('path', '')}` "
                f"({(p.get('tests') or {}).get('cases', 0)} cases) — rerun with "
                f"`python -m runner verify --target <file> --tests <manifest>`"
            )
        else:
            proof_lines.append(
                "- **no workbook returned** — a failed or unavailable proof ships nothing"
            )
        if p.get("engine_error"):
            proof_lines.append(f"- engine: {p['engine_error']}")
    else:
        proof_lines.append("Prove unit never ran (budget or failure) — the workbook is UNVERIFIED.")

    sections = [
        ("Workbook teardown", "\n".join(rows)),
        ("Findings verdict", "\n".join(f"- {v}" for v in verdict)),
    ]
    if graph_lines:
        sections.append(("Dependency graph", "\n".join(graph_lines)))
    sections += [
        ("Proof", "\n".join(proof_lines)),
        (
            "What this could not check",
            "\n".join(
                [
                    "- **Intent.** The engine proves the file computes what its formulas "
                    "say; it cannot prove the formulas say what you meant. Short ranges "
                    "and lone edited cells stay proposals with cell addresses.",
                    "- **Surgery only.** The input file is never modified; the proven "
                    "copy differs only in the parts the report lists, byte-for-byte.",
                ]
            ),
        ),
    ]
    out = {
        "sections": sections,
        "sheets": len(sheets),
        "error_cells": tot("error_cells"),
        "sum_mismatches": tot("sum_mismatches"),
    }
    if proofs:
        out["verdict"] = proofs[0].get("verdict")
        out["output"] = proofs[0].get("output")
    return out


def plan_items(run_id):
    f = store.findings(run_id)
    codes = {x["code"] for x in f}
    ev = lambda c: [x["ref"] for x in f if x["code"] == c][:6]
    items = []
    if "XL-SUMMISMATCH" in codes:
        items.append(
            dict(
                topic="Totals that lie",
                why="A SUM's cached value disagrees with the range it sums.",
                evidence=ev("XL-SUMMISMATCH"),
                read="Cached values vs formulas: the two-load rule",
                check="Open one flagged cell and re-enter the formula; watch the number move.",
                question="How would you detect this class of error across the whole workbook?",
            )
        )
    if "XL-ERRCELL" in codes:
        items.append(
            dict(
                topic="Error values shipped in the file",
                why="#REF!/#VALUE! are saved, not transient.",
                evidence=ev("XL-ERRCELL"),
                read="Excel error types and what each one means",
                check="The finding names the error chain and its root — open "
                "the root cell and confirm the diagnosis.",
                question="Why did the break propagate exactly as far as it did?",
            )
        )
    if "XL-CIRCULAR" in codes:
        items.append(
            dict(
                topic="Circular references",
                why="A cell depends, directly or through a chain, on itself.",
                evidence=ev("XL-CIRCULAR"),
                read="Iterative calculation and why 'it shows a number' is not 'it works'",
                check="Follow the named cycle in the finding; pick the one cell "
                "that should take an outside input.",
                question="With iteration on, what number does Excel converge to — and why is it wrong?",
            )
        )
    if "XL-SHORTRANGE" in codes:
        items.append(
            dict(
                topic="Ranges that stop short",
                why="A total excludes a row that holds data.",
                evidence=ev("XL-SHORTRANGE"),
                read="Why value checks cannot catch a wrong range",
                check="Add a row at the bottom of one table and see which totals move.",
                question="How would you make the total structurally unable to miss a row?",
            )
        )
    if "XL-INCONSISTENT" in codes:
        items.append(
            dict(
                topic="The lone edited cell",
                why="One formula differs inside a row of identical ones.",
                evidence=ev("XL-INCONSISTENT"),
                read="Spreadsheet error rates and consistency checking",
                check="Compare the flagged cell with its neighbours.",
                question="Was it deliberate, and how would the sheet tell you next time?",
            )
        )
    if "XL-VOLATILE" in codes or "XL-IOS-DYNAMIC" in codes:
        items.append(
            dict(
                topic="Your own iOS-safe engine rules",
                why="Volatile or dynamic-array functions are present.",
                evidence=ev("XL-VOLATILE") + ev("XL-IOS-DYNAMIC"),
                read="Your CHOOSE-based defined-name pattern that replaced INDIRECT",
                check="Open the workbook on the phone and watch what fails to render.",
                question="Why does INDIRECT cost you more than a CHOOSE lookup?",
            )
        )
    if "XL-OOXML-ORDER" in codes:
        items.append(
            dict(
                topic="OOXML element ordering",
                why="Child order is violated in at least one sheet part.",
                evidence=ev("XL-OOXML-ORDER"),
                read="The ordering rule: sheetData → mergeCells → conditionalFormatting → dataValidations → pageMargins",
                check="Unzip the file and read the sheet XML directly.",
                question="Which readers forgive bad ordering, and which one does not?",
            )
        )
    if "XL-MAGIC" in codes:
        items.append(
            dict(
                topic="Assumptions buried in formulas",
                why="Numbers are typed inside formulas instead of referenced.",
                evidence=ev("XL-MAGIC"),
                read="Every assumption in its own labelled cell",
                check="Pick one and find every other formula that should share it.",
                question="If that rate changes, how many cells must you edit?",
            )
        )
    if "XL-PROOF-FAILED" in codes or "XL-UNVERIFIED" in codes:
        items.append(
            dict(
                topic="What a proof actually is",
                why="The run could not certify the workbook — the report says which gate failed.",
                evidence=ev("XL-PROOF-FAILED") + ev("XL-UNVERIFIED"),
                read="The six proof gates and why absence of evidence is never cleanliness",
                check="Fix one failed gate and rerun; watch the verdict change.",
                question="Which failures can surgery repair, and which need a human?",
            )
        )
    return items
