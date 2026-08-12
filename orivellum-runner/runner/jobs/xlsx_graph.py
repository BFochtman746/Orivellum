"""The workbook dependency graph — what turns a linter into an auditor.

Built once per workbook from tokenized formulas (xlsx_formula). Nodes are
cells ("Sheet!A1"); an edge runs from a formula to each cell it reads. The
graph answers the questions the flat checks never could:

  · circular references — the one defect that yields a plausible wrong
    number forever, detected as strongly-connected components
  · precedent tracing — an error cell's chain is NAMED in the finding,
    not asked about in the training plan
  · orphans — formulas nothing depends on, numeric inputs no formula reads

HONESTY RULES
  · INDIRECT/OFFSET compute their references at run time. Cells behind them
    are marked `computed_ref`; the graph reports them as partially mapped
    rather than pretending the edges it can see are all the edges there are.
  · Whole-column/row references are clipped to the sheet's used range. A
    range is expanded to individual precedent cells only up to a cap; past
    it, membership is still tested against known cells (containment), so
    cycle detection never silently loses an edge.
"""

from collections import defaultdict

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from . import xlsx_formula as fx

EXPAND_CAP = 4096  # cells per referenced area expanded eagerly


def _key(sheet, col, row):
    return f"{sheet}!{get_column_letter(col)}{row}"


class WorkbookGraph:
    def __init__(self):
        self.formulas = {}  # key -> formula text
        self.facts = {}  # key -> analyze() dict
        self.precedents = defaultdict(set)  # key -> set of precedent keys
        self.dependents = defaultdict(set)  # inverse
        self.referenced = set()  # cells read via expanded (small) areas
        self.areas = defaultdict(list)  # key -> [(sheet,c1,r1,c2,r2)]
        # rectangle index for areas ABOVE the expansion cap:
        # {sheet: [(c1, r1, c2, r2), ...] deduped} — membership queries stay
        # exact for whole-column AND whole-row refs without materializing keys
        self.large_areas = {}
        self.computed_ref_cells = []  # cells whose refs are partly dynamic
        self.names = {}
        self.name_inventory = []
        self.tables = {}
        self.used_names = set()  # upper-case defined names actually referenced
        self.unresolved_names = set()  # names/refs the graph could not resolve to cells
        self.sheet_index = {}  # lower→canonical sheet names

    # ---------------------------------------------------------------- build
    @classmethod
    def build(cls, wbf, wbv=None):
        g = cls()
        g.names, g.name_inventory = fx.extract_names(wbf)
        g.tables = fx.extract_tables(wbf)
        g.sheet_index = {s.upper(): s for s in wbf.sheetnames}
        used = {}  # sheet -> (max_col, max_row)

        cells = []
        for ws in wbf.worksheets:
            used[ws.title] = (ws.max_column or 1, ws.max_row or 1)
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        cells.append((ws.title, c.column, c.row, c.value))

        for sheet, col, row, formula in cells:
            key = _key(sheet, col, row)
            facts = fx.analyze(formula, sheet, g.names, g.tables, cell=(col, row))
            g.formulas[key] = formula
            g.facts[key] = facts
            if facts["computed_ref"]:
                g.computed_ref_cells.append(key)
            g.used_names.update(n.upper() for n in facts["name_refs"])
            g.unresolved_names.update(facts["unresolved_names"])
            for area in facts["areas"]:
                asheet, c1, r1, c2, r2 = area
                asheet = g.sheet_index.get(str(asheet).upper(), asheet)
                mc, mr = used.get(asheet, (fx.MAX_COL, fx.MAX_ROW))
                c2, r2 = min(c2, mc), min(r2, mr)
                if c2 < c1 or r2 < r1:
                    continue
                g.areas[key].append((asheet, c1, r1, c2, r2))
                if (c2 - c1 + 1) * (r2 - r1 + 1) <= EXPAND_CAP:
                    for rr in range(r1, r2 + 1):
                        for cc in range(c1, c2 + 1):
                            # pk == key is a direct self-reference — keep the
                            # edge, it IS the circular reference
                            pk = _key(asheet, cc, rr)
                            g.precedents[key].add(pk)
                            g.referenced.add(pk)

        # Containment pass: every formula cell inside ANY referenced area is a
        # precedent even when the area was too big to expand — cycle edges are
        # never lost to the cap. Formula cells are bucketed per (sheet, col)
        # AND per (sheet, row); each capped rectangle walks its NARROW
        # dimension, so whole-column refs cost O(cells in those columns) and
        # whole-row refs cost O(cells in those rows) — never 16,384 iterations.
        by_col = defaultdict(list)  # (sheet, col) -> [(key, row)]
        by_row = defaultdict(list)  # (sheet, row) -> [(key, col)]
        for k in g.formulas:
            sheet, _, coord = k.partition("!")
            cl, rw = coordinate_from_string(coord)
            ci = column_index_from_string(cl)
            by_col[(sheet, ci)].append((k, rw))
            by_row[(sheet, rw)].append((k, ci))
        large_rects = defaultdict(set)
        for key, areas in g.areas.items():
            for asheet, c1, r1, c2, r2 in areas:
                if (c2 - c1 + 1) * (r2 - r1 + 1) <= EXPAND_CAP:
                    continue
                large_rects[asheet].add((c1, r1, c2, r2))
                if (c2 - c1) <= (r2 - r1):  # tall rectangle: walk its columns
                    for cc in range(c1, c2 + 1):
                        for fk, fr in by_col.get((asheet, cc), ()):
                            if r1 <= fr <= r2:
                                g.precedents[key].add(fk)
                                g.referenced.add(fk)
                else:  # wide rectangle: walk its rows
                    for rr in range(r1, r2 + 1):
                        for fk, fc in by_row.get((asheet, rr), ()):
                            if c1 <= fc <= c2:
                                g.precedents[key].add(fk)
                                g.referenced.add(fk)
        g.large_areas = {s: sorted(v) for s, v in large_rects.items()}

        for key, precs in g.precedents.items():
            for p in precs:
                g.dependents[p].add(key)
        return g

    # --------------------------------------------------------------- cycles
    def cycles(self, limit=20):
        """Strongly-connected components of size > 1, plus direct self-loops.
        Iterative Tarjan — recursion depth is workbook-controlled otherwise."""
        index, low, onstack = {}, {}, set()
        stack, out, counter = [], [], [0]
        nodes = [k for k in self.formulas if self.precedents.get(k)]
        prec = {k: [p for p in self.precedents.get(k, ()) if p in self.formulas] for k in nodes}

        for root in nodes:
            if root in index:
                continue
            work = [(root, iter(prec.get(root, ())))]
            while work:
                node, it = work[-1]
                if node not in index:
                    index[node] = low[node] = counter[0]
                    counter[0] += 1
                    stack.append(node)
                    onstack.add(node)
                advanced = False
                for nxt in it:
                    if nxt not in index:
                        work.append((nxt, iter(prec.get(nxt, ()))))
                        advanced = True
                        break
                    if nxt in onstack:
                        low[node] = min(low[node], index[nxt])
                if advanced:
                    continue
                work.pop()
                if work:
                    parent = work[-1][0]
                    low[parent] = min(low[parent], low[node])
                if low[node] == index[node]:
                    comp = []
                    while True:
                        w = stack.pop()
                        onstack.discard(w)
                        comp.append(w)
                        if w == node:
                            break
                    if len(comp) > 1:
                        out.append(sorted(comp))
                    elif node in prec.get(node, ()):
                        out.append([node])
                if len(out) >= limit:
                    return out
        return out

    # --------------------------------------------------------------- traces
    def trace_precedents(self, key, depth=4, width=6):
        """A named precedent chain: ["Model!B6", "← Model!B4, Model!B5", …].
        Formula precedents first — they are where a defect propagates from."""
        lines, frontier, seen = [], [key], {key}
        for _ in range(depth):
            nxt = []
            for k in frontier:
                for p in sorted(
                    self.precedents.get(k, ()),
                    key=lambda x: (x not in self.formulas, x),
                ):
                    if p not in seen:
                        seen.add(p)
                        nxt.append(p)
            if not nxt:
                break
            shown = nxt[:width]
            more = f" (+{len(nxt) - width} more)" if len(nxt) > width else ""
            lines.append("← " + ", ".join(shown) + more)
            frontier = [n for n in shown if n in self.formulas]
            if not frontier:
                break
        return lines

    def trace_error_root(self, key, error_values, depth=8):
        """Follow error-valued precedents down to the deepest erroring cell —
        the root that actually broke."""
        cur, chain, seen = key, [key], {key}
        for _ in range(depth):
            nxt = [
                p
                for p in self.precedents.get(cur, ())
                if p in error_values and p not in seen
            ]
            if not nxt:
                break
            cur = sorted(nxt)[0]
            seen.add(cur)
            chain.append(cur)
        return chain

    # -------------------------------------------------------------- orphans
    def orphan_formulas(self):
        """Formula cells no other formula reads. Terminal outputs land here
        legitimately — reported as an aggregate observation, never a defect."""
        return sorted(k for k in self.formulas if not self.dependents.get(k))

    def is_referenced(self, sheet, col, row):
        """Exact membership: expanded edges OR the rectangle index of areas
        too large to expand — a cell inside =SUM(A:A) is READ, always."""
        if _key(sheet, col, row) in self.referenced:
            return True
        return any(
            c1 <= col <= c2 and r1 <= row <= r2
            for c1, r1, c2, r2 in self.large_areas.get(sheet, ())
        )

    def unread_inputs(self, wbv):
        """Numeric input cells (per sheet) that no formula anywhere reads.
        Text cells are labels; only unread NUMBERS suggest dead data.

        Capped-rectangle membership is batched PER ROW: the active column
        intervals are computed once when the row changes, so the cost is
        O(rows × rects + cells × active intervals) — never a full rectangle
        scan per numeric cell."""
        out = defaultdict(list)
        for ws in wbv.worksheets:
            rects = self.large_areas.get(ws.title, ())
            cur_row, intervals = None, ()
            for row in ws.iter_rows():
                for c in row:
                    if not isinstance(c.value, (int, float)) or isinstance(c.value, bool):
                        continue
                    k = _key(ws.title, c.column, c.row)
                    if k in self.formulas or k in self.referenced:
                        continue
                    if c.row != cur_row:
                        cur_row = c.row
                        spans = sorted(
                            (rc1, rc2) for rc1, rr1, rc2, rr2 in rects if rr1 <= cur_row <= rr2
                        )
                        merged = []
                        for a, b in spans:
                            if merged and a <= merged[-1][1] + 1:
                                merged[-1][1] = max(merged[-1][1], b)
                            else:
                                merged.append([a, b])
                        intervals = merged
                    if any(a <= c.column <= b for a, b in intervals):
                        continue
                    out[ws.title].append(k)
        return dict(out)

    # ---------------------------------------------------------------- names
    def name_audit(self):
        """(orphaned names, shadowing names). Orphan: defined but never
        referenced by any formula. Shadow: sheet-scoped name spelled the same
        as a workbook-scoped one — formulas resolve differently per sheet."""
        workbook_scope = {n.upper() for n, scope in self.name_inventory if scope is None}
        orphans, shadows = [], []
        for name, scope in self.name_inventory:
            if name.upper() not in self.used_names:
                orphans.append((name, scope))
            if scope is not None and name.upper() in workbook_scope:
                shadows.append((name, scope))
        return orphans, shadows

    def stats(self):
        return {
            "formula_cells": len(self.formulas),
            "edges": sum(len(v) for v in self.precedents.values()),
            "referenced_cells": len(self.referenced),
            "computed_ref_cells": len(self.computed_ref_cells),
            "defined_names": len(self.name_inventory),
            "tables": len(self.tables),
        }
