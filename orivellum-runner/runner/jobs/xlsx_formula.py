"""Formula analysis on a real tokenizer — the regexes are retired.

`openpyxl.formula.tokenizer` is a proper Excel formula tokenizer and openpyxl
is already a dependency. Tokenizing instead of regex-matching removes an
entire class of silent false findings:

  · string literals   =IF(A1="See B2:C9",…)   no phantom B2:C9 reference
  · table references  =SUM(Table1[@Amount])    resolved via the table's range
  · defined names     =SUM(Q1_Revenue)         resolved via the name registry
  · array constants   {1,2;3,4}                data, never "magic numbers"

Every function here is pure: formula text in, structured facts out. Resolution
of names and tables happens against maps the caller extracts from the
workbook, so this module never opens a file.
"""

import re
from functools import lru_cache

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter

# A1-style piece inside an operand (anchors optional)
_CELL = re.compile(r"^(\$?)([A-Z]{1,3})(\$?)(\d{1,7})$")
_COLSPAN = re.compile(r"^(\$?)([A-Z]{1,3}):(\$?)([A-Z]{1,3})$")
_ROWSPAN = re.compile(r"^(\$?)(\d{1,7}):(\$?)(\d{1,7})$")

MAX_ROW = 1_048_576
MAX_COL = 16_384

# Functions whose reference argument is COMPUTED at run time — the static
# graph cannot see through them and says so instead of guessing.
COMPUTED_REF_FUNCS = {"INDIRECT", "OFFSET"}


@lru_cache(maxsize=4096)
def _tokens(formula):
    try:
        return tuple(Tokenizer(formula).items)
    except Exception:  # noqa: BLE001 — an untokenizable formula yields no facts
        return ()


def _split_sheet(operand):
    """'My Sheet'!A1:B2 → ("My Sheet", "A1:B2"); A1 → (None, "A1")."""
    if "!" in operand:
        prefix, _, rest = operand.rpartition("!")
        prefix = prefix.strip()
        if prefix.startswith("'") and prefix.endswith("'"):
            prefix = prefix[1:-1].replace("''", "'")
        return prefix, rest
    return None, operand


def _is_external(operand):
    """External-workbook refs carry a [bookindex] or [Book.xlsx] prefix on the
    sheet part — distinct from Table1[Col], where the bracket FOLLOWS a name."""
    prefix, _rest = _split_sheet(operand)
    return (prefix or operand).startswith("[")


def _parse_area(ref):
    """One un-sheeted A1-style ref → (c1, r1, c2, r2, anchors) or None.
    anchors is the tuple of '$' flags, preserved for the anchoring checks."""
    ref = ref.strip()
    if ":" in ref:
        a, _, b = ref.partition(":")
        ma, mb = _CELL.match(a), _CELL.match(b)
        if ma and mb:
            c1, r1 = column_index_from_string(ma.group(2)), int(ma.group(4))
            c2, r2 = column_index_from_string(mb.group(2)), int(mb.group(4))
            anchors = (bool(ma.group(1)), bool(ma.group(3)), bool(mb.group(1)), bool(mb.group(3)))
            return (min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2), anchors)
        mc = _COLSPAN.match(ref)
        if mc:
            c1 = column_index_from_string(mc.group(2))
            c2 = column_index_from_string(mc.group(4))
            anchors = (bool(mc.group(1)), True, bool(mc.group(3)), True)
            return (min(c1, c2), 1, max(c1, c2), MAX_ROW, anchors)
        mr = _ROWSPAN.match(ref)
        if mr:
            r1, r2 = int(mr.group(2)), int(mr.group(4))
            anchors = (True, bool(mr.group(1)), True, bool(mr.group(3)))
            return (1, min(r1, r2), MAX_COL, max(r1, r2), anchors)
        return None
    m = _CELL.match(ref)
    if not m:
        return None
    c, r = column_index_from_string(m.group(2)), int(m.group(4))
    return (c, r, c, r, (bool(m.group(1)), bool(m.group(3)), bool(m.group(1)), bool(m.group(3))))


def _split_top(s):
    """Split on commas at bracket depth zero: '[#Headers],[Amt]' → 2 parts."""
    parts, depth, cur = [], 0, ""
    for ch in s:
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append(cur)
            cur = ""
        else:
            cur += ch
    if cur.strip():
        parts.append(cur)
    return [p.strip() for p in parts]


def _table_area(table, spec, cell=None):
    """Resolve a structured reference Table1[...] EXACTLY, or refuse.

    Returns a LIST of (sheet, c1, r1, c2, r2) areas — disjoint selections
    like [[#Headers],[#Totals]] yield one rectangle per region, never a
    bounding box that would invent data-body dependencies. Item specifiers
    map to their true rows: #Headers → header row(s), #Totals → totals
    row(s), #Data → data body, #All → everything, and @ → the CALLING
    formula's row (needs `cell`; only valid inside the data body). Anything
    the parser cannot pin down returns None — the caller discloses it as
    graph-partial instead of silently substituting the data body.
    """
    sheet, c1, r1, c2, r2, header = table[:6]
    col_names = table[6] if len(table) > 6 else None
    totals = table[7] if len(table) > 7 else 0
    data_r1, data_r2 = r1 + header, r2 - totals
    inner = spec.strip()
    if inner.startswith("[") and inner.endswith("]"):
        inner = inner[1:-1].strip()

    # col_groups holds (lo_name, hi_name) pairs: a range keeps its span, a
    # single column is (name, name) — NON-ADJACENT selections stay separate
    # rectangles so unselected columns never enter the graph
    specs, col_groups, at = set(), [], False
    for part in _split_top(inner):
        p = part.strip()
        if "]:[" in p:  # column range [Col1]:[Col2]
            sides = [s.strip().strip("[]").strip() for s in p.split(":")]
            if len(sides) != 2 or not all(sides):
                return None
            col_groups.append((sides[0], sides[1]))
            continue
        if p.startswith("[") and p.endswith("]"):
            p = p[1:-1].strip()
        if not p:
            continue
        if p.startswith("#"):
            specs.add(p.upper())
        elif p.startswith("@"):
            at = True
            rest = p[1:].strip().strip("[]").strip()
            if rest:
                col_groups.append((rest, rest))
        else:
            col_groups.append((p, p))

    # rows — a LIST of spans; adjacent spans merge, disjoint ones stay apart
    if at:
        if specs or cell is None:
            return None  # @ mixed with specifiers, or no calling cell known
        crow = cell[1]
        if not (data_r1 <= crow <= data_r2):
            return None  # @ outside the data body is a #VALUE! in Excel
        row_spans = [(crow, crow)]
    elif specs:
        row_spans = []
        for s in specs:
            if s == "#ALL":
                row_spans.append((r1, r2))
            elif s == "#DATA":
                row_spans.append((data_r1, data_r2))
            elif s == "#HEADERS":
                if header <= 0:
                    return None  # no header row → #REF! in Excel
                row_spans.append((r1, r1 + header - 1))
            elif s == "#TOTALS":
                if totals <= 0:
                    return None  # no totals row → #REF! in Excel
                row_spans.append((data_r2 + 1, r2))
            else:
                return None  # #This Row spelled out, or unknown specifier
        row_spans.sort()
        merged = [list(row_spans[0])]
        for a, b in row_spans[1:]:
            if a <= merged[-1][1] + 1:
                merged[-1][1] = max(merged[-1][1], b)
            else:
                merged.append([a, b])
        row_spans = [tuple(m) for m in merged]
    else:
        row_spans = [(data_r1, data_r2)]
    if any(lo > hi for lo, hi in row_spans):
        return None

    # columns — like rows, a LIST of spans; adjacent merge, disjoint stay
    # apart: T[[A],[C]] must never invent a dependency on column B
    if col_groups:
        if not col_names:
            return None
        spans = []
        try:
            for lo_name, hi_name in col_groups:
                i, j = col_names.index(lo_name), col_names.index(hi_name)
                spans.append((c1 + min(i, j), c1 + max(i, j)))
        except ValueError:
            return None  # unknown column name — refuse, don't guess
        spans.sort()
        cmerged = [list(spans[0])]
        for a, b in spans[1:]:
            if a <= cmerged[-1][1] + 1:
                cmerged[-1][1] = max(cmerged[-1][1], b)
            else:
                cmerged.append([a, b])
        col_spans = [tuple(m) for m in cmerged]
    else:
        col_spans = [(c1, c2)]
    return [
        (sheet, cs, lo, ce, hi) for lo, hi in row_spans for cs, ce in col_spans
    ]


def analyze(formula, sheet, names=None, tables=None, cell=None):
    """Structured facts about one formula.

    names:  {NAME_UPPER: [(sheet, "A1:B2"), ...]}       (workbook scope)
            {(SHEET_UPPER, NAME_UPPER): [...]}          (sheet scope)
    tables: {NAME_UPPER: (sheet, c1, r1, c2, r2, header_rows, [cols], totals)}
    cell:   (col, row) of the formula itself — required to resolve @ (this
            row) structured references exactly

    Returns dict:
      functions     set of upper-case function names actually CALLED
      areas         [(sheet, c1, r1, c2, r2)] every statically resolvable ref
      anchors       [((sheet, c1, r1, c2, r2), anchor_flags)] for direct refs
      literals      numeric literal strings used as operands (never from
                    string text, never from array constants)
      external      operand strings that reference another workbook
      table_refs    table names referenced
      name_refs     defined names referenced (resolved ones)
      unresolved_names  names referenced but absent from the registry
      computed_ref  True when INDIRECT/OFFSET make some refs uncomputable
    """
    names = names or {}
    tables = tables or {}
    out = {
        "functions": set(),
        "areas": [],
        "anchors": [],
        "literals": [],
        "external": [],
        "table_refs": [],
        "name_refs": [],
        "unresolved_names": [],
        "computed_ref": False,
    }
    array_depth = 0
    for tok in _tokens(formula):
        if tok.type == "ARRAY":
            array_depth += 1 if tok.subtype == "OPEN" else -1
            continue
        if tok.type == "FUNC" and tok.subtype == "OPEN":
            fn = tok.value[:-1].rpartition("!")[-1].upper()  # strip '(' and any sheet prefix
            out["functions"].add(fn)
            if fn in COMPUTED_REF_FUNCS:
                out["computed_ref"] = True
            continue
        if tok.type != "OPERAND":
            continue
        if tok.subtype == "NUMBER":
            if array_depth == 0:
                out["literals"].append(tok.value)
            continue
        if tok.subtype != "RANGE":
            continue  # TEXT, LOGICAL, ERROR: never references
        operand = tok.value
        if _is_external(operand):
            out["external"].append(operand)
            continue
        if "[" in operand:  # structured table reference
            tname, _, spec = operand.partition("[")
            key = tname.strip().upper()
            if key in tables:
                out["table_refs"].append(tname.strip())
                resolved = _table_area(tables[key], "[" + spec, cell)
                if resolved is not None:
                    out["areas"].extend(resolved)
                else:
                    # exact resolution refused (unknown column, @ without a
                    # calling cell, missing totals/header row) — disclose,
                    # never substitute the data body
                    out["unresolved_names"].append(operand)
            else:
                out["unresolved_names"].append(operand)
            continue
        ref_sheet, rest = _split_sheet(operand)
        area = _parse_area(rest)
        if area:
            c1, r1, c2, r2, anch = area
            resolved = ((ref_sheet or sheet), c1, r1, c2, r2)
            out["areas"].append(resolved)
            out["anchors"].append((resolved, anch))
            continue
        # a bare identifier: defined name. Recognition is separate from
        # resolution — constant and formula-defined names ARE used names even
        # though they resolve to no cell area (the graph is a lower bound
        # there, and says so via unresolved_names).
        key = rest.upper()
        skey = (str(ref_sheet or sheet).upper(), key)
        if skey in names:
            scoped = names[skey]
        elif key in names:
            scoped = names[key]
        else:
            scoped = None
        if scoped is None:
            out["unresolved_names"].append(rest)
            continue
        out["name_refs"].append(rest)
        if not scoped:  # known name, but no cell destination (constant/formula)
            out["unresolved_names"].append(rest)
        for nsheet, nref in scoped:
            narea = _parse_area(nref.split("!")[-1])
            if narea:
                out["areas"].append((nsheet, *narea[:4]))
    return out


def normalize(formula, row, col):
    """Offset-normalised shape so sibling formulas compare equal:
    =B2+C2 in row 2 and =B3+C3 in row 3 → one shape. Tokenized, so cell-like
    text inside string literals is never rewritten (the regex version's bug)."""
    return _shape(formula, row, col, keep_anchors=True)


def skeleton(formula):
    """The formula with every reference collapsed to ⟨R⟩ — structural
    identity regardless of which cells it points at."""
    return _shape(formula, 0, 0, placeholder=True)


def anchor_signature(formula):
    """Tuple of the $-anchor flags of each direct reference, in order —
    ((col$, row$), …). Two structurally identical formulas whose signatures
    differ are anchored inconsistently."""
    sig = []
    for tok in _tokens(formula):
        if tok.type == "OPERAND" and tok.subtype == "RANGE" and "[" not in tok.value:
            _s, rest = _split_sheet(tok.value)
            area = _parse_area(rest)
            if area:
                sig.append(area[4])
    return tuple(sig)


def _shape(formula, row, col, keep_anchors=False, placeholder=False):
    parts = []
    for tok in _tokens(formula):
        if tok.type == "OPERAND" and tok.subtype == "RANGE" and "[" not in tok.value:
            sheet_prefix, rest = _split_sheet(tok.value)
            if placeholder:
                parts.append("⟨R⟩")
                continue
            pieces = []
            ok = True
            for piece in rest.split(":"):
                m = _CELL.match(piece.strip())
                if not m:
                    ok = False
                    break
                abs_c, c, abs_r, r = m.group(1), m.group(2), m.group(3), int(m.group(4))
                ci = column_index_from_string(c)
                cc = f"C{ci}" if abs_c else f"C[{ci - col}]"
                rr = f"R{r}" if abs_r else f"R[{r - row}]"
                pieces.append(rr + cc)
            if ok:
                prefix = f"{sheet_prefix}!" if sheet_prefix else ""
                parts.append(prefix + ":".join(pieces))
            else:
                parts.append(tok.value)
        else:
            parts.append(tok.value)
    return "".join(parts)


def extract_names(wb):
    """{NAME_UPPER: [(sheet, ref)]} plus {(SHEET_UPPER, NAME_UPPER): …} for
    sheet-scoped names, and the list of (name, scope_sheet) for audits."""
    names, inventory = {}, []

    def _add(key, dn):
        try:
            dests = [(s, r) for s, r in dn.destinations]
        except Exception:  # noqa: BLE001 — constants/formulas have no destinations
            dests = []
        # ALWAYS register the name — a constant or formula-defined name is
        # still a name formulas legitimately use; an empty destination list
        # just means the graph cannot resolve it to cells.
        names[key] = dests

    for name, dn in getattr(wb.defined_names, "items", lambda: [])():
        _add(name.upper(), dn)
        inventory.append((name, None))
    for ws in wb.worksheets:
        local = getattr(ws, "defined_names", None)
        if not local:
            continue
        for name, dn in local.items():
            _add((ws.title.upper(), name.upper()), dn)
            inventory.append((name, ws.title))
    return names, inventory


def extract_tables(wb):
    """{NAME_UPPER: (sheet, c1, r1, c2, r2, header_rows, [cols], totals_rows)}."""
    from openpyxl.utils import range_boundaries

    out = {}
    for ws in wb.worksheets:
        for tname, tbl in getattr(ws, "tables", {}).items():
            try:
                ref = tbl.ref if hasattr(tbl, "ref") else str(tbl)
                c1, r1, c2, r2 = range_boundaries(ref)
                header = getattr(tbl, "headerRowCount", 1) or 0
                totals = getattr(tbl, "totalsRowCount", 0) or 0
                cols = [c.name for c in getattr(tbl, "tableColumns", [])] or [
                    get_column_letter(i) for i in range(c1, c2 + 1)
                ]
                # Excel resolves structured refs against the HEADER TEXT, and
                # keeps tableColumns in sync with it — files written by other
                # tools may not, so the header cells are authoritative
                if header > 0:
                    for i, cc in enumerate(range(c1, c2 + 1)):
                        v = ws.cell(row=r1, column=cc).value
                        if isinstance(v, str) and v.strip() and i < len(cols):
                            cols[i] = v.strip()
                out[tname.upper()] = (ws.title, c1, r1, c2, r2, header, cols, totals)
            except Exception:  # noqa: BLE001
                continue
    return out
