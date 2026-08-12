"""The auditor layer: tokenized formula analysis + the dependency graph.

Two promises are pinned here:

  1. NO PHANTOM FINDINGS — string literals, table references, defined names
     and array constants (the regex era's false-positive classes) produce
     zero bogus references, zero bogus magic numbers, zero bogus ext-links.
  2. THE GRAPH ANSWERS — circular references are named, error findings carry
     their precedent chain, orphans and name hygiene are reported.
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from runner import store  # noqa: E402
from runner.jobs import xlsx  # noqa: E402
from runner.jobs import xlsx_formula as fx  # noqa: E402
from runner.jobs import xlsx_graph as depgraph  # noqa: E402


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    from runner.config import CFG

    monkeypatch.setattr(CFG, "runs_dir", str(tmp_path / "runs"))
    monkeypatch.setattr(CFG, "db", str(tmp_path / "runs" / "runner.db"))
    monkeypatch.setattr(CFG, "mock", True)
    xlsx._drop_cache()
    yield
    xlsx._drop_cache()


def _make_run(target="x"):
    store.init()
    return store.start_run("xlsx", str(target), "test", {})


def _codes(run_id):
    return [(f["code"], f["ref"]) for f in store.findings(run_id)]


# ------------------------------------------------------ tokenizer: analyze()


def test_string_literals_produce_no_phantom_references():
    facts = fx.analyze('=IF(A1="See B2:C9","ok","no")', "S")
    assert facts["areas"] == [("S", 1, 1, 1, 1)]  # A1 only — never B2:C9
    assert facts["functions"] == {"IF"}
    assert not facts["literals"] and not facts["external"]


def test_array_constants_are_data_not_magic_numbers():
    facts = fx.analyze("=SUM({5,7;9,11})+B2*365", "S")
    assert facts["literals"] == ["365"]  # array members never counted
    assert facts["areas"] == [("S", 2, 2, 2, 2)]


def test_table_reference_is_not_an_external_link():
    tables = {"TABLE1": ("Data", 1, 1, 2, 5, 1, ["Amount", "Qty"], 0)}
    facts = fx.analyze("=SUM(Table1[Amount])", "S", tables=tables)
    assert not facts["external"]
    assert facts["table_refs"] == ["Table1"]
    # resolved to the Amount column's data body: col 1, rows 2-5
    assert facts["areas"] == [("Data", 1, 2, 1, 5)]


def test_structured_ref_item_specifiers_resolve_exactly():
    # Data sheet: table A1:B5, 1 header row, 1 totals row → data body rows 2-4
    t = {"T": ("Data", 1, 1, 2, 5, 1, ["Amount", "Qty"], 1)}
    assert fx.analyze("=COUNTA(T[#Headers])", "S", tables=t)["areas"] == [("Data", 1, 1, 2, 1)]
    assert fx.analyze("=SUM(T[#Totals])", "S", tables=t)["areas"] == [("Data", 1, 5, 2, 5)]
    assert fx.analyze("=SUM(T[#Data])", "S", tables=t)["areas"] == [("Data", 1, 2, 2, 4)]
    assert fx.analyze("=SUM(T[#All])", "S", tables=t)["areas"] == [("Data", 1, 1, 2, 5)]
    # [[#Headers],[Qty]] → the Qty header cell only
    facts = fx.analyze("=T[[#Headers],[Qty]]", "S", tables=t)
    assert facts["areas"] == [("Data", 2, 1, 2, 1)]
    # disjoint union: headers + totals are TWO areas — the data body (rows
    # 2-4) must NOT be swallowed by a bounding box
    disjoint = fx.analyze("=COUNTA(T[[#Headers],[#Totals]])", "S", tables=t)
    assert disjoint["areas"] == [("Data", 1, 1, 2, 1), ("Data", 1, 5, 2, 5)]
    assert not any(a[2] <= 3 <= a[4] for a in disjoint["areas"])  # no data rows
    # adjacent union: headers + data merge into one exact rectangle
    adjacent = fx.analyze("=COUNTA(T[[#Headers],[#Data]])", "S", tables=t)
    assert adjacent["areas"] == [("Data", 1, 1, 2, 4)]


def test_structured_ref_non_contiguous_columns_stay_disjoint():
    """T[[Alpha],[Gamma]] must never invent a dependency on column Beta."""
    t = {"T": ("Data", 1, 1, 3, 5, 1, ["Alpha", "Beta", "Gamma"], 0)}
    facts = fx.analyze("=SUM(T[[Alpha],[Gamma]])", "S", tables=t)
    # two areas: col 1 and col 3 data bodies — col 2 (Beta) excluded
    assert facts["areas"] == [("Data", 1, 2, 1, 5), ("Data", 3, 2, 3, 5)]
    assert not any(a[1] <= 2 <= a[3] for a in facts["areas"])
    # combined with a specifier: two header cells, still no Beta
    hdr = fx.analyze("=T[[#Headers],[Alpha],[Gamma]]", "S", tables=t)
    assert hdr["areas"] == [("Data", 1, 1, 1, 1), ("Data", 3, 1, 3, 1)]
    # adjacent columns still merge into one rectangle
    adj = fx.analyze("=SUM(T[[Alpha],[Beta]])", "S", tables=t)
    assert adj["areas"] == [("Data", 1, 2, 2, 5)]
    # explicit range spans contiguously by definition
    rng = fx.analyze("=SUM(T[[Alpha]:[Gamma]])", "S", tables=t)
    assert rng["areas"] == [("Data", 1, 2, 3, 5)]


def test_structured_ref_this_row_uses_calling_cell():
    t = {"T": ("Data", 1, 1, 2, 5, 1, ["Amount", "Qty"], 1)}
    # formula sits in data-body row 3 → @Amount is exactly Data!A3
    facts = fx.analyze("=T[@Amount]*2", "Data", tables=t, cell=(4, 3))
    assert facts["areas"] == [("Data", 1, 3, 1, 3)]
    # without a calling cell, @ is unresolvable — refused, disclosed
    blind = fx.analyze("=T[@Amount]*2", "Data", tables=t)
    assert not blind["areas"] and blind["unresolved_names"] == ["T[@Amount]"]
    # @ outside the data body is a #VALUE! in Excel — refused, not guessed
    outside = fx.analyze("=T[@Amount]*2", "Data", tables=t, cell=(4, 9))
    assert not outside["areas"] and outside["unresolved_names"]


def test_structured_ref_refuses_missing_targets():
    # no totals row, unknown column: refuse and disclose, never substitute
    t = {"T": ("Data", 1, 1, 2, 5, 1, ["Amount", "Qty"], 0)}
    no_tot = fx.analyze("=SUM(T[#Totals])", "S", tables=t)
    assert not no_tot["areas"] and no_tot["unresolved_names"] == ["T[#Totals]"]
    bad_col = fx.analyze("=SUM(T[Nope])", "S", tables=t)
    assert not bad_col["areas"] and bad_col["unresolved_names"] == ["T[Nope]"]


def test_defined_name_resolves_to_its_destination():
    names = {"Q1_REVENUE": [("Data", "$B$2:$B$5")]}
    facts = fx.analyze("=SUM(Q1_Revenue)", "S", names=names)
    assert facts["name_refs"] == ["Q1_Revenue"]
    assert facts["areas"] == [("Data", 2, 2, 2, 5)]
    assert not facts["unresolved_names"]


def test_external_workbook_reference_detected():
    facts = fx.analyze("='[Book2.xlsx]Sheet1'!A1+[1]Sheet1!C3", "S")
    assert len(facts["external"]) == 2
    assert not facts["areas"]  # external refs never enter the local graph


def test_computed_reference_flag():
    assert fx.analyze('=INDIRECT("A"&C1)', "S")["computed_ref"]
    assert fx.analyze("=OFFSET(A1,1,1)", "S")["computed_ref"]
    assert not fx.analyze("=SUM(A1:A9)", "S")["computed_ref"]


def test_normalize_ignores_cell_text_inside_strings():
    # regex-era bug: "B2" inside the string literal got rewritten
    a = fx.normalize('=IF(B2="B2",C2,0)', 2, 4)
    b = fx.normalize('=IF(B3="B2",C3,0)', 3, 4)
    assert a == b


def test_anchor_signature_distinguishes_fills():
    assert fx.anchor_signature("=A1*$B$1") != fx.anchor_signature("=A1*B1")
    assert fx.skeleton("=A1*$B$1") == fx.skeleton("=A4*B4")


# ------------------------------------------------- fixtures for whole files


def _wb_false_positive_classes(path):
    """Every regex-era false-positive class in one workbook."""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Amount", "Qty"])  # table header
    for amt, qty in ((10.5, 1), (20.25, 2), (30.75, 3)):
        ws.append([amt, qty])
    ws.add_table(Table(displayName="Table1", ref="A1:B4"))
    ws["D2"] = 100.0
    ws["D3"] = 200.0
    ws["D4"] = 300.0
    wb.defined_names["Q1_Revenue"] = DefinedName("Q1_Revenue", attr_text="Data!$D$2:$D$4")

    calc = wb.create_sheet("Calc")
    calc["A1"] = '=IF(Data!A2=1,"See B2:C9","no")'  # string literal trap
    calc["A2"] = "=SUM(Table1[Amount])"  # structured table ref
    calc["A3"] = "=SUM(Q1_Revenue)"  # defined name
    calc["A4"] = "=SUM({5,7;9,11})"  # array constant
    wb.save(path)
    return path


def test_no_phantom_findings_on_false_positive_classes(tmp_path):
    p = _wb_false_positive_classes(tmp_path / "fp.xlsx")
    run_id = _make_run(p)
    for sheet in ("Data", "Calc"):
        xlsx.sheet_unit(run_id, {"sheet": sheet, "target": str(p)})
    xlsx.workbook_unit(run_id, {"target": str(p)})
    codes = [c for c, _ in _codes(run_id)]
    assert "XL-EXTLINK" not in codes, "table ref misread as an external link"
    assert "XL-MAGIC" not in codes, "string/array content misread as magic numbers"
    assert "XL-ERRCELL" not in codes and "XL-CIRCULAR" not in codes
    # and the names/table actually resolved: nothing reported unresolved,
    # the defined name is USED (no orphan finding for it)
    orphan = [r for c, r in _codes(run_id) if c == "XL-NAME-ORPHAN"]
    assert not orphan


def test_graph_resolves_names_and_tables_into_edges(tmp_path):
    p = _wb_false_positive_classes(tmp_path / "fp2.xlsx")
    from openpyxl import load_workbook

    wbf = load_workbook(p, data_only=False)
    g = depgraph.WorkbookGraph.build(wbf)
    # =SUM(Table1[Amount]) reads Data!A2:A4; =SUM(Q1_Revenue) reads Data!D2:D4
    assert "Data!A2" in g.precedents["Calc!A2"]
    assert "Data!D3" in g.precedents["Calc!A3"]
    assert g.dependents["Data!D2"] == {"Calc!A3"}
    wbf.close()


# ------------------------------------------------------------ circular refs


def test_circular_reference_is_named(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = "=B1+1"
    ws["B1"] = "=C1*2"
    ws["C1"] = "=A1-3"  # A1 → B1 → C1 → A1
    ws["E5"] = "=E5+1"  # direct self-loop
    p = tmp_path / "circ.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    assert digest["circular"] == 2
    circ = [f for f in store.findings(run_id) if f["code"] == "XL-CIRCULAR"]
    assert len(circ) == 2
    details = " | ".join(f["detail"] for f in circ)
    for cell in ("M!A1", "M!B1", "M!C1", "M!E5"):
        assert cell in details, f"{cell} missing from named cycles"
    # the displayed chain must follow ACTUAL directed edges (a reads b),
    # not just list sorted SCC members
    g = xlsx._graph(p)
    for cyc in g.cycles():
        for a, b in zip(cyc, cyc[1:] + [cyc[0]], strict=True):
            assert b in g.precedents[a], f"{a} → {b} is not a real edge"


def test_clean_workbook_has_no_circular_findings(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1*2"
    ws["A3"] = "=A2+A1"
    p = tmp_path / "clean.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    assert digest["circular"] == 0
    assert not [f for f in store.findings(run_id) if f["code"] == "XL-CIRCULAR"]


# ------------------------------------------------------- precedent tracing


def test_error_finding_names_its_precedent_chain(tmp_path):
    """B1 errors; C1 consumes it. C1's finding must name B1 as the root."""
    from openpyxl import Workbook
    from runner.jobs import xlsx_surgery as surgery

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = 1
    ws["B1"] = "=A1/0"
    ws["C1"] = "=B1*2"
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    # bake cached error values in by surgery (openpyxl saves no cache)
    import zipfile

    with zipfile.ZipFile(raw) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode()
    fixed, changed = surgery.refresh_cached(xml, {"B1": "#DIV/0!", "C1": "#DIV/0!"})
    assert sorted(changed) == ["B1", "C1"]
    p = tmp_path / "err.xlsx"
    surgery.apply(raw, p, {"xl/worksheets/sheet1.xml": fixed})

    run_id = _make_run(p)
    xlsx.sheet_unit(run_id, {"sheet": "M", "target": str(p)})
    errs = {f["ref"]: f for f in store.findings(run_id) if f["code"] == "XL-ERRCELL"}
    assert set(errs) == {"M!B1", "M!C1"}
    assert "root: M!B1" in errs["M!C1"]["detail"]
    assert "M!B1" in errs["M!C1"]["fix"]


def test_trace_precedents_walks_the_chain(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["B1"], ws["B2"], ws["B3"] = 8, 6.5, 7
    ws["B4"] = "=SUM(B1:B3)"
    ws["B5"] = 42.5
    ws["B6"] = "=B4*B5"
    p = tmp_path / "m.xlsx"
    wb.save(p)
    g = xlsx._graph(p)
    lines = g.trace_precedents("Model!B6")
    assert lines and "Model!B4" in lines[0] and "Model!B5" in lines[0]
    assert any("Model!B1" in ln for ln in lines[1:])  # second hop reaches inputs


# ----------------------------------------------------- new check classes


def test_merged_cells_inside_summed_range(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    for i in range(1, 5):
        ws.cell(row=i, column=1, value=i * 10)
    ws.merge_cells("A2:A3")
    ws["B1"] = "=SUM(A1:A4)"
    p = tmp_path / "merged.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.sheet_unit(run_id, {"sheet": "M", "target": str(p)})
    hits = [f for f in store.findings(run_id) if f["code"] == "XL-MERGED-RANGE"]
    assert len(hits) == 1 and hits[0]["ref"] == "M!B1"
    assert "A2:A3" in hits[0]["title"]


def test_anchor_drift_flags_the_lone_unanchored_fill(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    for r in range(1, 5):
        ws.cell(row=r, column=1, value=r)
    ws["B5"] = 2
    ws["D1"] = "=A1*$B$5"
    ws["D2"] = "=A2*$B$5"
    ws["D3"] = "=A3*$B$5"
    ws["D4"] = "=A4*B8"  # anchor dropped — points somewhere else per row
    p = tmp_path / "anchor.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.sheet_unit(run_id, {"sheet": "M", "target": str(p)})
    hits = [f for f in store.findings(run_id) if f["code"] == "XL-ANCHOR-DRIFT"]
    assert [h["ref"] for h in hits] == ["M!D4"]


def test_date_text_mixing_in_a_column(tmp_path):
    import datetime

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = datetime.date(2026, 1, 5)
    ws["A2"] = datetime.date(2026, 2, 5)
    ws["A3"] = "3/5/2026"  # text pretending to be a date
    p = tmp_path / "dates.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.sheet_unit(run_id, {"sheet": "M", "target": str(p)})
    hits = [f for f in store.findings(run_id) if f["code"] == "XL-DATE-MIX"]
    assert len(hits) == 1 and "A" in hits[0]["title"]


def test_orphaned_and_shadowing_defined_names(tmp_path):
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = 5
    ws["B1"] = "=UsedName*2"
    wb.defined_names["UsedName"] = DefinedName("UsedName", attr_text="M!$A$1")
    wb.defined_names["DeadName"] = DefinedName("DeadName", attr_text="M!$A$9")
    ws.defined_names["UsedName"] = DefinedName("UsedName", attr_text="M!$A$2")  # shadows
    p = tmp_path / "names.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.workbook_unit(run_id, {"target": str(p)})
    fs = store.findings(run_id)
    orphan = next(f for f in fs if f["code"] == "XL-NAME-ORPHAN")
    assert "DeadName" in orphan["detail"] and "UsedName (sheet" not in orphan["detail"].replace(
        "UsedName (sheet: M)", ""
    )
    shadow = [f for f in fs if f["code"] == "XL-NAME-SHADOW"]
    assert len(shadow) == 1 and "UsedName" in shadow[0]["title"]


def test_unread_inputs_and_orphan_formulas(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = 10
    ws["A2"] = 20  # read by the SUM
    ws["B1"] = "=SUM(A1:A2)"
    ws["D9"] = 777  # nothing reads this
    p = tmp_path / "orph.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    assert digest["unread_inputs"] == 1
    unread = next(f for f in store.findings(run_id) if f["code"] == "XL-UNREAD-INPUT")
    assert "M!D9" in unread["detail"]
    assert digest["orphan_formulas"] == 1  # B1 is terminal — observation, not defect
    assert not [
        f
        for f in store.findings(run_id)
        if f["ref"] == "M!B1" and f["code"].startswith("XL-ORPHAN")
    ]


def test_whole_column_reference_marks_inputs_as_read(tmp_path):
    """5,000 numeric cells under =SUM(A:A): a capped area must still count as
    READ — the interval index, not expansion, answers membership."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    for r in range(1, 5001):
        ws.cell(row=r, column=1, value=r * 1.5)
    ws["C1"] = "=SUM(A:A)"
    p = tmp_path / "bigcol.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    assert digest["unread_inputs"] == 0
    assert not [f for f in store.findings(run_id) if f["code"] == "XL-UNREAD-INPUT"]
    g = xlsx._graph(p)
    assert g.is_referenced("M", 1, 1) and g.is_referenced("M", 1, 5000)
    assert not g.is_referenced("M", 2, 1)


def test_wide_capped_area_uses_row_buckets(tmp_path, monkeypatch):
    """A whole-row reference above the cap: membership stays exact and the
    containment pass walks row buckets, not 16,384 columns."""
    from openpyxl import Workbook

    monkeypatch.setattr(depgraph, "EXPAND_CAP", 4)  # force the capped path
    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    for c in range(1, 11):
        ws.cell(row=3, column=c, value=c)
    ws["A5"] = "=SUM(3:3)"
    ws["J3"] = "=A1+1"  # formula INSIDE the wide area → must become an edge
    p = tmp_path / "widerow.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    g = xlsx._graph(p)
    assert g.is_referenced("M", 5, 3) and not g.is_referenced("M", 5, 4)
    assert "M!J3" in g.precedents["M!A5"]  # containment via row buckets
    assert digest["unread_inputs"] == 0


def test_many_capped_rects_and_inputs_stay_fast(tmp_path, monkeypatch):
    """Membership against capped areas is batched per row — many distinct
    capped ranges times many numeric inputs must not blow up."""
    import time

    from openpyxl import Workbook

    monkeypatch.setattr(depgraph, "EXPAND_CAP", 4)
    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    for r in range(1, 401):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=r * c)
    # 100 distinct capped column ranges, plus full-column coverage of A-E
    for i in range(100):
        col = "ABCDE"[i % 5]
        ws.cell(row=401 + i, column=7, value=f"=SUM({col}{1 + i}:{col}400)")
    for j, col in enumerate("ABCDE"):
        ws.cell(row=520 + j, column=7, value=f"=SUM({col}1:{col}400)")
    p = tmp_path / "many.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    t0 = time.monotonic()
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    elapsed = time.monotonic() - t0
    assert elapsed < 10, f"graph audit took {elapsed:.1f}s"
    # rows 100+ of every column are read by the deepest-starting ranges;
    # row 1 of col A is read by =SUM(A1:A400)
    assert digest["unread_inputs"] == 0


def test_3d_sheet_span_resolves_per_sheet_never_false_unread(tmp_path):
    """=SUM(S1:S3!A1) reads A1 on EVERY sheet in the span — those inputs must
    never be reported unread, and the edges must be real per-sheet edges."""
    from openpyxl import Workbook

    # analyze-level: exact expansion with sheet order, refusal without it
    facts = fx.analyze("=SUM(S1:S3!A1)", "Sum", sheets=["S1", "S2", "S3", "Sum"])
    assert facts["areas"] == [("S1", 1, 1, 1, 1), ("S2", 1, 1, 1, 1), ("S3", 1, 1, 1, 1)]
    blind = fx.analyze("=SUM(S1:S3!A1)", "Sum")
    assert not blind["areas"] and blind["unresolved_names"] == ["S1:S3!A1"]
    missing = fx.analyze("=SUM(S1:S9!A1)", "Sum", sheets=["S1", "S2", "S3"])
    assert not missing["areas"] and missing["unresolved_names"] == ["S1:S9!A1"]
    # quoted span with spaces
    q = fx.analyze("=SUM('Q 1:Q 2'!B2)", "Sum", sheets=["Q 1", "Q 2"])
    assert q["areas"] == [("Q 1", 2, 2, 2, 2), ("Q 2", 2, 2, 2, 2)]

    # workbook-level: no false unread inputs, real per-sheet edges
    wb = Workbook()
    wb.active.title = "S1"
    wb["S1"]["A1"] = 10
    for n in ("S2", "S3"):
        wb.create_sheet(n)["A1"] = 20
    wb.create_sheet("Sum")["A1"] = "=SUM(S1:S3!A1)"
    p = tmp_path / "threed.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    digest = xlsx.workbook_unit(run_id, {"target": str(p)})
    assert digest["unread_inputs"] == 0
    g = xlsx._graph(p)
    assert {"S1!A1", "S2!A1", "S3!A1"} <= g.precedents["Sum!A1"]
    assert not g.unresolved_names
    assert not [f for f in store.findings(run_id) if f["code"] == "XL-GRAPH-PARTIAL"]


def test_constant_defined_name_is_not_an_orphan(tmp_path):
    """A constant name (=0.21) has no cell destination but IS a used name —
    the audit must not call it orphaned, and must disclose partial graph."""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = 100
    ws["B1"] = "=A1*TaxRate"
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text="0.21")
    p = tmp_path / "const.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.workbook_unit(run_id, {"target": str(p)})
    fs = store.findings(run_id)
    assert not [f for f in fs if f["code"] == "XL-NAME-ORPHAN"]
    partial = [f for f in fs if f["code"] == "XL-GRAPH-PARTIAL"]
    assert len(partial) == 1 and "TaxRate" in partial[0]["detail"]


def test_graph_partial_when_indirect_present(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws["A1"] = '=INDIRECT("B"&ROW())'
    p = tmp_path / "ind.xlsx"
    wb.save(p)
    run_id = _make_run(p)
    xlsx.workbook_unit(run_id, {"target": str(p)})
    hits = [f for f in store.findings(run_id) if f["code"] == "XL-GRAPH-PARTIAL"]
    assert len(hits) == 1 and "M!A1" in hits[0]["detail"]
