"""The prove pipeline: detect → repair by surgery → recalculate → gate → emit.

Every test here answers one question: can this system hand back a workbook
and honestly call it proven?
"""
import json
import re
import sys
import zipfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from runner.jobs import xlsx, xlsx_engine as engine, xlsx_surgery as surgery  # noqa: E402
from runner import store  # noqa: E402

DEMO = Path(__file__).resolve().parents[1] / "targets" / "demo_defects.xlsx"

pytestmark = pytest.mark.skipif(not engine.available(),
                                reason="formulas engine not installed")


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    """Each test gets its own runs dir + checkpoint DB and mock mode."""
    from runner.config import CFG
    monkeypatch.setattr(CFG, "runs_dir", str(tmp_path / "runs"))
    monkeypatch.setattr(CFG, "db", str(tmp_path / "runs" / "runner.db"))
    monkeypatch.setattr(CFG, "mock", True)
    xlsx._drop_cache()
    yield
    xlsx._drop_cache()


def _make_run(job="xlsx", target="x"):
    store.init()
    return store.start_run(job, str(target), "test", {})


def _make_wb(path, credits=False):
    """A small clean model — but openpyxl writes NO cached values, so the
    file ships 'never calculated'. The prove pass must repair that."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "hours"; ws["B1"] = 8
    ws["A2"] = "hours"; ws["B2"] = 6.5
    ws["A3"] = "hours"; ws["B3"] = 7
    ws["A4"] = "total"; ws["B4"] = "=SUM(B1:B3)"
    ws["A5"] = "rate";  ws["B5"] = 42.5
    ws["A6"] = "pay";   ws["B6"] = "=B4*B5"
    wb.save(path)
    return path


# ---------------------------------------------------------------- engine

def test_engine_recalculates_and_flags_missing_cache(tmp_path):
    p = _make_wb(tmp_path / "clean.xlsx")
    recalc = engine.recalculate(p)
    assert recalc["available"] and recalc["checked"] >= 2
    cmp0 = engine.compare(p, recalc)
    # No cached values at all: every formula is a mismatch, none uncovered
    assert not cmp0["uncovered"]
    refs = {m["ref"] for m in cmp0["mismatches"]}
    assert refs == {"Model!B4", "Model!B6"}
    b4 = next(m for m in cmp0["mismatches"] if m["cell"] == "B4")
    assert b4["computed"] == 21.5 and b4["cached"] is None


def test_engine_catches_stale_cache_in_demo_workbook():
    recalc = engine.recalculate(DEMO)
    assert recalc["available"]
    cmp0 = engine.compare(DEMO, recalc)
    # The demo ships deliberate defects; the recalculation must disagree
    # with at least one saved value — a stale total cannot certify itself.
    assert cmp0["mismatches"], "demo defects went undetected by recalculation"


def test_engine_unavailable_is_honest(monkeypatch, tmp_path):
    p = _make_wb(tmp_path / "c.xlsx")
    monkeypatch.setattr(engine, "available", lambda: False)
    out = engine.recalculate(p)
    assert out["available"] is False and "not installed" in out["error"]


# --------------------------------------------------------------- surgery

BAD_SHEET = (
    '<?xml version="1.0"?>'
    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    "<pageMargins left=\"0.7\"/>"
    "<mergeCells count=\"1\"><mergeCell ref=\"A1:B1\"/></mergeCells>"
    "<sheetData><row r=\"1\"><c r=\"A1\"><v>1</v></c></row></sheetData>"
    "<dataValidations count=\"1\"><dataValidation sqref=\"A2\"/></dataValidations>"
    "</worksheet>")


def test_reorder_sheet_xml_produces_canonical_order():
    assert surgery.sheet_order_violations(BAD_SHEET)
    fixed = surgery.reorder_sheet_xml(BAD_SHEET)
    assert not surgery.sheet_order_violations(fixed)
    seq = [t for t in re.findall(r"<(\w+)[ >/]", fixed)
           if t in surgery.SHEET_ORDER]
    assert seq == ["sheetData", "mergeCells", "dataValidations", "pageMargins"]
    # content survives intact
    assert "<mergeCell ref=\"A1:B1\"/>" in fixed and "<v>1</v>" in fixed


def test_reorder_is_identity_on_clean_xml():
    clean = surgery.reorder_sheet_xml(BAD_SHEET)
    assert surgery.reorder_sheet_xml(clean) == clean


def test_refresh_cached_inserts_and_replaces_values():
    xml = ('<worksheet><sheetData><row r="1">'
           '<c r="A1"><v>5</v></c>'                         # input: untouched
           '<c r="B1"><f>SUM(A1)</f></c>'                   # never calculated
           '<c r="C1" t="e"><f>A1/0</f><v>#DIV/0!</v></c>'  # stale error
           '</row></sheetData></worksheet>')
    out, changed = surgery.refresh_cached(xml, {"A1": 99, "B1": 5.0, "C1": 5.0})
    assert sorted(changed) == ["B1", "C1"]
    assert "<c r=\"A1\"><v>5</v></c>" in out          # inputs never edited
    assert "<f>SUM(A1)</f><v>5</v>" in out
    assert 't="e"' not in out and "#DIV/0!" not in out


def test_apply_and_parts_diff_prove_containment(tmp_path):
    p = _make_wb(tmp_path / "a.xlsx")
    out = tmp_path / "b.xlsx"
    with zipfile.ZipFile(p) as z:
        part = "xl/worksheets/sheet1.xml"
        xml = z.read(part).decode()
    touched = surgery.apply(p, out, {part: xml.replace("42.5", "43.5")})
    assert touched == [part]
    diff = surgery.parts_diff(p, out)
    assert diff["changed"] == [part] and not diff["added"] and not diff["removed"]


def test_matches_is_strict():
    assert engine._matches(None, None)
    assert not engine._matches(0.0, None)          # missing cache is a mismatch
    assert not engine._matches(True, 1)            # bool never equals a number
    assert not engine._matches(False, None)
    assert engine._matches(True, True)
    assert engine._matches(21.5, 21.5 + 1e-12)


def test_reorder_refuses_unknown_top_level_content():
    with_ext = BAD_SHEET.replace(
        "</worksheet>", "<extLst><ext uri=\"x\"><foo/></ext></extLst></worksheet>")
    fixed = surgery.reorder_sheet_xml(with_ext)      # extLst is known: fixable
    assert not surgery.sheet_order_violations(fixed)
    unknown = BAD_SHEET.replace(
        "</worksheet>", "<mysteryTag><a/></mysteryTag></worksheet>")
    assert surgery.reorder_sheet_xml(unknown) == unknown         # refused
    commented = BAD_SHEET.replace("</worksheet>", "<!-- note --></worksheet>")
    assert surgery.reorder_sheet_xml(commented) == commented     # refused


def test_refresh_skips_shared_and_inline_string_cells():
    xml = ('<worksheet><sheetData><row r="1">'
           '<c r="A1" t="s"><f>X()</f><v>3</v></c>'
           '<c r="B1" t="inlineStr"><f>Y()</f><is><t>hi</t></is></c>'
           '</row></sheetData></worksheet>')
    out, changed = surgery.refresh_cached(xml, {"A1": 9.0, "B1": 9.0})
    assert out == xml and changed == []


# ------------------------------------------------------------ prove unit

def test_prove_repairs_never_calculated_workbook(tmp_path):
    p = _make_wb(tmp_path / "clean.xlsx")
    run_id = _make_run(target=p)
    digest = xlsx.prove_unit(run_id, {"target": str(p)})
    assert digest["verdict"] == "PROVEN", digest
    assert all(digest["gates"].values()), digest["gates"]
    out = Path(digest["output"])
    assert out.exists() and out.name == "PROVEN_clean.xlsx"
    # the input file was never modified
    assert surgery.parts_diff(p, p)["changed"] == []
    # cached values now present and correct
    from openpyxl import load_workbook
    wv = load_workbook(out, data_only=True)
    assert wv["Model"]["B4"].value == 21.5
    assert wv["Model"]["B6"].value == pytest.approx(21.5 * 42.5)
    wv.close()
    # surgery touched only the sheet part
    diff = surgery.parts_diff(p, out)
    assert diff["changed"] == ["xl/worksheets/sheet1.xml"]
    # and the run built a test manifest
    manifest = json.loads(Path(digest["tests"]["path"]).read_text())
    assert manifest["formula_cells"] == 2
    assert engine.run_manifest(str(out), manifest)["status"] == "PASS"


def test_prove_fails_and_ships_nothing_on_true_errors(tmp_path):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws["A1"] = 1; ws["B1"] = "=A1/0"       # genuinely errors on recalculation
    p = tmp_path / "bad.xlsx"; wb.save(p)
    run_id = _make_run(target=p)
    digest = xlsx.prove_unit(run_id, {"target": str(p)})
    assert digest["verdict"] == "FAILED PROOF"
    assert digest["output"] is None
    from runner.config import CFG
    assert not list(Path(CFG.runs_dir).glob("**/PROVEN_*.xlsx"))
    codes = {f["code"] for f in store.findings(run_id)}
    assert "XL-PROOF-FAILED" in codes


def test_prove_unverified_without_engine(tmp_path, monkeypatch):
    p = _make_wb(tmp_path / "c.xlsx")
    run_id = _make_run(target=p)
    monkeypatch.setattr(engine, "available", lambda: False)
    digest = xlsx.prove_unit(run_id, {"target": str(p)})
    assert digest["verdict"] == "UNVERIFIED" and digest["output"] is None
    codes = {f["code"] for f in store.findings(run_id)}
    assert "XL-UNVERIFIED" in codes


def test_manifest_catches_a_later_bad_edit(tmp_path):
    p = _make_wb(tmp_path / "clean.xlsx")
    run_id = _make_run(target=p)
    digest = xlsx.prove_unit(run_id, {"target": str(p)})
    out = Path(digest["output"])
    manifest = json.loads(Path(digest["tests"]["path"]).read_text())
    # someone edits an input by surgery — totals must move, manifest must fail
    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode()
    broken = tmp_path / "edited.xlsx"
    surgery.apply(out, broken, {"xl/worksheets/sheet1.xml": xml.replace("<v>8</v>", "<v>9</v>")})
    res = engine.run_manifest(str(broken), manifest)
    assert res["status"] == "FAIL"
    assert any(f["cell"] in ("B4", "B6") for f in res["failed"])


def test_manifest_reruns_structural_rules(tmp_path):
    p = _make_wb(tmp_path / "clean.xlsx")
    run_id = _make_run(target=p)
    digest = xlsx.prove_unit(run_id, {"target": str(p)})
    out = Path(digest["output"])
    manifest = json.loads(Path(digest["tests"]["path"]).read_text())
    # break the OOXML order without touching any value: formulas still pass,
    # the STRUCTURAL rules must catch it
    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode()
    m = re.search(r"<pageMargins[^>]*/>", xml)
    assert m
    # move pageMargins to just inside <worksheet ...> — before sheetData
    ws_open = re.search(r"<worksheet[^>]*>", xml)
    head_end = ws_open.end()
    reordered = xml[:head_end] + m.group(0) + xml[head_end:].replace(m.group(0), "")
    broken = tmp_path / "misordered.xlsx"
    surgery.apply(out, broken, {"xl/worksheets/sheet1.xml": reordered})
    res = engine.run_manifest(str(broken), manifest)
    assert res["status"] == "FAIL" and res["structural"]


def test_no_proven_file_left_when_gates_crash(tmp_path, monkeypatch):
    p = _make_wb(tmp_path / "clean.xlsx")
    run_id = _make_run(target=p)
    monkeypatch.setattr(surgery, "parts_diff",
                        lambda a, b: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(RuntimeError):
        xlsx.prove_unit(run_id, {"target": str(p)})
    from runner.config import CFG
    leftovers = list(Path(CFG.runs_dir).glob("**/PROVEN_*")) + \
                list(Path(CFG.runs_dir).glob("**/.candidate_*"))
    assert leftovers == [], leftovers


# ------------------------------------------------------------- full run

def test_full_run_on_demo_defects_reports_not_proven(tmp_path):
    """The shipped demo has five deliberate defects; a full pipeline run must
    finish, find them, and refuse to certify what it cannot repair — or
    repair the stale caches and certify. Either way the verdict is explicit."""
    from runner import harness
    run_id = _make_run(target=DEMO)
    plan = xlsx.plan(str(DEMO), tmp_path)
    store.add_units(run_id, plan["units"])
    res = harness.execute(run_id, xlsx, xlsx.unit_worker, xlsx.final_pass)
    assert res["status"] == "done"
    summary = res["totals"]["summary"]
    assert summary.get("verdict") in ("PROVEN", "PROVEN WITH WARNINGS", "FAILED PROOF")
    findings = store.findings(run_id)
    assert findings, "the demo's deliberate defects produced no findings"
    if summary["verdict"] == "FAILED PROOF":
        assert summary.get("output") is None
    else:
        assert Path(summary["output"]).exists()
